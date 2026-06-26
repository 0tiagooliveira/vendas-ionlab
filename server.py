from __future__ import annotations

import csv
import copy
import json
import math
import re
import sys
import threading
import time
import traceback
import unicodedata
import hashlib
import zipfile
import xml.etree.ElementTree as ET
from collections import Counter, defaultdict
from datetime import date, datetime, timedelta
from email import policy
from email.parser import BytesParser
from http import HTTPStatus
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from io import BytesIO, TextIOWrapper
from pathlib import Path
from urllib.parse import parse_qs, unquote, urlparse

try:
    import pandas as pd
except ImportError as exc:
    raise SystemExit(
        "Este CRM precisa de pandas e openpyxl para importar planilhas .xlsx.\n"
        "Instale com: python -m pip install pandas openpyxl"
    ) from exc

try:
    from PIL import Image
except ImportError:
    Image = None


BASE_DIR = Path(__file__).resolve().parent
STATIC_DIR = BASE_DIR / "static"
DATA_DIR = BASE_DIR / "data"
SERVER_LOG_FILE = DATA_DIR / "server_logs" / "server.log"
JSON_CACHE = {}
PAYLOAD_CACHE = {}
SHARED_PAYLOAD_CACHE = {}
RUNTIME_AUTH_SESSIONS = {}
SAVE_JSON_LOCK = threading.Lock()

COMPANIES = {
    "ionlab": "Ionlab",
    "ciorbrasil": "CiorBrasil",
    "even": "Even",
    "onix": "Onix",
    "vitralab": "Vitralab",
    "ambarlab": "Ambarlab",
    "nativalab": "Nativalab",
}

DASHBOARD_COMPANY_GROUPS = {
    "ionlab_cior": {
        "name": "Ionlab+Cior",
        "companies": ["ionlab", "ciorbrasil"],
    },
}
MERCADO_LIVRE_COMPANIES = ("onix", "vitralab", "nativalab")

COMPANY_NAME_ALIASES = {
    "ionlab": ["Ionlab", "Ionlab Equip Lab", "Ionlab Equipamentos"],
    "ciorbrasil": ["CiorBrasil", "Cior Brasil"],
    "even": ["Even"],
    "onix": ["Onix"],
    "vitralab": ["Vitralab", "Vitralab Equipamentos"],
    "ambarlab": ["Ambarlab"],
    "nativalab": ["Nativalab", "Nativa Lab", "Nativa Lab Produtos"],
}

BRAZIL_UFS = [
    "AC", "AL", "AP", "AM", "BA", "CE", "DF", "ES", "GO",
    "MA", "MT", "MS", "MG", "PA", "PB", "PR", "PE", "PI",
    "RJ", "RN", "RS", "RO", "RR", "SC", "SP", "SE", "TO",
]

START_YEAR = 2021
CURRENT_YEAR = max(2026, date.today().year)

GROUP_COMPANY_NAMES = sorted({
    alias
    for aliases in COMPANY_NAME_ALIASES.values()
    for alias in aliases
})

COMPANY_SALE_EXCLUSIONS = {
    company_id: [
        alias
        for other_company_id, aliases in COMPANY_NAME_ALIASES.items()
        if other_company_id != company_id
        for alias in aliases
    ]
    for company_id in COMPANIES
}

CONTENT_TYPES = {
    ".html": "text/html; charset=utf-8",
    ".css": "text/css; charset=utf-8",
    ".js": "application/javascript; charset=utf-8",
    ".json": "application/json; charset=utf-8",
    ".svg": "image/svg+xml",
}

SALES_TABLE_COLUMNS = [
    ("ID_NF", "ID NF"),
    ("NF_NUM", "Nota"),
    ("NF_SER", "Serie"),
    ("NF_EMI", "Emissao"),
    ("CL_NOM", "Cliente"),
    ("CL_UF", "UF"),
    ("VN_NOM", "Vendedor"),
    ("PR_COD", "Cod. produto"),
    ("PR_DES", "Produto"),
    ("AGRP", "AGRP"),
    ("Origem", "Origem"),
    ("PR_QTD", "Qtd."),
    ("PR_SBT", "Valor"),
    ("LUCRO_REAIS", "Lucro"),
]

CLIENT_TABLE_COLUMNS = [
    ("ID", "ID"),
    ("CGC", "Documento"),
    ("NOM", "Nome"),
    ("FAN", "Fantasia"),
    ("MAI", "Email"),
    ("TEL", "Telefone"),
    ("CID", "Cidade"),
    ("UF", "UF"),
    ("CAT", "Categoria"),
    ("CC_DES", "Tipo"),
]

STOCK_TABLE_COLUMNS = [
    ("Referencia", "Referencia"),
    ("Descricao", "Descricao"),
    ("AGRP", "AGRP"),
    ("Origem", "Origem"),
    ("Unidade", "Unidade"),
    ("Quantidade", "Quantidade"),
    ("V.Unitario", "V.Unitario"),
    ("Total", "Total"),
    ("Ultima Saida", "Ultima Saida"),
    ("Dias sem movimento", "Dias sem movimento"),
    ("C.Fiscal", "C.Fiscal"),
    ("Id.Estoque", "Id.Estoque"),
    ("Local", "Local"),
    ("End. Fisico", "End. Fisico"),
]

IN_TRANSIT_TABLE_COLUMNS = [
    ("Referencia", "Referencia"),
    ("Descricao", "Descricao"),
    ("QTY", "QTY"),
    ("Processo", "Processo"),
    ("Previsao", "Previsao"),
]

COSTING_TABLE_COLUMNS = [
    ("Referencia", "Referencia"),
    ("Descricao item", "Descricao item"),
    ("AGRP", "AGRP"),
    ("Origem", "Origem"),
    ("Preco Dollar", "Preco Dollar"),
    ("Custo BRL Indice", "Custo BRL Indice"),
    ("Custo BRL Ultima Importacao", "Custo BRL Ultima Importacao"),
]

PRICE_TABLE_COLUMNS = [
    ("ID_TAB", "ID_TAB"),
    ("ID_EST", "ID_EST"),
    ("LC_COD", "LC_COD"),
    ("PR_COD", "PR_COD"),
    ("PR_DES", "PR_DES"),
    ("AGRP", "AGRP"),
    ("Origem", "Origem"),
    ("VAL_VEND", "VAL_VEND"),
    ("PER_DESC", "PER_DESC"),
    ("PER_COM", "PER_COM"),
    ("Aliquota IPI", "Aliquota IPI"),
    ("Custo BRL Indice", "Custo BRL Indice"),
]

PRODUCT_TABLE_COLUMNS = [
    ("ID", "ID"),
    ("Referencia", "Referencia"),
    ("Descricao resumida", "Descricao resumida"),
    ("Descricao", "Descricao"),
    ("Dados tecnicos", "Dados tecnicos"),
    ("Descricao completa", "Descricao completa"),
    ("Marca", "Marca"),
    ("So se usuario especificar", "So se usuario especificar"),
    ("Palavra chave", "Palavra chave"),
    ("Palavras positivas", "Palavras positivas"),
    ("Palavras negativas", "Palavras negativas"),
    ("Regras de variaveis", "Regras de variaveis"),
    ("Score minimo", "Score minimo"),
]

PRODUCT_CASE_FIELDS = {
    "Descricao resumida",
    "Descricao",
    "Dados tecnicos",
    "Descricao completa",
    "Marca",
    "So se usuario especificar",
    "Palavra chave",
    "Palavras positivas",
    "Palavras negativas",
    "Regras de variaveis",
}

PRODUCT_ACRONYM_REPLACEMENTS = {
    "ph": "pH",
    "usb": "USB",
    "led": "LED",
    "lcd": "LCD",
    "uv": "UV",
    "ir": "IR",
    "rpm": "RPM",
    "ncm": "NCM",
    "sku": "SKU",
    "c.a": "C.A",
    "ca": "CA",
    "p.a": "P.A",
    "pa": "PA",
    "ptfe": "PTFE",
    "pvc": "PVC",
    "br": "BR",
    "co2": "CO2",
    "o2": "O2",
    "h2o": "H2O",
    "bod": "BOD",
    "dqo": "DQO",
    "dbo": "DBO",
}

VENDOR_TABLE_COLUMNS = [
    ("foto_vendedor", "Foto"),
    ("status", "Status"),
    ("tipo_usuario", "Tipo usuario"),
    ("nome_completo", "Nome completo"),
    ("usuario_vinculado_nome", "Usuario vinculado"),
    ("login_acesso", "Login"),
    ("acesso_configurado", "Acesso"),
    ("pagina_vendedor", "Pagina"),
    ("telefone", "Telefone"),
    ("whatsapp", "WhatsApp"),
    ("email", "Email"),
    ("empresas_acesso_nomes", "Empresas de acesso"),
    ("origem", "Origem"),
]

BLOCKED_CLIENT_COLUMNS = [
    ("codigo_cliente", "Codigo"),
    ("nome_cliente", "Cliente"),
    ("motivo", "Motivo"),
    ("empresa_nome", "Empresa"),
    ("bloqueado_em", "Bloqueado em"),
]

PROSPECT_COLUMNS = [
    ("ID", "Codigo"),
    ("NOM", "Cliente"),
    ("UF", "UF"),
    ("CID", "Cidade"),
    ("status", "Status"),
    ("compras_periodo", "Compras"),
    ("faturamento_liquido", "Faturamento liquido"),
    ("ultima_compra", "Ultima compra"),
]


def company_dir(company_id: str) -> Path:
    return DATA_DIR / company_id


def sales_file(company_id: str) -> Path:
    return company_dir(company_id) / "vendas.json"


def import_log_file(company_id: str) -> Path:
    return company_dir(company_id) / "importacoes.json"


def clients_file(company_id: str) -> Path:
    return company_dir(company_id) / "clientes.json"


def stock_file(company_id: str) -> Path:
    return company_dir(company_id) / "estoque.json"


def in_transit_file(company_id: str) -> Path:
    return company_dir(company_id) / "em_transito.json"


def slow_items_file(company_id: str) -> Path:
    return company_dir(company_id) / "itens_sem_giro.json"


def costing_file(company_id: str) -> Path:
    return company_dir(company_id) / "custeio_importados.json"


def costing_fabricated_file(company_id: str) -> Path:
    return company_dir(company_id) / "custeio_fabricados.json"


def price_table_file(company_id: str) -> Path:
    return company_dir(company_id) / "tabela_precos.json"


def products_file(company_id: str) -> Path:
    return company_dir(company_id) / "cadastro_produtos.json"


def quotes_file(company_id: str) -> Path:
    return company_dir(company_id) / "orcamentos.json"


def orders_file(company_id: str) -> Path:
    return company_dir(company_id) / "pedidos.json"




def mercado_livre_ads_file(company_id: str) -> Path:
    return company_dir(company_id) / "mercado_livre_anuncios.json"


def mercado_livre_dir(company_id: str) -> Path:
    return company_dir(company_id) / "mercado_livre"


def mercado_livre_imports_file(company_id: str) -> Path:
    return mercado_livre_dir(company_id) / "importacoes.json"


def mercado_livre_import_file(company_id: str, import_id: str) -> Path:
    return mercado_livre_dir(company_id) / f"{import_id}.json"


def mercado_livre_sales_file(company_id: str) -> Path:
    return mercado_livre_dir(company_id) / "vendas.json"


def costing_config_file(company_id: str) -> Path:
    return company_dir(company_id) / "custeio_config.json"


def vendors_file(company_id: str) -> Path:
    return company_dir(company_id) / "vendedores.json"


def vendor_goals_file(company_id: str) -> Path:
    return company_dir(company_id) / "metas_objetivos_vendedores.json"


def vendor_monthly_items_file(company_id: str) -> Path:
    return company_dir(company_id) / "itens_equipamentos_mes.json"


def vendor_day_by_day_file(company_id: str) -> Path:
    return company_dir(company_id) / "day_by_day_vendedores.json"


def region_assignments_file(company_id: str) -> Path:
    return company_dir(company_id) / "regioes_vendedores.json"


def blocked_clients_file(company_id: str) -> Path:
    return company_dir(company_id) / "clientes_bloqueados.json"


def block_reasons_file() -> Path:
    return DATA_DIR / "motivos_bloqueio.json"


def users_file() -> Path:
    return DATA_DIR / "usuarios.json"


def auth_sessions_file() -> Path:
    return DATA_DIR / "sessoes_usuarios.json"


def client_import_log_file(company_id: str) -> Path:
    return company_dir(company_id) / "importacoes_clientes.json"


def stock_import_log_file(company_id: str) -> Path:
    return company_dir(company_id) / "importacoes_estoque.json"


def costing_import_log_file(company_id: str) -> Path:
    return company_dir(company_id) / "importacoes_custeio.json"


BAD_CHAR = "\ufffd"
MOJIBAKE_MARKERS = (
    "\u00c3\u00a7", "\u00c3\u00a3", "\u00c3\u00a1", "\u00c3\u00a9", "\u00c3\u00aa",
    "\u00c3\u00b4", "\u00c3\u00b3", "\u00c3\u00ad", "\u00c3\u00ba", "\u00c3\u0087",
    "\u00c3\u00b5", "\u00c3\u00c1", "\u00c3\u00a2", "\u00c3\u00a0", "\u00c3\u0080",
    "\u00c2\u00b0", "\u00c2\u00b7", "\u00c3\u201a", "\u00c3\u0192", "\u00c3",
    "\u00e2\u20ac", "\u00e2\u2020", "\u00e2\u0152", BAD_CHAR,
)


def mojibake_score(text: str) -> int:
    return (
        sum(text.count(marker) * 3 for marker in MOJIBAKE_MARKERS if marker != BAD_CHAR)
        + text.count(BAD_CHAR) * 5
    )


def clean_text_encoding(value: str) -> str:
    text = str(value)
    if not any(marker in text for marker in MOJIBAKE_MARKERS):
        return text

    best = text
    best_score = mojibake_score(text)
    for _ in range(3):
        improved = False
        for encoding in ("latin1", "cp1252"):
            try:
                candidate = best.encode(encoding).decode("utf-8")
            except (UnicodeEncodeError, UnicodeDecodeError):
                continue
            score = mojibake_score(candidate)
            if score < best_score:
                best = candidate
                best_score = score
                improved = True
        if not improved:
            break

    replacements = {
        f"Pe{BAD_CHAR}a": "Pe\u00e7a",
        f"pe{BAD_CHAR}a": "pe\u00e7a",
        f"Reposi{BAD_CHAR}{BAD_CHAR}o": "Reposi\u00e7\u00e3o",
        f"reposi{BAD_CHAR}{BAD_CHAR}o": "reposi\u00e7\u00e3o",
        f"Medi{BAD_CHAR}{BAD_CHAR}o": "Medi\u00e7\u00e3o",
        f"medi{BAD_CHAR}{BAD_CHAR}o": "medi\u00e7\u00e3o",
        f"Descri{BAD_CHAR}{BAD_CHAR}o": "Descri\u00e7\u00e3o",
        f"descri{BAD_CHAR}{BAD_CHAR}o": "descri\u00e7\u00e3o",
        f"Prote{BAD_CHAR}{BAD_CHAR}o": "Prote\u00e7\u00e3o",
        f"prote{BAD_CHAR}{BAD_CHAR}o": "prote\u00e7\u00e3o",
        f"Calibra{BAD_CHAR}{BAD_CHAR}o": "Calibra\u00e7\u00e3o",
        f"calibra{BAD_CHAR}{BAD_CHAR}o": "calibra\u00e7\u00e3o",
        f"Opera{BAD_CHAR}{BAD_CHAR}o": "Opera\u00e7\u00e3o",
        f"opera{BAD_CHAR}{BAD_CHAR}o": "opera\u00e7\u00e3o",
        f"Importa{BAD_CHAR}{BAD_CHAR}o": "Importa\u00e7\u00e3o",
        f"importa{BAD_CHAR}{BAD_CHAR}o": "importa\u00e7\u00e3o",
        f"Agita{BAD_CHAR}{BAD_CHAR}o": "Agita\u00e7\u00e3o",
        f"agita{BAD_CHAR}{BAD_CHAR}o": "agita\u00e7\u00e3o",
        f"A{BAD_CHAR}o": "A\u00e7o",
        f"a{BAD_CHAR}o": "a\u00e7o",
        f"N{BAD_CHAR}": "N\u00b0",
        f"n{BAD_CHAR}": "n\u00b0",
        "\u00c3\u0081": "\u00c1",
        "\u00c3\u0192O": "\u00c3O",
        "\u00c3\u0192o": "\u00c3o",
    }
    for broken, fixed in replacements.items():
        best = best.replace(broken, fixed)
    best = re.sub(f"(?<=\\d){BAD_CHAR}C", "\u00b0C", best)
    best = best.replace(f"{BAD_CHAR}{BAD_CHAR}O", "\u00c7\u00c3O")
    best = best.replace(f"{BAD_CHAR}{BAD_CHAR}o", "\u00e7\u00e3o")
    best = best.replace(f"{BAD_CHAR}{BAD_CHAR}A", "\u00c7A")
    best = best.replace(f"{BAD_CHAR}{BAD_CHAR}a", "\u00e7a")
    best = best.replace(f"AT{BAD_CHAR}", "AT\u00c9")
    best = best.replace(f"at{BAD_CHAR}", "at\u00e9")
    best = best.replace(f"FOR{BAD_CHAR}", "FOR\u00c7")
    best = best.replace(f"for{BAD_CHAR}", "for\u00e7")
    return best

def normalize_record(value):
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None
    if pd.isna(value):
        return None
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return value.isoformat()
    value = value.item() if hasattr(value, "item") else value
    if isinstance(value, str):
        return clean_text_encoding(value)
    return value


def product_sentence_case(value):
    value = normalize_record(value)
    if value in (None, ""):
        return ""

    text = re.sub(r"\s+", " ", str(value)).strip()
    if not text:
        return ""

    text = text.lower()
    text = re.sub(
        r"(^|[.!?;:]\s+|[•]\s*)([a-záàâãéêíóôõúç])",
        lambda match: match.group(1) + match.group(2).upper(),
        text,
    )
    text = re.sub(r"(?<=\d)\s*°\s*c\b", "°C", text, flags=re.IGNORECASE)
    text = re.sub(r"(?<=\d)\s*º\s*c\b", "°C", text, flags=re.IGNORECASE)

    for source, replacement in PRODUCT_ACRONYM_REPLACEMENTS.items():
        text = re.sub(rf"\b{re.escape(source)}\b", replacement, text, flags=re.IGNORECASE)

    return text


def normalize_product_record_case(record: dict) -> dict:
    normalized = dict(record)
    for field in PRODUCT_CASE_FIELDS:
        if field in normalized:
            normalized[field] = product_sentence_case(normalized.get(field))
    return normalized


def normalize_identifier(value) -> str:
    value = normalize_record(value)
    if value in (None, ""):
        return ""

    if isinstance(value, float) and value.is_integer():
        value = int(value)

    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return re.sub(r"\D", "", text) or text.upper()


def load_json(path: Path, default):
    if not path.exists():
        return default
    stat = path.stat()
    cache_key = str(path.resolve())
    signature = (stat.st_mtime_ns, stat.st_size)
    cached = JSON_CACHE.get(cache_key)
    if cached and cached["signature"] == signature:
        return copy.deepcopy(cached["payload"])
    payload = json.loads(path.read_text(encoding="utf-8-sig"))
    JSON_CACHE[cache_key] = {"signature": signature, "payload": payload}
    return copy.deepcopy(payload)


def save_json(path: Path, payload):
    path.parent.mkdir(parents=True, exist_ok=True)
    content = json.dumps(payload, ensure_ascii=False, indent=2)
    with SAVE_JSON_LOCK:
        last_error = None
        for attempt in range(20):
            temp_path = path.with_name(
                f".{path.name}.{threading.get_ident()}.{int(time.time() * 1000)}.{attempt}.tmp"
            )
            try:
                temp_path.write_text(content, encoding="utf-8")
                temp_path.replace(path)
                last_error = None
                break
            except PermissionError as exc:
                last_error = exc
                try:
                    if temp_path.exists():
                        temp_path.unlink()
                except Exception:
                    pass
                try:
                    path.write_text(content, encoding="utf-8")
                    last_error = None
                    break
                except PermissionError as direct_exc:
                    last_error = direct_exc
                time.sleep(0.5)
        if last_error:
            raise last_error
    try:
        stat = path.stat()
        JSON_CACHE[str(path.resolve())] = {
            "signature": (stat.st_mtime_ns, stat.st_size),
            "payload": copy.deepcopy(payload),
        }
    except Exception:
        JSON_CACHE.pop(str(path.resolve()), None)
    PAYLOAD_CACHE.clear()
    SHARED_PAYLOAD_CACHE.clear()


def file_signature(paths: list[Path]) -> tuple:
    signature = []
    for path in paths:
        if path.exists():
            stat = path.stat()
            signature.append((str(path.resolve()), stat.st_mtime_ns, stat.st_size))
        else:
            signature.append((str(path.resolve()), 0, 0))
    return tuple(signature)


def cached_payload(cache_key: tuple, dependencies: list[Path], builder):
    signature = file_signature(dependencies)
    cached = PAYLOAD_CACHE.get(cache_key)
    if cached and cached["signature"] == signature:
        return copy.deepcopy(cached["payload"])
    payload = builder()
    PAYLOAD_CACHE[cache_key] = {"signature": signature, "payload": copy.deepcopy(payload)}
    return payload


def shared_cached_payload(cache_key: tuple, dependencies: list[Path], builder):
    signature = file_signature(dependencies)
    cached = SHARED_PAYLOAD_CACHE.get(cache_key)
    if cached and cached["signature"] == signature:
        return cached["payload"]
    payload = builder()
    SHARED_PAYLOAD_CACHE[cache_key] = {"signature": signature, "payload": payload}
    return payload


def common_company_dependencies(company_id: str) -> list[Path]:
    return [
        sales_file(company_id),
        clients_file(company_id),
        stock_file(company_id),
        costing_file(company_id),
        costing_fabricated_file(company_id),
        price_table_file(company_id),
        vendors_file(company_id),
        vendor_goals_file(company_id),
        vendor_monthly_items_file(company_id),
        region_assignments_file(company_id),
        blocked_clients_file(company_id),
    ]


def dashboard_dependencies(company_id: str) -> list[Path]:
    company_ids = dashboard_company_ids(company_id)
    dependencies = []
    for current_company_id in company_ids:
        dependencies.extend([
            sales_file(current_company_id),
            clients_file(current_company_id),
            costing_file(current_company_id),
            costing_fabricated_file(current_company_id),
            blocked_clients_file(current_company_id),
        ])
    return dependencies


def valid_dashboard_company_id(company_id: str) -> bool:
    return company_id in COMPANIES or company_id == "agrupado" or company_id in DASHBOARD_COMPANY_GROUPS


def dashboard_company_ids(company_id: str) -> list[str]:
    if company_id == "agrupado":
        return list(COMPANIES)
    if company_id in DASHBOARD_COMPANY_GROUPS:
        return DASHBOARD_COMPANY_GROUPS[company_id]["companies"]
    return [company_id]


def dashboard_company_name(company_id: str) -> str:
    if company_id == "agrupado":
        return "Agrupado"
    if company_id in DASHBOARD_COMPANY_GROUPS:
        return DASHBOARD_COMPANY_GROUPS[company_id]["name"]
    return COMPANIES[company_id]


def vendor_sales_company_ids(company_id: str) -> list[str]:
    if company_id in {"ionlab", "ciorbrasil"}:
        return ["ionlab", "ciorbrasil"]
    return [company_id]


def vendor_dashboard_dependencies(company_id: str) -> list[Path]:
    dependencies = [
        clients_file(company_id),
        vendors_file(company_id),
        region_assignments_file(company_id),
        vendor_goals_file(company_id),
        vendor_monthly_items_file(company_id),
        vendor_day_by_day_file(company_id),
        blocked_clients_file(company_id),
    ]
    dependencies.extend(sales_file(current_company_id) for current_company_id in vendor_sales_company_ids(company_id))
    return dependencies


def normalize_text(value) -> str:
    text = clean_text_encoding(str(value or ""))
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    replacements = {
        "ß": "SS",
        "Æ": "AE",
        "æ": "AE",
        "Œ": "OE",
        "œ": "OE",
        "Ø": "O",
        "ø": "O",
        "Ð": "D",
        "ð": "D",
        "Þ": "TH",
        "þ": "TH",
        "Ł": "L",
        "ł": "L",
    }
    for source, target in replacements.items():
        text = text.replace(source, target)
    return re.sub(r"[^A-Z0-9]", "", text.upper())


SEARCH_STOP_WORDS = {"A", "O", "OS", "AS", "E", "DE", "DA", "DO", "DAS", "DOS", "PARA", "POR", "COM"}


def normalize_search_tokens(value) -> list[str]:
    text = clean_text_encoding(str(value or ""))
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    raw_tokens = re.split(r"[^A-Za-z0-9]+", text.upper())
    tokens = []
    for token in raw_tokens:
        normalized = normalize_text(token)
        if normalized and normalized not in SEARCH_STOP_WORDS:
            tokens.append(normalized)
    return tokens


def search_text_matches(query, haystack) -> bool:
    query_key = normalize_text(query)
    if not query_key:
        return True
    haystack_key = normalize_text(haystack)
    if query_key in haystack_key:
        return True
    tokens = normalize_search_tokens(query)
    return bool(tokens) and all(token in haystack_key for token in tokens)


def number_value(value) -> float:
    value = normalize_record(value)
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        text = str(value).strip().replace(".", "").replace(",", ".")
        try:
            return float(text)
        except ValueError:
            return 0.0


def optional_number_value(value):
    value = normalize_record(value)
    if value in (None, ""):
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        if "," in value and "." in value:
            value = value.replace(".", "").replace(",", ".")
        elif "," in value:
            value = value.replace(",", ".")
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def get_field(record: dict, *names):
    normalized_lookup = {
        normalize_text(key): value
        for key, value in record.items()
    }
    for name in names:
        normalized_name = normalize_text(name)
        if normalized_name in normalized_lookup:
            return normalized_lookup[normalized_name]
        if normalized_name.startswith("REFER"):
            for key, value in normalized_lookup.items():
                if key.startswith("REFER"):
                    return value
        if "DOLLAR" in normalized_name:
            for key, value in normalized_lookup.items():
                if "DOLLAR" in key:
                    return value
        if normalized_name.startswith("DESCR"):
            for key, value in normalized_lookup.items():
                if key.startswith("DESCR"):
                    return value
        if normalized_name.startswith("CUSTOBRLULTIMA"):
            for key, value in normalized_lookup.items():
                if key.startswith("CUSTOBRLULTIMA"):
                    return value
    return None


def record_year(value):
    value = normalize_record(value)
    if value in (None, ""):
        return None

    if isinstance(value, str):
        match = re.match(r"^(\d{4})", value)
        if match:
            return int(match.group(1))
        try:
            return pd.to_datetime(value).year
        except Exception:
            return None

    if isinstance(value, (datetime, date, pd.Timestamp)):
        return value.year

    return None


def record_month(value):
    value = normalize_record(value)
    if value in (None, ""):
        return None

    if isinstance(value, str):
        match = re.match(r"^\d{4}[-/](\d{1,2})", value)
        if match:
            return int(match.group(1))
        try:
            return pd.to_datetime(value).month
        except Exception:
            return None

    if isinstance(value, (datetime, date, pd.Timestamp)):
        return value.month

    return None


def record_date_value(value):
    value = normalize_record(value)
    if value in (None, ""):
        return None

    if isinstance(value, pd.Timestamp):
        return value.date()
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    if isinstance(value, str):
        text = value.strip()
        iso_match = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})", text)
        if iso_match:
            try:
                return date(int(iso_match.group(1)), int(iso_match.group(2)), int(iso_match.group(3)))
            except ValueError:
                return None

        br_match = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})", text)
        if br_match:
            try:
                return date(int(br_match.group(3)), int(br_match.group(2)), int(br_match.group(1)))
            except ValueError:
                return None

        for dayfirst in (False, True):
            try:
                parsed = pd.to_datetime(text, dayfirst=dayfirst)
                if pd.isna(parsed):
                    return None
                return parsed.date()
            except Exception:
                continue

    return None


def period_index(year: int, month: int) -> int:
    return year * 12 + month


def sale_period_index(sale: dict):
    year = record_year(sale.get("NF_EMI"))
    month = record_month(sale.get("NF_EMI"))
    if not year or not month:
        return None
    return period_index(year, month)


def sale_matches_reference(sale: dict, reference: str) -> bool:
    if not reference:
        return True

    raw_reference = str(reference or "").strip()
    candidates = [raw_reference]
    if " - " in raw_reference:
        before, after = raw_reference.split(" - ", 1)
        candidates.extend([before, after])

    product_code = normalize_text(sale.get("PR_COD"))
    product_description = normalize_text(sale.get("PR_DES"))
    product_haystack = normalize_text(f"{sale.get('PR_COD') or ''} {sale.get('PR_DES') or ''}")
    return any(
        candidate
        and (
            candidate in product_code
            or candidate in product_description
            or candidate in product_haystack
        )
        for candidate in (normalize_text(item) for item in candidates)
    )


def sale_net_revenue(sale: dict) -> float:
    return (
        number_value(sale.get("PR_SBT"))
        - number_value(sale.get("V_FRETE_NF"))
        - number_value(sale.get("V_DESCONTO"))
    )


def is_excluded_group_sale(company_id: str, customer_name) -> bool:
    normalized_customer = normalize_text(customer_name)
    if not normalized_customer:
        return False

    return any(
        normalize_text(company_name) in normalized_customer
        for company_name in COMPANY_SALE_EXCLUSIONS.get(company_id, [])
    )


def is_group_company_client(client: dict) -> bool:
    normalized_name = normalize_text(client.get("NOM"))
    normalized_fantasy = normalize_text(client.get("FAN"))
    if not normalized_name and not normalized_fantasy:
        return False
    return any(
        normalize_text(company_name) in normalized_name or normalize_text(company_name) in normalized_fantasy
        for company_name in GROUP_COMPANY_NAMES
    )


def client_ddd(client: dict) -> str:
    digits = re.sub(r"\D", "", str(client.get("TEL") or ""))
    return digits[:2] if len(digits) >= 2 else ""


def economic_client_key(client_id: str, client: dict | None = None, sale: dict | None = None) -> str:
    document = normalize_identifier((client or {}).get("CGC"))
    if len(document) == 14:
        return f"CNPJROOT:{document[:8]}"
    if document:
        return f"DOC:{document}"
    sale_document = normalize_identifier((sale or {}).get("CL_CFIN"))
    if len(sale_document) == 14:
        return f"CNPJROOT:{sale_document[:8]}"
    if sale_document:
        return f"DOC:{sale_document}"
    return f"ID:{normalize_identifier(client_id)}"


def vendor_can_manage(role: str) -> bool:
    return role in ("master", "supervisor")


def value_matches(record: dict, query: str) -> bool:
    if not query:
        return True

    for key, value in record.items():
        if key.startswith("_") or value is None:
            continue
        if search_text_matches(query, value):
            return True
    return False


def format_days_without_movement(days: int) -> str:
    if days < 0:
        return ""
    if days <= 30:
        return f"{days} dia" if days == 1 else f"{days} dias"

    months = days // 30
    remaining_days = days % 30
    month_label = "mes" if months == 1 else "meses"
    day_label = "dia" if remaining_days == 1 else "dias"
    return f"{months} {month_label} e {remaining_days} {day_label}"


def stock_records_with_last_exit(company_id: str):
    return cached_payload(
        ("stock_records_with_last_exit", company_id, date.today().isoformat()),
        [stock_file(company_id), sales_file(company_id)],
        lambda: build_stock_records_with_last_exit(company_id),
    )


def build_stock_records_with_last_exit(company_id: str):
    records = load_json(stock_file(company_id), [])
    latest_exit_by_reference = {}
    for sale in load_json(sales_file(company_id), []):
        key = normalize_text(sale.get("PR_COD"))
        if not key:
            continue
        sale_date = record_date_value(sale.get("NF_EMI"))
        if not sale_date:
            continue
        current_date = latest_exit_by_reference.get(key)
        if current_date is None or sale_date > current_date:
            latest_exit_by_reference[key] = sale_date

    enriched_records = []
    today = date.today()
    for record in records:
        enriched = dict(record)
        latest_date = latest_exit_by_reference.get(normalize_text(record.get("Referencia")))
        enriched["Ultima Saida"] = latest_date.isoformat() if latest_date else ""
        enriched["Dias sem movimento"] = format_days_without_movement((today - latest_date).days) if latest_date else ""
        enriched_records.append(enriched)
    return enriched_records


def slow_items_payload(company_id: str, query: str = ""):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    saved_percentages = load_json(slow_items_file(company_id), {})
    if not isinstance(saved_percentages, dict):
        saved_percentages = {}

    today = date.today()
    price_by_reference = {
        normalize_text(record.get("PR_COD")): number_value(record.get("VAL_VEND"))
        for record in load_json(price_table_file(company_id), [])
        if normalize_text(record.get("PR_COD"))
    }
    rows = []
    for record in stock_records_with_last_exit(company_id):
        reference = normalize_record(record.get("Referencia"))
        if not reference:
            continue
        quantity = number_value(record.get("Quantidade"))
        if quantity <= 0:
            continue

        agrp = normalize_record(record.get("AGRP"))
        if is_slow_item_agrp_excluded(agrp):
            continue

        last_exit = record_date_value(record.get("Ultima Saida"))
        days_without_movement = (today - last_exit).days if last_exit else 999999
        if last_exit and days_without_movement <= 365:
            continue

        key = normalize_text(reference)
        has_price = key in price_by_reference and price_by_reference.get(key, 0) > 0
        unit_price = price_by_reference.get(key, 0.0) if has_price else 0.0
        saved = saved_percentages.get(key, {})
        row = {
            "referencia": reference,
            "descricao": normalize_record(record.get("Descricao") or record.get("Descricao item")),
            "agrp": agrp,
            "origem": normalize_record(record.get("Origem")),
            "quantidade": round(quantity, 3),
            "valor_unitario": round(unit_price, 3),
            "sem_preco_tabela": not has_price,
            "total": round(quantity * unit_price, 2),
            "ultima_saida": display_date_br(last_exit) if last_exit else "",
            "dias_sem_movimento": format_days_without_movement(days_without_movement) if last_exit else "Nunca vendeu",
            "percentual": round(number_value(saved.get("percentual")), 3),
            "_dias_sem_movimento": days_without_movement,
        }
        if value_matches(row, query):
            rows.append(row)

    rows.sort(key=lambda item: (-number_value(item.get("_dias_sem_movimento")), normalize_text(item.get("referencia"))))

    return {
        "empresa": COMPANIES[company_id],
        "empresa_id": company_id,
        "total": len(rows),
        "columns": [
            {"key": "referencia", "label": "Referencia"},
            {"key": "descricao", "label": "Descricao"},
            {"key": "agrp", "label": "AGRP"},
            {"key": "origem", "label": "Origem"},
            {"key": "quantidade", "label": "Quantidade"},
            {"key": "valor_unitario", "label": "V.Unitario"},
            {"key": "total", "label": "Total"},
            {"key": "ultima_saida", "label": "Ultima Saida"},
            {"key": "dias_sem_movimento", "label": "Dias sem movimento"},
            {"key": "percentual", "label": "% comissao"},
        ],
        "rows": rows,
    }


def is_slow_item_agrp_excluded(value) -> bool:
    agrp_key = normalize_text(value)
    return agrp_key == "MATERIAPRIMA" or "PECADEREPOSICAO" in agrp_key


def save_slow_items_payload(payload: dict):
    company_id = normalize_record(payload.get("company")) or "ionlab"
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    items = payload.get("items") or []
    if not isinstance(items, list):
        raise ValueError("Lista de itens invalida.")

    saved_percentages = load_json(slow_items_file(company_id), {})
    if not isinstance(saved_percentages, dict):
        saved_percentages = {}

    now = datetime.now().isoformat(timespec="seconds")
    updated = 0
    for item in items:
        reference = normalize_record(item.get("referencia"))
        if not reference:
            continue
        saved_percentages[normalize_text(reference)] = {
            "referencia": reference,
            "percentual": round(max(0, number_value(item.get("percentual"))), 3),
            "atualizado_em": now,
        }
        updated += 1

    save_json(slow_items_file(company_id), saved_percentages)
    result = slow_items_payload(company_id, normalize_record(payload.get("query")))
    result["message"] = f"{updated} percentual(is) salvo(s)."
    return result


def slow_items_export_xlsx(company_id: str, query: str = "") -> bytes:
    payload = slow_items_payload(company_id, query)
    columns = [
        ("referencia", "Referencia"),
        ("descricao", "Descricao"),
        ("agrp", "AGRP"),
        ("origem", "Origem"),
        ("quantidade", "Quantidade"),
        ("valor_unitario", "V.Unitario"),
        ("total", "Total"),
        ("ultima_saida", "Ultima Saida"),
        ("dias_sem_movimento", "Dias sem movimento"),
        ("percentual", "% comissao"),
        ("sem_preco_tabela_label", "Sem preco na tabela"),
    ]
    rows = []
    for row in payload["rows"]:
        export_row = dict(row)
        export_row["sem_preco_tabela_label"] = "Sim" if row.get("sem_preco_tabela") else "Nao"
        rows.append({
            label: export_row.get(key)
            for key, label in columns
        })
    return export_records_xlsx(rows, [label for _key, label in columns], "Itens Sem Giro")


def table_sort_value(record: dict, key: str):
    if key in ("Quantidade", "QTY", "V.Unitario", "Total", "Preco Dollar", "Custo BRL Indice", "Custo BRL Ultima Importacao", "PR_QTD", "PR_SBT", "LUCRO_REAIS", "Aliquota IPI", "Aliquota de IPI"):
        return number_value(record.get(key))
    if key in ("Ultima Saida", "Previsao", "NF_EMI", "importado_em", "bloqueado_em"):
        parsed = record_date_value(record.get(key))
        return parsed.toordinal() if parsed else -1
    if key == "Dias sem movimento":
        last_exit = record_date_value(record.get("Ultima Saida"))
        return (date.today() - last_exit).days if last_exit else -1
    return normalize_text(record.get(key))


def table_payload(company_id: str, table_name: str, query: str, limit: int, sort_key: str = "", sort_dir: str = ""):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    if table_name == "sales":
        records = load_json(sales_file(company_id), [])
        columns = SALES_TABLE_COLUMNS
    elif table_name == "clients":
        records = load_json(clients_file(company_id), [])
        columns = CLIENT_TABLE_COLUMNS
    elif table_name == "stock":
        records = stock_records_with_last_exit(company_id)
        columns = STOCK_TABLE_COLUMNS
    elif table_name == "in-transit":
        records = load_json(in_transit_file(company_id), [])
        columns = IN_TRANSIT_TABLE_COLUMNS
    elif table_name == "costing":
        records = load_json(costing_file(company_id), [])
        config = load_json(costing_config_file(company_id), {})
        columns = COSTING_TABLE_COLUMNS
    elif table_name == "costing-fabricated":
        records = load_json(costing_fabricated_file(company_id), [])
        columns = COSTING_TABLE_COLUMNS
    elif table_name == "prices":
        records = price_table_records_with_cost(company_id, save=True)
        columns = PRICE_TABLE_COLUMNS
    elif table_name == "products":
        records = load_json(products_file(company_id), [])
        columns = PRODUCT_TABLE_COLUMNS
    else:
        raise ValueError("Tabela invalida.")

    limit = max(1, min(limit, 500))
    matches = [record for record in records if value_matches(record, query)]
    allowed_sort_keys = {key for key, _label in columns}
    if sort_key not in allowed_sort_keys:
        sort_key = "Total" if table_name == "stock" else ""
    reverse_sort = sort_dir == "desc" or (table_name == "stock" and sort_key == "Total" and sort_dir != "asc")
    if sort_key:
        matches = sorted(
            matches,
            key=lambda record: (
                table_sort_value(record, sort_key),
                normalize_text(record.get("Referencia") or record.get("ID") or record.get("NF_NUM") or record.get(sort_key)),
            ),
            reverse=reverse_sort,
        )
    rows = [
        {
            key: record.get(key)
            for key, _label in columns
        }
        for record in matches[:limit]
    ]

    payload = {
        "empresa": COMPANIES[company_id],
        "empresa_id": company_id,
        "tabela": table_name,
        "consulta": query,
        "total_registros": len(records),
        "total_filtrado": len(matches),
        "limite": limit,
        "sort_key": sort_key,
        "sort_dir": "desc" if reverse_sort else "asc",
        "columns": [
            {"key": key, "label": label}
            for key, label in columns
        ],
        "rows": rows,
    }
    if table_name == "costing":
        payload["indice_custeio"] = optional_number_value(config.get("indice_custeio")) or 0
    return payload


def mercado_livre_import_id(value: str) -> str:
    base = normalize_text(value) or "IMPORTACAO"
    base = re.sub(r"[^A-Z0-9]+", "-", base).strip("-").lower()
    return base[:70] or "importacao"


MERCADO_LIVRE_IMPORT_TYPES = {
    "anuncios": "Anuncios",
    "ficha_tecnica": "Ficha Tecnica",
    "resumo": "Resumo",
}


def mercado_livre_import_type(value: str) -> str:
    normalized = normalize_text(value)
    if normalized in ("FICHATECNICA", "FICHA-TECNICA", "TECHNICALSHEET"):
        return "ficha_tecnica"
    if normalized == "RESUMO":
        return "resumo"
    return "anuncios"


def mercado_livre_load_imports(company_id: str) -> list[dict]:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    imports = load_json(mercado_livre_imports_file(company_id), [])
    if isinstance(imports, list) and imports:
        changed = False
        for item in imports:
            if not item.get("tipo"):
                item["tipo"] = "anuncios"
                item["tipo_nome"] = MERCADO_LIVRE_IMPORT_TYPES["anuncios"]
                changed = True
            elif not item.get("tipo_nome"):
                item["tipo_nome"] = MERCADO_LIVRE_IMPORT_TYPES.get(item.get("tipo"), item.get("tipo"))
                changed = True
        if changed:
            save_json(mercado_livre_imports_file(company_id), imports)
        return imports

    legacy_records = load_json(mercado_livre_ads_file(company_id), [])
    if not legacy_records:
        return []

    legacy_id = "anuncios"
    legacy_file = mercado_livre_import_file(company_id, legacy_id)
    if not legacy_file.exists():
        save_json(legacy_file, legacy_records)
    imported_at = normalize_record((legacy_records[0] or {}).get("_importado_em")) if legacy_records else ""
    imports = [{
        "id": legacy_id,
        "nome": "Anuncios",
        "arquivo": "Anuncios",
        "aba": "Anuncios",
        "tipo": "anuncios",
        "tipo_nome": MERCADO_LIVRE_IMPORT_TYPES["anuncios"],
        "linhas": len(legacy_records),
        "importado_em": imported_at or datetime.now().isoformat(timespec="seconds"),
    }]
    save_json(mercado_livre_imports_file(company_id), imports)
    return imports


def mercado_livre_overview_payload():
    companies_payload = []
    for company_id in MERCADO_LIVRE_COMPANIES:
        companies_payload.append({
            "id": company_id,
            "name": COMPANIES.get(company_id, company_id),
            "imports": mercado_livre_load_imports(company_id),
        })
    return {"companies": companies_payload}


def mercado_livre_ads_payload(company_id: str, query: str = "", limit: int = 300, import_id: str = "", import_type: str = ""):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    imports = mercado_livre_load_imports(company_id)
    selected_type = mercado_livre_import_type(import_type) if import_type else ""
    if selected_type == "resumo":
        return mercado_livre_summary_payload(company_id, query, limit)
    typed_imports = [item for item in imports if not selected_type or item.get("tipo", "anuncios") == selected_type]
    selected_import = None
    if import_id:
        selected_import = next((item for item in typed_imports if item.get("id") == import_id), None)
    if selected_import is None and typed_imports:
        selected_import = typed_imports[0]
    records = load_json(mercado_livre_import_file(company_id, selected_import.get("id")), []) if selected_import else []
    columns = []
    for record in records:
        for key in record:
            if not str(key).startswith("_") and key not in columns:
                columns.append(key)
        if columns:
            break

    matches = [record for record in records if value_matches(record, query)]
    limit = max(1, min(int(limit or 300), 1000))
    return {
        "empresa": COMPANIES[company_id],
        "empresa_id": company_id,
        "imports": imports,
        "import_id": selected_import.get("id") if selected_import else "",
        "import_name": selected_import.get("nome") if selected_import else "",
        "tipo": selected_import.get("tipo") if selected_import else selected_type,
        "tipo_nome": selected_import.get("tipo_nome") if selected_import else MERCADO_LIVRE_IMPORT_TYPES.get(selected_type, ""),
        "columns": [{"key": key, "label": key} for key in columns],
        "rows": [
            {key: record.get(key) for key in columns}
            for record in matches[:limit]
        ],
        "total_filtrado": len(matches),
        "total_registros": len(records),
        "limite": limit,
    }


def prospect_options_payload(company_id: str, uf: str = ""):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    clients = load_json(clients_file(company_id), [])
    sales = load_json(sales_file(company_id), [])
    blocked_clients = blocked_client_map(company_id)
    selected_uf = uf.strip().upper()

    ufs = sorted({
        str(client.get("UF") or "").strip().upper()
        for client in clients
        if normalize_identifier(client.get("ID")) not in blocked_clients
        if str(client.get("UF") or "").strip()
    })
    cities = sorted({
        str(client.get("CID") or "").strip()
        for client in clients
        if normalize_identifier(client.get("ID")) not in blocked_clients
        if str(client.get("CID") or "").strip()
        and (not selected_uf or str(client.get("UF") or "").strip().upper() == selected_uf)
    }, key=lambda item: normalize_text(item))

    references = {}
    for sale in sales:
        code = str(sale.get("PR_COD") or "").strip()
        description = str(sale.get("PR_DES") or "").strip()
        if not code and not description:
            continue
        label = f"{code} - {description}" if code and description else code or description
        references[normalize_text(label)] = label

    return {
        "ufs": ufs,
        "cities": cities,
        "references": sorted(references.values(), key=lambda item: normalize_text(item))[:1000],
    }


def prospect_client_search_payload(company_id: str, query: str, uf: str = "", city: str = "", limit: int = 20):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    clients = load_json(clients_file(company_id), [])
    blocked_clients = blocked_client_map(company_id)
    normalized_query = normalize_text(query)
    selected_uf = uf.strip().upper()
    selected_city = normalize_text(city)
    matches = []

    for client in clients:
        if selected_uf and str(client.get("UF") or "").strip().upper() != selected_uf:
            continue
        if selected_city and normalize_text(client.get("CID")) != selected_city:
            continue

        client_id = normalize_identifier(client.get("ID"))
        if client_id in blocked_clients:
            continue
        client_name = str(client.get("NOM") or "")
        haystack = f"{client_id} {client_name}"
        if normalized_query and not search_text_matches(query, haystack):
            continue

        matches.append({
            "id": client_id,
            "name": client_name,
            "uf": client.get("UF"),
            "city": client.get("CID"),
            "label": f"{client_id} - {client_name}",
        })
        if len(matches) >= limit:
            break

    return {"clients": matches}


def prospect_payload(params):
    company_id = params.get("company", [""])[0]
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    start_month = int(params.get("start_month", ["0"])[0])
    end_month = int(params.get("end_month", ["0"])[0])
    start_year = int(params.get("start_year", ["0"])[0])
    end_year = int(params.get("end_year", ["0"])[0])
    if not all([start_month, end_month, start_year, end_year]):
        raise ValueError("Preencha mes inicial, mes final, ano inicial e ano final.")
    if not (1 <= start_month <= 12 and 1 <= end_month <= 12):
        raise ValueError("Mes invalido.")

    start_index = period_index(start_year, start_month)
    end_index = period_index(end_year, end_month)
    if start_index > end_index:
        raise ValueError("Periodo inicial nao pode ser maior que periodo final.")

    selected_uf = params.get("uf", [""])[0].strip().upper()
    selected_city = normalize_text(params.get("city", [""])[0])
    selected_client = normalize_identifier(params.get("client_id", [""])[0])
    selected_client_query = normalize_text(params.get("client_query", [""])[0])
    reference = params.get("reference", [""])[0].strip()
    client_type = params.get("client_type", ["all"])[0]

    clients = load_json(clients_file(company_id), [])
    sales = load_json(sales_file(company_id), [])
    blocked_clients = blocked_client_map(company_id)
    client_map = {}
    last_purchase_by_client = {}
    last_purchase_revenue_by_client = {}

    for client in clients:
        client_id = normalize_identifier(client.get("ID"))
        if not client_id:
            continue
        if client_id in blocked_clients:
            continue
        if selected_uf and str(client.get("UF") or "").strip().upper() != selected_uf:
            continue
        if selected_city and normalize_text(client.get("CID")) != selected_city:
            continue
        if selected_client and client_id != selected_client:
            continue
        if selected_client_query and not selected_client:
            client_haystack = normalize_text(f"{client_id} {client.get('NOM') or ''}")
            if selected_client_query not in client_haystack:
                continue
        client_map[client_id] = {
            "client": client,
            "sales": [],
            "revenue": 0.0,
            "last_purchase": None,
        }

    for sale in sales:
        if is_excluded_group_sale(company_id, sale.get("CL_NOM")):
            continue

        client_id = normalize_identifier(sale.get("ID_CL"))
        sale_date = normalize_record(sale.get("NF_EMI"))
        if client_id and sale_date:
            previous_date = last_purchase_by_client.get(client_id)
            if previous_date is None or str(sale_date) > str(previous_date):
                last_purchase_by_client[client_id] = sale_date
                last_purchase_revenue_by_client[client_id] = sale_net_revenue(sale)
            elif str(sale_date) == str(previous_date):
                last_purchase_revenue_by_client[client_id] = (
                    last_purchase_revenue_by_client.get(client_id, 0.0) + sale_net_revenue(sale)
                )

        sale_index = sale_period_index(sale)
        if sale_index is None or sale_index < start_index or sale_index > end_index:
            continue
        if not sale_matches_reference(sale, reference):
            continue

        if client_id not in client_map:
            continue

        entry = client_map[client_id]
        entry["sales"].append(sale)
        entry["revenue"] += sale_net_revenue(sale)
        if sale_date and (entry["last_purchase"] is None or str(sale_date) > str(entry["last_purchase"])):
            entry["last_purchase"] = sale_date

    rows = []
    active_count = 0
    inactive_count = 0
    never_bought_count = 0
    for client_id, entry in client_map.items():
        sales_count = len(entry["sales"])
        if reference and sales_count == 0:
            continue
        has_historical_purchase = client_id in last_purchase_by_client
        status = "Ativo" if sales_count > 0 else "Inativo"
        display_status = "Nunca Comprou" if not has_historical_purchase else status

        if client_type == "active" and status != "Ativo":
            continue
        if client_type == "inactive" and status != "Inativo":
            continue
        if client_type == "never" and has_historical_purchase:
            continue

        if status == "Ativo":
            active_count += 1
        else:
            inactive_count += 1
        if not has_historical_purchase:
            never_bought_count += 1

        client = entry["client"]
        rows.append({
            "ID": client.get("ID"),
            "NOM": client.get("NOM"),
            "UF": client.get("UF"),
            "CID": client.get("CID"),
            "status": display_status,
            "compras_periodo": sales_count,
            "faturamento_liquido": round(
                entry["revenue"] if status == "Ativo" else last_purchase_revenue_by_client.get(client_id, 0.0),
                2,
            ),
            "ultima_compra": entry["last_purchase"] if status == "Ativo" else last_purchase_by_client.get(client_id, "Nunca Comprou"),
        })

    rows.sort(key=lambda row: (row["status"], normalize_text(row.get("UF")), normalize_text(row.get("NOM"))))

    return {
        "empresa": COMPANIES[company_id],
        "periodo": {
            "mes_inicial": start_month,
            "ano_inicial": start_year,
            "mes_final": end_month,
            "ano_final": end_year,
        },
        "filters": {
            "uf": selected_uf,
            "city": params.get("city", [""])[0],
            "client_id": selected_client,
            "client_query": params.get("client_query", [""])[0],
            "reference": reference,
            "client_type": client_type,
        },
        "total_base": len(rows),
        "ativos": active_count,
        "inativos": inactive_count,
        "nunca_comprou": never_bought_count,
        "total_resultado": len(rows),
        "faturamento_liquido": round(sum(row["faturamento_liquido"] for row in rows), 2),
        "columns": [{"key": key, "label": label} for key, label in PROSPECT_COLUMNS],
        "rows": rows[:500],
        "limit": 500,
    }


def clients_by_uf_payload(company_id: str):
    if not valid_dashboard_company_id(company_id):
        raise ValueError("Empresa invalida.")

    years = list(range(START_YEAR, CURRENT_YEAR + 1))
    counts = {uf: 0 for uf in BRAZIL_UFS}
    client_ids_by_uf = {uf: set() for uf in BRAZIL_UFS}
    active_client_ids_by_uf_year = {
        uf: {str(year): set() for year in years}
        for uf in BRAZIL_UFS
    }
    revenue = {
        uf: {str(year): 0.0 for year in years}
        for uf in BRAZIL_UFS
    }
    without_uf = 0
    sales_without_uf = 0
    excluded_sales = 0
    total_clients = 0

    company_ids = dashboard_company_ids(company_id)

    for current_company_id in company_ids:
        clients = load_json(clients_file(current_company_id), [])
        total_clients += len(clients)
        for client in clients:
            uf = str(client.get("UF") or "").strip().upper()
            if uf in counts:
                counts[uf] += 1
                client_id = normalize_identifier(client.get("ID"))
                if client_id:
                    client_ids_by_uf[uf].add(f"{current_company_id}|{client_id}")
            else:
                without_uf += 1

        for sale in load_json(sales_file(current_company_id), []):
            if is_excluded_group_sale(current_company_id, sale.get("CL_NOM")):
                excluded_sales += 1
                continue

            year = record_year(sale.get("NF_EMI"))
            if year not in years:
                continue

            uf = str(sale.get("CL_UF") or "").strip().upper()
            if uf not in revenue:
                sales_without_uf += 1
                continue

            client_id = normalize_identifier(sale.get("ID_CL"))
            year_key = str(year)
            scoped_client_id = f"{current_company_id}|{client_id}"
            if client_id and scoped_client_id in client_ids_by_uf[uf]:
                active_client_ids_by_uf_year[uf][year_key].add(scoped_client_id)

            revenue[uf][str(year)] += sale_net_revenue(sale)

    rows = [
        {
            "uf": uf,
            "clientes": counts[uf],
            "faturamento_liquido": {
                str(year): round(revenue[uf][str(year)], 2)
                for year in years
            },
            "faturamento_total": round(sum(revenue[uf].values()), 2),
            "atividade": {
                str(year): {
                    "ativos": len(active_client_ids_by_uf_year[uf][str(year)]),
                    "inativos": max(0, counts[uf] - len(active_client_ids_by_uf_year[uf][str(year)])),
                }
                for year in years
            },
        }
        for uf in BRAZIL_UFS
    ]
    top = max(rows, key=lambda item: item["clientes"], default={"uf": "", "clientes": 0})
    top_revenue = max(rows, key=lambda item: item["faturamento_total"], default={"uf": "", "faturamento_total": 0})
    active_ufs = sum(1 for row in rows if row["clientes"] > 0)
    totals_by_year = {
        str(year): round(sum(revenue[uf][str(year)] for uf in BRAZIL_UFS), 2)
        for year in years
    }

    return {
        "empresa": dashboard_company_name(company_id),
        "empresa_id": company_id,
        "years": years,
        "total_clientes": total_clients,
        "ufs_com_clientes": active_ufs,
        "sem_uf": without_uf,
        "vendas_sem_uf": sales_without_uf,
        "vendas_grupo_excluidas": excluded_sales,
        "maior_uf": top,
        "maior_faturamento_uf": top_revenue,
        "faturamento_total": round(sum(totals_by_year.values()), 2),
        "faturamento_por_ano": totals_by_year,
        "calculo_faturamento": "PR_SBT - V_FRETE_NF - V_DESCONTO",
        "rows": rows,
    }


def product_origin_matches(origin: str, selected_origin: str) -> bool:
    if selected_origin == "all":
        return True
    normalized_origin = normalize_text(origin)
    is_imported = "IMPORTADO" in normalized_origin
    is_fabricated = "FABRICADO" in normalized_origin
    if selected_origin == "imported":
        return is_imported
    if selected_origin == "fabricated_br":
        return is_fabricated
    return False


def dashboard_origin_options() -> list[dict]:
    return [
        {"id": "all", "name": "Todas as origens"},
        {"id": "imported", "name": "Importado"},
        {"id": "fabricated_br", "name": "Fabricados BR"},
    ]


def dashboard_sales_annual_payload(company_id: str, selected_year: str = "all", selected_origin: str = "all", selected_agrp: str = "all"):
    if not valid_dashboard_company_id(company_id):
        raise ValueError("Empresa invalida.")

    years = list(range(START_YEAR, CURRENT_YEAR + 1))
    company_ids = dashboard_company_ids(company_id)
    selected_year = str(selected_year or "all")
    selected_origin = str(selected_origin or "all")
    selected_agrp = str(selected_agrp or "all")
    if selected_year != "all":
        try:
            selected_year_int = int(selected_year)
        except ValueError as exc:
            raise ValueError("Ano invalido.") from exc
        if selected_year_int not in years:
            raise ValueError("Ano invalido.")
    if selected_origin not in {"all", "imported", "fabricated_br"}:
        raise ValueError("Origem invalida.")

    companies_payload = []
    grouped_years = {str(year): 0.0 for year in years}
    grouped_months = {
        str(year): {str(month): 0.0 for month in range(1, 13)}
        for year in years
    }
    families = set()
    excluded_sales = 0

    for current_company_id in company_ids:
        costing_by_ref = all_costing_info_by_reference(current_company_id)
        yearly = {str(year): 0.0 for year in years}
        monthly = {
            str(year): {str(month): 0.0 for month in range(1, 13)}
            for year in years
        }
        for sale in load_json(sales_file(current_company_id), []):
            if is_excluded_group_sale(current_company_id, sale.get("CL_NOM")):
                excluded_sales += 1
                continue
            year = record_year(sale.get("NF_EMI"))
            month = record_month(sale.get("NF_EMI"))
            if year not in years or not month or not (1 <= month <= 12):
                continue
            costing_info = classification_for_reference(sale.get("PR_COD"), costing_by_ref)
            family = normalize_record(sale.get("AGRP")) or costing_info.get("AGRP") or "Sem AGRP"
            origin = normalize_record(sale.get("Origem")) or costing_info.get("Origem") or "Sem origem"
            if not product_origin_matches(origin, selected_origin):
                continue
            families.add(family)
            if selected_agrp != "all" and family != selected_agrp:
                continue
            value = sale_net_revenue(sale)
            year_key = str(year)
            month_key = str(month)
            yearly[year_key] += value
            monthly[year_key][month_key] += value
            grouped_years[year_key] += value
            grouped_months[year_key][month_key] += value

        companies_payload.append({
            "empresa_id": current_company_id,
            "empresa": COMPANIES[current_company_id],
            "faturamento_por_ano": {
                year: round(value, 2)
                for year, value in yearly.items()
            },
            "faturamento_por_mes": {
                year: {
                    month: round(value, 2)
                    for month, value in months.items()
                }
                for year, months in monthly.items()
            },
        })

    return {
        "empresa": dashboard_company_name(company_id),
        "empresa_id": company_id,
        "years": years,
        "selected_year": selected_year,
        "selected_origin": selected_origin,
        "selected_agrp": selected_agrp,
        "origins": dashboard_origin_options(),
        "families": sorted(families, key=lambda item: normalize_text(item)),
        "companies": companies_payload,
        "agrupado": {
            "empresa": dashboard_company_name(company_id),
            "faturamento_por_ano": {
                year: round(value, 2)
                for year, value in grouped_years.items()
            },
            "faturamento_por_mes": {
                year: {
                    month: round(value, 2)
                    for month, value in months.items()
                }
                for year, months in grouped_months.items()
            },
        },
        "vendas_grupo_excluidas": excluded_sales,
        "calculo_faturamento": "PR_SBT - V_FRETE_NF - V_DESCONTO",
    }


def dashboard_sales_family_payload(company_id: str, selected_year: str = "all", selected_origin: str = "all"):
    if not valid_dashboard_company_id(company_id):
        raise ValueError("Empresa invalida.")

    years = list(range(START_YEAR, CURRENT_YEAR + 1))
    company_ids = dashboard_company_ids(company_id)
    selected_year = str(selected_year or "all")
    selected_origin = str(selected_origin or "all")
    if selected_year != "all":
        try:
            selected_year_int = int(selected_year)
        except ValueError as exc:
            raise ValueError("Ano invalido.") from exc
        if selected_year_int not in years:
            raise ValueError("Ano invalido.")
    if selected_origin not in {"all", "imported", "fabricated_br"}:
        raise ValueError("Origem invalida.")

    grouped = {}
    company_family = {}
    excluded_sales = 0

    for current_company_id in company_ids:
        costing_by_ref = costing_info_by_reference(current_company_id)
        company_family[current_company_id] = {}
        for sale in load_json(sales_file(current_company_id), []):
            if is_excluded_group_sale(current_company_id, sale.get("CL_NOM")):
                excluded_sales += 1
                continue
            year = record_year(sale.get("NF_EMI"))
            if year not in years:
                continue
            costing_info = classification_for_reference(sale.get("PR_COD"), costing_by_ref)
            origin = normalize_record(sale.get("Origem")) or costing_info.get("Origem") or "Sem origem"
            if not product_origin_matches(origin, selected_origin):
                continue
            family = normalize_record(sale.get("AGRP"))
            if not family:
                family = costing_info.get("AGRP") or "Sem AGRP"
            year_key = str(year)
            value = sale_net_revenue(sale)
            grouped.setdefault(family, {str(item): 0.0 for item in years})
            company_family[current_company_id].setdefault(family, {str(item): 0.0 for item in years})
            grouped[family][year_key] += value
            company_family[current_company_id][family][year_key] += value

    families = sorted(grouped, key=lambda item: normalize_text(item))
    rows = []
    for family in families:
        yearly = grouped[family]
        rows.append({
            "agrp": family,
            "faturamento_por_ano": {
                year: round(value, 2)
                for year, value in yearly.items()
            },
            "total": round(sum(yearly.values()), 2),
            "empresas": [
                {
                    "empresa_id": current_company_id,
                    "empresa": COMPANIES[current_company_id],
                    "faturamento_por_ano": {
                        year: round(company_family.get(current_company_id, {}).get(family, {}).get(year, 0.0), 2)
                        for year in map(str, years)
                    },
                }
                for current_company_id in company_ids
            ],
        })

    return {
        "empresa": dashboard_company_name(company_id),
        "empresa_id": company_id,
        "years": years,
        "selected_year": selected_year,
        "selected_origin": selected_origin,
        "origins": dashboard_origin_options(),
        "families": families,
        "rows": rows,
        "vendas_grupo_excluidas": excluded_sales,
        "calculo_faturamento": "PR_SBT - V_FRETE_NF - V_DESCONTO",
    }


def empty_region_summary(label: str, key: str, years: list, months: list) -> dict:
    return {
        "key": key,
        "label": label,
        "clientes": 0,
        "ativos": 0,
        "inativos": 0,
        "nunca_comprou": 0,
        "vendas_anos": {str(year): 0.0 for year in years},
        "vendas_meses": {str(month): 0.0 for month in months},
        "ativos_anos": {str(year): 0 for year in years},
        "ativos_meses": {str(month): 0 for month in months},
        "inativos_anos": {str(year): 0 for year in years},
        "inativos_meses": {str(month): 0 for month in months},
        "nunca_anos": {str(year): 0 for year in years},
        "nunca_meses": {str(month): 0 for month in months},
        "vendas_meses_anos": {
            str(year): {str(month): 0.0 for month in range(1, 13)}
            for year in years
        },
        "_client_ids": set(),
    }


def finalize_region_summary(row: dict) -> dict:
    payload = dict(row)
    payload.pop("_client_ids", None)
    payload["vendas_anos"] = {period: round(value, 2) for period, value in payload["vendas_anos"].items()}
    payload["vendas_meses"] = {period: round(value, 2) for period, value in payload["vendas_meses"].items()}
    payload["vendas_meses_anos"] = {
        year: {month: round(value, 2) for month, value in months.items()}
        for year, months in payload.get("vendas_meses_anos", {}).items()
    }
    return payload


def vendor_region_client_context(company_id: str, vendor_id_value: str):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    if not vendor_id_value:
        raise ValueError("Selecione um vendedor.")

    vendors, _inserted = sync_vendors_from_sales(company_id)
    vendor = next((item for item in vendors if item.get("id") == vendor_id_value), None)
    if not vendor:
        raise ValueError("Vendedor nao encontrado.")

    years = list(range(START_YEAR, CURRENT_YEAR + 1))
    closed_month_count = max(1, date.today().month - 1) if CURRENT_YEAR == date.today().year else 12
    months = list(range(1, date.today().month + 1)) if CURRENT_YEAR == date.today().year else list(range(1, 13))
    client_index = {}
    for client_company_id in reversed(vendor_sales_company_ids(company_id)):
        for client in valid_clients_for_assignment(client_company_id):
            client_id = normalize_identifier(client.get("ID"))
            if client_id:
                client_index[client_id] = client

    region_client_ids = clients_for_vendor_region(company_id, vendor_id_value, client_index, vendor)

    return vendor, client_index, region_client_ids


def active_region_rules(company_id: str) -> list[dict]:
    rules = load_json(region_assignments_file(company_id), [])
    return [rule for rule in rules if rule.get("status", "Ativo") == "Ativo"]


REGION_PRIORITY = {"uf": 1, "ddd": 2, "cidade": 3, "cliente": 4, "exclusao_ddd": 5, "exclusao_cidade": 6, "exclusao_cliente": 7}


def rule_matches_client(rule: dict, client_id: str, client: dict) -> bool:
    rule_type = str(rule.get("tipo") or "").strip().lower()
    uf = str(client.get("UF") or "").strip().upper()
    city = normalize_text(client.get("CID"))
    ddd = client_ddd(client)
    if rule_type in ("cliente", "exclusao_cliente"):
        return normalize_identifier(rule.get("cliente_id")) == client_id
    if str(rule.get("uf") or "").strip().upper() != uf:
        return False
    if rule_type == "uf":
        return True
    if rule_type in ("ddd", "exclusao_ddd"):
        return re.sub(r"\D", "", str(rule.get("ddd") or ""))[:2] == ddd
    if rule_type in ("cidade", "exclusao_cidade"):
        rule_ddd = re.sub(r"\D", "", str(rule.get("ddd") or ""))[:2]
        return normalize_text(rule.get("cidade")) == city and (not rule_ddd or rule_ddd == ddd)
    return False


def region_rule_sort_key(rule: dict):
    return (
        REGION_PRIORITY.get(str(rule.get("tipo") or "").lower(), 0),
        str(rule.get("_atualizado_em") or rule.get("_criado_em") or ""),
    )


def vendor_legacy_region_client_ids(vendor: dict, client_index: dict) -> set:
    assignments = vendor.get("clientes_atendidos") or {}
    assigned_ufs = {
        str(uf or "").strip().upper()
        for uf in assignments.get("ufs", [])
        if str(uf or "").strip().upper()
    }
    assigned_cities = {
        (str(item.get("uf") or "").strip().upper(), normalize_text(item.get("cidade"))): item
        for item in assignments.get("cidades", [])
        if item.get("cidade")
    }
    specific_client_ids = {
        normalize_identifier(item.get("id"))
        for item in assignments.get("clientes_especificos", [])
        if normalize_identifier(item.get("id"))
    }
    region_client_ids = set()
    for client_id, client in client_index.items():
        uf = str(client.get("UF") or "").strip().upper()
        city_key = normalize_text(client.get("CID"))
        ddd = client_ddd(client)
        city_assignment = assigned_cities.get((uf, city_key))
        if client_id in specific_client_ids:
            region_client_ids.add(client_id)
            continue
        if uf in assigned_ufs:
            region_client_ids.add(client_id)
            continue
        if city_assignment:
            selected_ddds = [str(value) for value in city_assignment.get("ddds", []) if str(value)]
            if not selected_ddds or not ddd or ddd in selected_ddds:
                region_client_ids.add(client_id)
    return region_client_ids


def clients_for_vendor_region(company_id: str, vendor_id_value: str, client_index: dict, vendor: dict) -> set:
    rules = active_region_rules(company_id)
    if not rules:
        return vendor_legacy_region_client_ids(vendor, client_index)
    region_client_ids = set()
    for client_id, client in client_index.items():
        matching = [rule for rule in rules if rule_matches_client(rule, client_id, client)]
        if not matching:
            continue
        selected_rule = sorted(matching, key=region_rule_sort_key)[-1]
        if selected_rule.get("vendor_id") == vendor_id_value and not str(selected_rule.get("tipo") or "").startswith("exclusao_"):
            region_client_ids.add(client_id)
    return region_client_ids


def vendor_regions_payload(company_id: str, vendor_id_value: str):
    vendor, client_index, region_client_ids = vendor_region_client_context(company_id, vendor_id_value)

    sales_by_client = {}
    historical_clients = set()
    years = list(range(START_YEAR, CURRENT_YEAR + 1))
    closed_month_count = max(1, date.today().month - 1) if CURRENT_YEAR == date.today().year else 12
    months = list(range(1, date.today().month + 1)) if CURRENT_YEAR == date.today().year else list(range(1, 13))
    for sale_company_id in vendor_sales_company_ids(company_id):
        for sale in load_json(sales_file(sale_company_id), []):
            if is_excluded_group_sale(sale_company_id, sale.get("CL_NOM")):
                continue
            client_id = normalize_identifier(sale.get("ID_CL"))
            if not client_id:
                continue
            historical_clients.add(client_id)
            year = record_year(sale.get("NF_EMI"))
            month = record_month(sale.get("NF_EMI"))
            net_revenue = sale_net_revenue(sale)
            entry = sales_by_client.setdefault(client_id, {
                "total": 0.0,
                "years": {str(year_value): 0.0 for year_value in years},
                "months": {str(month_value): 0.0 for month_value in months},
                "months_by_year": {
                    str(year_value): {str(month_value): 0.0 for month_value in range(1, 13)}
                    for year_value in years
                },
                "purchase_years": set(),
                "purchase_months": set(),
            })
            if year and year >= START_YEAR:
                entry["total"] += net_revenue
                entry["purchase_years"].add(year)
            if year in years:
                entry["years"][str(year)] += net_revenue
                if month:
                    entry["months_by_year"][str(year)][str(month)] += net_revenue
            if year == CURRENT_YEAR and month in months:
                entry["months"][str(month)] += net_revenue
                entry["purchase_months"].add(month)

    uf_rows = {}
    ddd_rows = {}
    city_rows = {}
    grouped_row = empty_region_summary("Agrupado", "agrupado", years, months)

    def include_client_in_summary(summary: dict, client_id: str):
        if client_id in summary["_client_ids"]:
            return
        summary["_client_ids"].add(client_id)
        summary["clientes"] += 1
        sale_summary = sales_by_client.get(client_id, {
            "total": 0.0,
            "years": {str(year_value): 0.0 for year_value in years},
            "months": {str(month_value): 0.0 for month_value in months},
            "months_by_year": {
                str(year_value): {str(month_value): 0.0 for month_value in range(1, 13)}
                for year_value in years
            },
            "purchase_years": set(),
            "purchase_months": set(),
        })
        purchase_years = sale_summary.get("purchase_years", set())
        purchase_months = sale_summary.get("purchase_months", set())
        if purchase_years:
            summary["ativos"] += 1
        elif client_id in historical_clients:
            summary["inativos"] += 1
        else:
            summary["nunca_comprou"] += 1
        for year_value in years:
            year_key = str(year_value)
            year_revenue = sale_summary["years"].get(year_key, 0.0)
            prior_year_purchase = any(prior_year < year_value for prior_year in purchase_years)
            summary["vendas_anos"][year_key] += year_revenue
            for month_value in range(1, 13):
                month_key = str(month_value)
                summary["vendas_meses_anos"][year_key][month_key] += sale_summary["months_by_year"].get(year_key, {}).get(month_key, 0.0)
            if year_value in purchase_years:
                summary["ativos_anos"][year_key] += 1
            elif prior_year_purchase:
                summary["inativos_anos"][year_key] += 1
            else:
                summary["nunca_anos"][year_key] += 1
        for month in months:
            month_key = str(month)
            month_revenue = sale_summary["months"].get(month_key, 0.0)
            prior_month_purchase = any(prior_year < CURRENT_YEAR for prior_year in purchase_years) or any(
                prior_month < month for prior_month in purchase_months
            )
            summary["vendas_meses"][month_key] += month_revenue
            if month in purchase_months:
                summary["ativos_meses"][month_key] += 1
            elif prior_month_purchase:
                summary["inativos_meses"][month_key] += 1
            else:
                summary["nunca_meses"][month_key] += 1

    for client_id in region_client_ids:
        client = client_index.get(client_id)
        if not client:
            continue
        uf = str(client.get("UF") or "").strip().upper() or "Sem UF"
        city = str(client.get("CID") or "").strip() or "Sem cidade"
        city_key = normalize_text(city)
        ddd = client_ddd(client) or "Sem DDD"

        uf_row = uf_rows.setdefault(uf, empty_region_summary(uf, uf, years, months))
        include_client_in_summary(uf_row, client_id)

        ddd_row = ddd_rows.setdefault(ddd, empty_region_summary(ddd, ddd, years, months))
        include_client_in_summary(ddd_row, client_id)

        city_row_key = f"{uf}|{city_key}"
        city_row = city_rows.setdefault(city_row_key, empty_region_summary(f"{city} - {uf}", city_row_key, years, months))
        include_client_in_summary(city_row, client_id)

        include_client_in_summary(grouped_row, client_id)

    return {
        "empresa": COMPANIES[company_id],
        "empresa_id": company_id,
        "vendedor": vendor_public_record(vendor),
        "ano_corrente": CURRENT_YEAR,
        "ano_inicial": START_YEAR,
        "anos": years,
        "meses": months,
        "meses_fechados": list(range(1, closed_month_count + 1)),
        "resumo": {
            "clientes": len(region_client_ids),
            "faturamento_total": round(sum(sales_by_client.get(client_id, {}).get("total", 0.0) for client_id in region_client_ids), 2),
        },
        "ufs": sorted((finalize_region_summary(row) for row in uf_rows.values()), key=lambda item: item["label"]),
        "ddds": sorted((finalize_region_summary(row) for row in ddd_rows.values()), key=lambda item: normalize_text(item["label"])),
        "cidades": sorted((finalize_region_summary(row) for row in city_rows.values()), key=lambda item: normalize_text(item["label"])),
        "agrupado": [finalize_region_summary(grouped_row)],
    }


def vendor_region_clients_payload(company_id: str, vendor_id_value: str, params):
    vendor, client_index, region_client_ids = vendor_region_client_context(company_id, vendor_id_value)
    selected_status = params.get("status", ["never"])[0]
    selected_uf = params.get("uf", [""])[0].strip().upper()
    selected_ddd = re.sub(r"\D", "", params.get("ddd", [""])[0])[:2]
    selected_city = normalize_text(params.get("city", [""])[0])
    query = str(params.get("q", [""])[0] or "").strip()

    sales_stats = {}
    for sale_company_id in vendor_sales_company_ids(company_id):
        for sale in load_json(sales_file(sale_company_id), []):
            if is_excluded_group_sale(sale_company_id, sale.get("CL_NOM")):
                continue
            client_id = normalize_identifier(sale.get("ID_CL"))
            if not client_id:
                continue
            sale_date = normalize_record(sale.get("NF_EMI"))
            net_revenue = sale_net_revenue(sale)
            entry = sales_stats.setdefault(client_id, {
                "historical": False,
                "current_year_purchases": 0,
                "current_year_revenue": 0.0,
                "last_purchase": None,
                "last_purchase_revenue": 0.0,
            })
            entry["historical"] = True
            if sale_date:
                if entry["last_purchase"] is None or str(sale_date) > str(entry["last_purchase"]):
                    entry["last_purchase"] = sale_date
                    entry["last_purchase_revenue"] = net_revenue
                elif str(sale_date) == str(entry["last_purchase"]):
                    entry["last_purchase_revenue"] += net_revenue
            if record_year(sale.get("NF_EMI")) == CURRENT_YEAR:
                entry["current_year_purchases"] += 1
                entry["current_year_revenue"] += net_revenue

    options = {"ufs": set(), "ddds": set(), "cities": {}}
    rows = []
    counts = {"active": 0, "inactive": 0, "never": 0}
    for client_id in region_client_ids:
        client = client_index.get(client_id)
        if not client:
            continue
        uf = str(client.get("UF") or "").strip().upper()
        city = str(client.get("CID") or "").strip()
        city_key = normalize_text(city)
        ddd = client_ddd(client)
        options["ufs"].add(uf)
        if ddd:
            options["ddds"].add(ddd)
        if uf or city:
            options["cities"][f"{uf}|{city_key}"] = {"uf": uf, "city": city, "label": f"{city} - {uf}".strip(" -")}

        stats = sales_stats.get(client_id, {})
        if stats.get("current_year_purchases", 0) > 0:
            status_key = "active"
            status_label = "Ativo"
        elif stats.get("historical"):
            status_key = "inactive"
            status_label = "Inativo"
        else:
            status_key = "never"
            status_label = "Nunca Comprou"
        counts[status_key] += 1

        if selected_status in ("active", "inactive", "never") and status_key != selected_status:
            continue
        if selected_uf and uf != selected_uf:
            continue
        if selected_ddd and ddd != selected_ddd:
            continue
        if selected_city and city_key != selected_city:
            continue
        if query and not search_text_matches(query, f"{client.get('ID')} {client.get('NOM')} {client.get('FAN')} {uf} {city} {ddd}"):
            continue

        revenue = stats.get("current_year_revenue", 0.0) if status_key == "active" else stats.get("last_purchase_revenue", 0.0)
        rows.append({
            "ID": client.get("ID"),
            "NOM": client.get("NOM"),
            "UF": uf,
            "CID": city,
            "DDD": ddd,
            "status": status_label,
            "ultima_compra": stats.get("last_purchase") or "Nunca Comprou",
            "faturamento_liquido": round(revenue, 2),
        })

    def vendor_client_sort_key(row: dict):
        if row["status"] == "Nunca Comprou":
            return (2, normalize_text(row.get("NOM")), row["UF"], normalize_text(row["CID"]))
        last_purchase = str(row.get("ultima_compra") or "")
        return (0 if row["status"] == "Ativo" else 1, "" if not last_purchase else "".join(chr(255 - ord(char)) for char in last_purchase), normalize_text(row.get("NOM")))

    rows.sort(key=vendor_client_sort_key)
    return {
        "empresa": COMPANIES[company_id],
        "empresa_id": company_id,
        "vendedor": vendor_public_record(vendor),
        "ano_corrente": CURRENT_YEAR,
        "filters": {
            "status": selected_status,
            "uf": selected_uf,
            "ddd": selected_ddd,
            "city": params.get("city", [""])[0],
        },
        "counts": counts,
        "total_carteira": len(region_client_ids),
        "total_resultado": len(rows),
        "columns": [
            {"key": "ID", "label": "Codigo"},
            {"key": "NOM", "label": "Cliente"},
            {"key": "UF", "label": "UF"},
            {"key": "CID", "label": "Cidade"},
            {"key": "DDD", "label": "DDD"},
            {"key": "status", "label": "Status"},
            {"key": "ultima_compra", "label": "Ultima compra"},
            {"key": "faturamento_liquido", "label": "Faturamento liquido"},
        ],
        "options": {
            "ufs": sorted(uf for uf in options["ufs"] if uf),
            "ddds": sorted(ddd for ddd in options["ddds"] if ddd),
            "cities": sorted(options["cities"].values(), key=lambda item: (item["uf"], normalize_text(item["city"]))),
        },
        "rows": rows[:1000],
        "limit": 1000,
    }


def display_date_br(value) -> str:
    parsed = record_date_value(value)
    if not parsed:
        return ""
    return parsed.strftime("%d/%m/%Y")


def day_by_day_date_value(day_key: str = "") -> date:
    parsed = record_date_value(day_key) if day_key else None
    return parsed or date.today()


def is_day_by_day_business_day(day_value: date) -> bool:
    return day_value.weekday() < 5


def next_day_by_day_business_day(day_value: date) -> date:
    next_day = day_value
    while not is_day_by_day_business_day(next_day):
        next_day += timedelta(days=1)
    return next_day


def day_by_day_weekend_payload(company_id: str, vendor_id_value: str, day_value: date) -> dict:
    next_business = next_day_by_day_business_day(day_value)
    vendor = vendor_record_by_id(company_id, vendor_id_value)
    return {
        "empresa": COMPANIES[company_id],
        "empresa_id": company_id,
        "vendedor": vendor,
        "data": day_value.isoformat(),
        "data_label": display_date_br(day_value),
        "dia_util": False,
        "proximo_dia_util": next_business.isoformat(),
        "proximo_dia_util_label": display_date_br(next_business),
        "mensagem": "O Day by Day considera somente dias úteis, de segunda-feira a sexta-feira.",
        "resumo": {
            "inativos": 0,
            "nunca_comprou": 0,
            "ja_contatados": 0,
            "total": 0,
            "candidatos_inativos": 0,
            "candidatos_nunca_comprou": 0,
            "candidatos_ja_contatados": 0,
            "ufs_consideradas": [],
        },
        "contagem": {
            "data": day_value.isoformat(),
            "meta_diaria": 0,
            "salvos": 0,
            "faltam": 0,
            "atingiu_meta": True,
            "fechamento_diario": "Sem rotina aos sabados e domingos",
        },
        "inativos": [],
        "nunca_comprou": [],
        "recontatos": [],
        "contatos_anteriores": [],
        "agrp_options": [],
    }


def vendor_client_sales_stats(company_id: str) -> dict:
    sales_stats = {}
    for sale_company_id in vendor_sales_company_ids(company_id):
        for sale in load_json(sales_file(sale_company_id), []):
            if is_excluded_group_sale(sale_company_id, sale.get("CL_NOM")):
                continue
            client_id = normalize_identifier(sale.get("ID_CL"))
            if not client_id:
                continue
            sale_date = normalize_record(sale.get("NF_EMI"))
            sale_year = record_year(sale.get("NF_EMI"))
            net_revenue = sale_net_revenue(sale)
            entry = sales_stats.setdefault(client_id, {
                "historical": False,
                "current_year_purchases": 0,
                "current_year_revenue": 0.0,
                "last_purchase": None,
                "last_purchase_revenue": 0.0,
                "purchase_notes": {},
            })
            entry["historical"] = True
            note_id = normalize_identifier(sale.get("ID_NF") or sale.get("NF_NUM") or f"{sale_date}|{sale.get('ID_PRNF')}")
            if note_id:
                note = entry["purchase_notes"].setdefault(note_id, {
                    "id_nf": sale.get("ID_NF") or "",
                    "nf_num": sale.get("NF_NUM") or "",
                    "data": sale_date,
                    "ano": sale_year,
                    "faturamento_liquido": 0.0,
                    "itens": 0,
                    "itens_detalhe": [],
                })
                note["faturamento_liquido"] += net_revenue
                note["itens"] += 1
                note.setdefault("itens_detalhe", []).append({
                    "referencia": normalize_record(sale.get("PR_COD")),
                    "descricao": normalize_record(sale.get("PR_DES")),
                    "quantidade": round(number_value(sale.get("PR_QTD")), 3),
                    "faturamento_liquido": round(net_revenue, 2),
                    "agrp": normalize_record(sale.get("AGRP")),
                    "origem": normalize_record(sale.get("Origem")),
                })
                if sale_date and (not note.get("data") or str(sale_date) > str(note.get("data"))):
                    note["data"] = sale_date
                if sale_year:
                    note["ano"] = sale_year
            if sale_date:
                if entry["last_purchase"] is None or str(sale_date) > str(entry["last_purchase"]):
                    entry["last_purchase"] = sale_date
                    entry["last_purchase_revenue"] = net_revenue
                elif str(sale_date) == str(entry["last_purchase"]):
                    entry["last_purchase_revenue"] += net_revenue
            if sale_year == CURRENT_YEAR:
                entry["current_year_purchases"] += 1
                entry["current_year_revenue"] += net_revenue
    for entry in sales_stats.values():
        notes = list((entry.get("purchase_notes") or {}).values())
        yearly_counts = {str(year): 0 for year in range(START_YEAR, CURRENT_YEAR + 1)}
        for note in notes:
            year = note.get("ano")
            if year and START_YEAR <= year <= CURRENT_YEAR:
                yearly_counts[str(year)] += 1
        notes.sort(key=lambda note: str(note.get("data") or ""), reverse=True)
        entry["purchase_counts_by_year"] = yearly_counts
        purchases_by_year = {str(year): [] for year in range(START_YEAR, CURRENT_YEAR + 1)}
        for note in notes:
            year_key = str(note.get("ano") or "")
            if year_key not in purchases_by_year:
                continue
            purchases_by_year[year_key].append({
                "data": display_date_br(note.get("data")),
                "data_iso": str(note.get("data") or ""),
                "nf_num": note.get("nf_num") or note.get("id_nf") or "",
                "faturamento_liquido": round(note.get("faturamento_liquido") or 0, 2),
                "itens": note.get("itens") or 0,
                "itens_detalhe": note.get("itens_detalhe") or [],
            })
        entry["purchases_by_year"] = purchases_by_year
    return sales_stats


def vendor_day_by_day_client_row(client_id: str, client: dict, stats: dict, status_key: str) -> dict:
    last_purchase = stats.get("last_purchase") or ""
    revenue = stats.get("last_purchase_revenue", 0.0) if status_key == "inactive" else 0.0
    return {
        "id": client_id,
        "codigo": client.get("ID") or client_id,
        "nome": client.get("NOM") or "",
        "documento": client.get("CGC") or client.get("CNPJ") or client.get("Documento") or "",
        "uf": str(client.get("UF") or "").strip().upper(),
        "cidade": str(client.get("CID") or "").strip(),
        "ddd": client_ddd(client),
        "telefone": client.get("TEL") or client.get("Telefone") or "",
        "email": client.get("EMAIL") or client.get("Email") or "",
        "status": "Inativo" if status_key == "inactive" else ("Nunca Comprou" if status_key == "never" else "Ja Contatado"),
        "ultima_compra": display_date_br(last_purchase) if last_purchase else "Nunca Comprou",
        "ultima_compra_iso": str(last_purchase or ""),
        "faturamento_liquido": round(revenue, 2),
        "compras_por_ano": stats.get("purchase_counts_by_year") or {str(year): 0 for year in range(START_YEAR, CURRENT_YEAR + 1)},
        "compras_ano_detalhe": stats.get("purchases_by_year") or {str(year): [] for year in range(START_YEAR, CURRENT_YEAR + 1)},
    }


def vendor_day_by_day_recontact_candidates(company_id: str, vendor_id_value: str, data: dict) -> list[dict]:
    _vendor, client_index, region_client_ids = vendor_region_client_context(company_id, vendor_id_value)
    sales_stats = vendor_client_sales_stats(company_id)
    blocked_clients = blocked_client_map(company_id)
    oldest_by_client = {}
    for record in data.get("historico") or []:
        if str(record.get("vendor_id") or "") != vendor_id_value:
            continue
        client_id = normalize_identifier(record.get("client_id"))
        if not client_id or client_id in blocked_clients or client_id not in region_client_ids:
            continue
        record_sort = (str(record.get("data") or ""), str(record.get("saved_at") or ""))
        current = oldest_by_client.get(client_id)
        current_sort = (str((current or {}).get("data") or ""), str((current or {}).get("saved_at") or ""))
        if not current or record_sort < current_sort:
            oldest_by_client[client_id] = record

    rows = []
    for client_id, record in oldest_by_client.items():
        client = client_index.get(client_id) or (record.get("cliente") or {})
        if not client:
            continue
        stats = sales_stats.get(client_id, {})
        row = vendor_day_by_day_client_row(client_id, client, stats, "recontact")
        form = record.get("form") or {}
        row.update({
            "primeiro_contato": record.get("data") or "",
            "primeiro_contato_label": display_date_br(record.get("data")),
            "primeiro_contato_salvo_em": record.get("saved_at") or "",
            "resumo_contato": record.get("resumo") or vendor_day_by_day_contact_summary(form),
            "probabilidade_compra_futura": form.get("probabilidade_compra_futura") or "",
            "data_agendamento_contato": form.get("data_agendamento_contato") or "",
        })
        rows.append(row)
    rows.sort(key=lambda item: (item.get("primeiro_contato") or "", item.get("primeiro_contato_salvo_em") or ""))
    return rows


def vendor_day_by_day_candidates(company_id: str, vendor_id_value: str):
    vendor, client_index, region_client_ids = vendor_region_client_context(company_id, vendor_id_value)
    sales_stats = vendor_client_sales_stats(company_id)
    blocked_clients = blocked_client_map(company_id)
    candidates = {"inactive": [], "never": []}
    for client_id in region_client_ids:
        if client_id in blocked_clients:
            continue
        client = client_index.get(client_id)
        if not client:
            continue
        stats = sales_stats.get(client_id, {})
        if stats.get("current_year_purchases", 0) > 0:
            continue
        status_key = "inactive" if stats.get("historical") else "never"
        candidates[status_key].append(vendor_day_by_day_client_row(client_id, client, stats, status_key))
    candidates["inactive"].sort(key=lambda row: normalize_text(row.get("nome")))
    candidates["inactive"].sort(key=lambda row: row.get("ultima_compra_iso") or "", reverse=True)
    candidates["never"].sort(key=lambda row: normalize_text(row.get("nome")))
    return vendor, candidates


def vendor_record_by_id(company_id: str, vendor_id_value: str) -> dict:
    return cached_payload(
        ("vendor_record_by_id", company_id, vendor_id_value),
        [vendors_file(company_id), sales_file(company_id)],
        lambda: build_vendor_record_by_id(company_id, vendor_id_value),
    )


def build_vendor_record_by_id(company_id: str, vendor_id_value: str) -> dict:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    vendors = load_json(vendors_file(company_id), [])
    vendor = next((item for item in vendors if item.get("id") == vendor_id_value), None)
    if not vendor:
        vendors, _inserted = sync_vendors_from_sales(company_id)
        vendor = next((item for item in vendors if item.get("id") == vendor_id_value), None)
    if not vendor:
        raise ValueError("Vendedor nao encontrado.")
    public = vendor_public_record(vendor)
    return {
        key: public.get(key)
        for key in (
            "id",
            "nome_completo",
            "tipo_usuario",
            "telefone",
            "whatsapp",
            "email",
            "status",
            "empresas_acesso",
            "empresa_padrao",
            "pagina_vendedor",
        )
    }


def stable_daily_sample(rows: list[dict], amount: int, company_id: str, vendor_id_value: str, day_key: str, status_key: str) -> list[str]:
    ordered = sorted(
        rows,
        key=lambda row: hashlib.sha256(f"{company_id}|{vendor_id_value}|{day_key}|{status_key}|{row.get('id')}".encode("utf-8")).hexdigest(),
    )
    return [row.get("id") for row in ordered[:amount] if row.get("id")]


def vendor_agrp_options(company_id: str) -> list[str]:
    dependencies = []
    for current_company_id in vendor_sales_company_ids(company_id):
        dependencies.extend([costing_file(current_company_id), costing_fabricated_file(current_company_id)])
    return cached_payload(
        ("vendor_agrp_options", company_id),
        dependencies,
        lambda: build_vendor_agrp_options(company_id),
    )


def build_vendor_agrp_options(company_id: str) -> list[str]:
    values = set()
    for current_company_id in vendor_sales_company_ids(company_id):
        for path in (costing_file(current_company_id), costing_fabricated_file(current_company_id)):
            for record in load_json(path, []):
                agrp = normalize_record(record.get("AGRP"))
                if agrp and not missing_product_classification(agrp):
                    values.add(agrp)
    return sorted(values, key=normalize_text)


def vendor_day_by_day_count_summary(data: dict, vendor_id_value: str, day_key: str) -> dict:
    counter = ((data.get("contagens") or {}).get(f"{day_key}|{vendor_id_value}") or {})
    events = counter.get("eventos") if isinstance(counter.get("eventos"), list) else []
    total = len(events)
    return {
        "data": day_key,
        "meta_diaria": DAY_BY_DAY_DAILY_TARGET,
        "salvos": total,
        "faltam": max(0, DAY_BY_DAY_DAILY_TARGET - total),
        "atingiu_meta": total >= DAY_BY_DAY_DAILY_TARGET,
        "fechamento_diario": "Fechado por data do atendimento",
    }


def vendor_day_by_day_monthly_count(company_id: str, vendor_id_value: str, year: int, month: int) -> int:
    data = load_json(vendor_day_by_day_file(company_id), {"listas": {}, "atendimentos": {}, "contagens": {}})
    total = 0
    for key, counter in (data.get("contagens") or {}).items():
        day_key = str((counter or {}).get("data") or key.split("|", 1)[0])
        if str((counter or {}).get("vendor_id") or "") != vendor_id_value:
            continue
        parsed = record_date_value(day_key)
        if parsed and not is_day_by_day_business_day(parsed):
            continue
        if parsed and parsed.year == year and parsed.month == month:
            events = counter.get("eventos") if isinstance(counter.get("eventos"), list) else []
            total += len(events)
    return total


def vendor_day_by_day_payload(company_id: str, vendor_id_value: str, day_key: str = ""):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    day_value = day_by_day_date_value(day_key)
    if not is_day_by_day_business_day(day_value):
        return day_by_day_weekend_payload(company_id, vendor_id_value, day_value)
    day_key = day_value.isoformat()
    data = load_json(vendor_day_by_day_file(company_id), {"listas": {}, "atendimentos": {}})
    blocked_clients = blocked_client_map(company_id)
    list_key = f"{day_key}|{vendor_id_value}"
    current_list = data.setdefault("listas", {}).get(list_key) or {}
    saved_details = current_list.get("details") or {}
    saved_schema_is_current = all(
        "compras_por_ano" in item and "compras_ano_detalhe" in item
        for rows in saved_details.values()
        for item in (rows or [])
    )
    if saved_details and saved_schema_is_current:
        records = data.get("atendimentos", {})

        def saved_rows_for(status_key: str) -> list[dict]:
            result = []
            for row in saved_details.get(status_key, []) or []:
                client_id = row.get("id")
                if normalize_identifier(client_id) in blocked_clients:
                    continue
                record_key = f"{day_key}|{vendor_id_value}|{client_id}"
                result.append({**row, "atendimento_salvo": bool(records.get(record_key))})
            return result

        inactive_rows = saved_rows_for("inactive")
        never_rows = saved_rows_for("never")
        recontact_rows = saved_rows_for("recontact")
        count_summary = vendor_day_by_day_count_summary(data, vendor_id_value, day_key)
        candidate_counts = current_list.get("candidate_counts") or {}
        if not inactive_rows and not never_rows and not recontact_rows and not candidate_counts.get("inactive") and not candidate_counts.get("never"):
            recontact_rows = vendor_day_by_day_recontact_candidates(company_id, vendor_id_value, data)[:DAY_BY_DAY_DAILY_TARGET]
        selected_ufs = sorted({row.get("uf") for row in inactive_rows + never_rows + recontact_rows if row.get("uf")})
        vendor = vendor_record_by_id(company_id, vendor_id_value)
        return {
            "empresa": COMPANIES[company_id],
            "empresa_id": company_id,
            "vendedor": vendor,
            "data": day_key,
            "data_label": display_date_br(day_key),
            "resumo": {
                "inativos": len(inactive_rows),
                "nunca_comprou": len(never_rows),
                "ja_contatados": len(recontact_rows),
                "total": len(inactive_rows) + len(never_rows) + len(recontact_rows),
                "candidatos_inativos": candidate_counts.get("inactive", len(inactive_rows)),
                "candidatos_nunca_comprou": candidate_counts.get("never", len(never_rows)),
                "candidatos_ja_contatados": candidate_counts.get("recontact", len(recontact_rows)),
                "ufs_consideradas": selected_ufs,
            },
            "contagem": count_summary,
            "inativos": inactive_rows,
            "nunca_comprou": never_rows,
            "recontatos": recontact_rows,
            "contatos_anteriores": vendor_day_by_day_previous_contacts(data, vendor_id_value, day_key) if count_summary.get("faltam") == 0 else [],
            "agrp_options": vendor_agrp_options(company_id),
        }

    vendor, candidates = vendor_day_by_day_candidates(company_id, vendor_id_value)
    changed = False

    selected = {}
    targets = {"inactive": 35, "never": 15}
    recontact_candidates = []
    if not candidates.get("inactive") and not candidates.get("never"):
        recontact_candidates = vendor_day_by_day_recontact_candidates(company_id, vendor_id_value, data)
        candidates["recontact"] = recontact_candidates
        targets = {"recontact": DAY_BY_DAY_DAILY_TARGET}
    candidate_by_status = {
        status: {row.get("id"): row for row in rows}
        for status, rows in candidates.items()
    }
    for status_key, amount in targets.items():
        if status_key == "inactive":
            saved_ids = [row.get("id") for row in candidates.get(status_key, [])[:amount] if row.get("id")]
        elif status_key == "recontact":
            saved_ids = [row.get("id") for row in candidates.get(status_key, [])[:amount] if row.get("id")]
        else:
            saved_ids = [
                client_id for client_id in current_list.get(status_key, [])
                if client_id in candidate_by_status.get(status_key, {})
            ]
        if status_key != "inactive" and len(saved_ids) < min(amount, len(candidate_by_status.get(status_key, {}))):
            extra_ids = stable_daily_sample(
                [row for row in candidates.get(status_key, []) if row.get("id") not in set(saved_ids)],
                amount - len(saved_ids),
                company_id,
                vendor_id_value,
                day_key,
                status_key,
            )
            saved_ids.extend(extra_ids)
        saved_ids = saved_ids[:amount]
        selected[status_key] = saved_ids
        if current_list.get(status_key) != saved_ids:
            changed = True

    selected_details = {
        status_key: [
            candidate_by_status.get(status_key, {}).get(client_id)
            for client_id in selected.get(status_key, [])
            if candidate_by_status.get(status_key, {}).get(client_id)
        ]
        for status_key in targets
    }

    save_warning = ""
    if changed or list_key not in data.get("listas", {}) or not current_list.get("details") or not saved_schema_is_current:
        data.setdefault("listas", {})[list_key] = {
            "data": day_key,
            "vendor_id": vendor_id_value,
            "inactive": selected.get("inactive", []),
            "never": selected.get("never", []),
            "recontact": selected.get("recontact", []),
            "inactive_strategy": "recent_purchase_first",
            "details": selected_details,
            "candidate_counts": {
                "inactive": len(candidates["inactive"]),
                "never": len(candidates["never"]),
                "recontact": len(candidates.get("recontact", [])),
            },
            "created_at": current_list.get("created_at") or datetime.now().isoformat(timespec="seconds"),
            "updated_at": datetime.now().isoformat(timespec="seconds"),
        }
        try:
            save_json(vendor_day_by_day_file(company_id), data)
        except PermissionError:
            save_warning = (
                "A fila foi exibida, mas o Windows nao permitiu gravar o arquivo do Day by Day neste momento. "
                "Feche outras janelas do CRM ou backups usando o arquivo e tente atualizar novamente."
            )

    records = data.get("atendimentos", {})

    def rows_for(status_key: str) -> list[dict]:
        result = []
        for client_id in selected.get(status_key, []):
            row = candidate_by_status.get(status_key, {}).get(client_id)
            if not row:
                continue
            record_key = f"{day_key}|{vendor_id_value}|{client_id}"
            result.append({**row, "atendimento_salvo": bool(records.get(record_key))})
        return result

    inactive_rows = rows_for("inactive")
    never_rows = rows_for("never")
    recontact_rows = rows_for("recontact")
    selected_ufs = sorted({row.get("uf") for row in inactive_rows + never_rows + recontact_rows if row.get("uf")})
    count_summary = vendor_day_by_day_count_summary(data, vendor_id_value, day_key)

    return {
        "empresa": COMPANIES[company_id],
        "empresa_id": company_id,
        "vendedor": vendor_public_record(vendor),
        "data": day_key,
        "data_label": display_date_br(day_key),
        "resumo": {
            "inativos": len(selected.get("inactive", [])),
            "nunca_comprou": len(selected.get("never", [])),
            "ja_contatados": len(selected.get("recontact", [])),
            "total": len(selected.get("inactive", [])) + len(selected.get("never", [])) + len(selected.get("recontact", [])),
            "candidatos_inativos": len(candidates["inactive"]),
            "candidatos_nunca_comprou": len(candidates["never"]),
            "candidatos_ja_contatados": len(candidates.get("recontact", [])),
            "ufs_consideradas": selected_ufs,
        },
        "contagem": count_summary,
        "inativos": inactive_rows,
        "nunca_comprou": never_rows,
        "recontatos": recontact_rows,
        "contatos_anteriores": vendor_day_by_day_previous_contacts(data, vendor_id_value, day_key) if count_summary.get("faltam") == 0 else [],
        "agrp_options": vendor_agrp_options(company_id),
        "aviso_gravacao": save_warning,
    }


def vendor_day_by_day_previous_contacts(data: dict, vendor_id_value: str, day_key: str) -> list[dict]:
    latest_by_client = {}
    for record in data.get("historico") or []:
        if str(record.get("vendor_id") or "") != vendor_id_value:
            continue
        if str(record.get("data") or "") == day_key:
            continue
        client_id = normalize_identifier(record.get("client_id"))
        if not client_id:
            continue
        current = latest_by_client.get(client_id)
        record_sort = (str(record.get("data") or ""), str(record.get("saved_at") or ""))
        current_sort = (str((current or {}).get("data") or ""), str((current or {}).get("saved_at") or ""))
        if not current or record_sort > current_sort:
            form = record.get("form") or {}
            latest_by_client[client_id] = {
                "data": record.get("data") or "",
                "data_label": display_date_br(record.get("data")),
                "saved_at": record.get("saved_at") or "",
                "client_id": client_id,
                "cliente_codigo": record.get("cliente_codigo") or client_id,
                "cliente_nome": record.get("cliente_nome") or "",
                "status": record.get("status_label") or record.get("status_key") or "",
                "resumo": record.get("resumo") or vendor_day_by_day_contact_summary(form),
                "probabilidade_compra_futura": form.get("probabilidade_compra_futura") or "",
                "data_agendamento_contato": form.get("data_agendamento_contato") or "",
            }
    rows = list(latest_by_client.values())
    rows.sort(key=lambda item: (item.get("data") or "", item.get("saved_at") or ""), reverse=True)
    return rows[:50]


def vendor_day_by_day_client_payload(company_id: str, vendor_id_value: str, client_id: str, day_key: str = ""):
    day_value = day_by_day_date_value(day_key)
    if not is_day_by_day_business_day(day_value):
        raise ValueError("O Day by Day considera somente dias uteis, de segunda-feira a sexta-feira.")
    day_key = day_value.isoformat()
    if normalize_identifier(client_id) in blocked_client_map(company_id):
        raise ValueError("Cliente bloqueado. Ele nao deve receber novas tratativas no Day by Day.")
    row = None
    status_key = ""
    data = load_json(vendor_day_by_day_file(company_id), {"listas": {}, "atendimentos": {}})
    daily_list = (data.get("listas") or {}).get(f"{day_key}|{vendor_id_value}") or {}
    for key in ("inactive", "never", "recontact"):
        row = next((item for item in (daily_list.get("details") or {}).get(key, []) if normalize_identifier(item.get("id")) == client_id), None)
        if row:
            status_key = key
            break
    if not row:
        _vendor, candidates = vendor_day_by_day_candidates(company_id, vendor_id_value)
        if not candidates.get("inactive") and not candidates.get("never"):
            candidates["recontact"] = vendor_day_by_day_recontact_candidates(company_id, vendor_id_value, data)
        for key in ("inactive", "never", "recontact"):
            row = next((item for item in candidates.get(key, []) if item.get("id") == client_id), None)
            if row:
                status_key = key
                break
    if not row:
        raise ValueError("Cliente nao encontrado na carteira do vendedor.")
    record_key = f"{day_key}|{vendor_id_value}|{client_id}"
    return {
        "empresa": COMPANIES[company_id],
        "empresa_id": company_id,
        "data": day_key,
        "data_label": display_date_br(day_key),
        "status_key": status_key,
        "cliente": row,
        "form": (data.get("atendimentos") or {}).get(record_key, {}).get("form", {}),
        "historico": vendor_day_by_day_history_rows(data, vendor_id_value, client_id),
        "agrp_options": vendor_agrp_options(company_id),
    }


def vendor_day_by_day_contact_summary(form: dict) -> str:
    if day_by_day_is_no(form.get("cliente_revenda")):
        return "Cliente informado como nao revenda."
    parts = []
    if form.get("nome_contato"):
        parts.append(f"Contato: {form.get('nome_contato')}")
    if form.get("motivo_sem_compra"):
        parts.append(f"Motivo sem compra: {form.get('motivo_sem_compra')}")
    if form.get("descritivo_problema_relatado"):
        parts.append(f"Problema: {form.get('descritivo_problema_relatado')}")
    if form.get("interesse_agrp"):
        parts.append("Interesse: " + ", ".join(form.get("interesse_agrp") or []))
    if form.get("descritivo_novo_contato"):
        parts.append(f"Proximo contato: {form.get('descritivo_novo_contato')}")
    if form.get("observacoes"):
        parts.append(f"Obs.: {form.get('observacoes')}")
    return " | ".join(parts) or "Atendimento salvo."


def day_by_day_is_yes(value) -> bool:
    return normalize_text(value) == "SIM"


def day_by_day_is_no(value) -> bool:
    return normalize_text(value) in {"NAO", "NÃO", "NO"}


def vendor_day_by_day_history_rows(data: dict, vendor_id_value: str, client_id: str) -> list[dict]:
    rows = []
    for record in (data.get("historico") or []):
        if str(record.get("vendor_id") or "") != vendor_id_value:
            continue
        if normalize_identifier(record.get("client_id")) != client_id:
            continue
        form = record.get("form") or {}
        rows.append({
            "data": record.get("data") or "",
            "data_label": display_date_br(record.get("data")),
            "saved_at": record.get("saved_at") or record.get("updated_at") or "",
            "status": record.get("status_label") or record.get("status_key") or "",
            "cliente_nome": record.get("cliente_nome") or "",
            "resumo": record.get("resumo") or vendor_day_by_day_contact_summary(form),
            "probabilidade_compra_futura": form.get("probabilidade_compra_futura") or "",
            "data_agendamento_contato": form.get("data_agendamento_contato") or "",
            "descritivo_novo_contato": form.get("descritivo_novo_contato") or "",
            "form": form,
        })
    rows.sort(key=lambda item: (item.get("data") or "", item.get("saved_at") or ""), reverse=True)
    return rows


def validate_vendor_day_by_day_form(form: dict, status_key: str = ""):
    if not normalize_record(form.get("cliente_revenda")):
        raise ValueError("Preencha todos os campos obrigatorios antes de salvar: Cliente Revenda.")
    if day_by_day_is_no(form.get("cliente_revenda")):
        return
    required = [
        ("cliente_ainda_operacao", "Cliente ainda em operacao"),
        ("nome_contato", "Nome do Contato"),
        ("telefone_ddd", "Telefone com DDD"),
        ("whatsapp", "WhatsApp"),
        ("emails_contato", "E-mails de contato"),
        ("vende_licitacao", "Cliente vende apenas em Licitacao"),
        ("tem_ecommerce", "Cliente tem ecommerce"),
        ("tem_site", "Cliente tem site"),
        ("produtos_site_ecommerce", "Cliente tem nossos produtos no site ou ecommerce"),
        ("relatou_problema_antes", "Cliente relatou algum problema ocorrido antes"),
        ("pertence_grupo_economico", "Cliente pertence a um grupo Economico"),
    ]
    if status_key == "inactive":
        required.append(("motivo_sem_compra", "Motivo de nao ter mais compra"))
    if day_by_day_is_no(form.get("cliente_ainda_operacao")):
        required.extend([
            ("empresa_baixada_cnpj", "Empresa baixada no CNPJ"),
            ("cliente_mudou_ramo", "Cliente mudou ramo de atividade"),
            ("observacoes", "Observacoes"),
        ])
    if day_by_day_is_yes(form.get("relatou_problema_antes")):
        required.append(("descritivo_problema_relatado", "Descritivo resumido do problema relatado"))
    required.extend([
        ("probabilidade_compra_futura", "Probabilidade de Compra Futura"),
        ("quer_agendar_contato", "Quer agendar uma data para entrar em contato?"),
    ])
    if day_by_day_is_yes(form.get("quer_agendar_contato")):
        required.extend([
            ("data_agendamento_contato", "Data do novo contato"),
            ("descritivo_novo_contato", "O que sera tratado no novo contato"),
        ])
        scheduled_date = parse_br_date(form.get("data_agendamento_contato"))
        if not scheduled_date:
            raise ValueError("Informe a data do novo contato no formato dd/mm/aaaa.")
        if scheduled_date <= date.today():
            raise ValueError("A data do novo contato precisa ser uma data futura.")
    missing = [label for key, label in required if not normalize_record(form.get(key))]
    if not form.get("interesse_agrp"):
        missing.append("Cliente relatou interesse em algum produto")
    if day_by_day_is_yes(form.get("pertence_grupo_economico")) and not normalize_record(form.get("cliente_matriz_grupo_id")):
        missing.append("Codigo do Cliente Matriz do grupo")
    if missing:
        raise ValueError("Preencha todos os campos obrigatorios antes de salvar: " + ", ".join(missing) + ".")


def save_vendor_day_by_day_client_payload(payload: dict):
    company_id = str(payload.get("company") or "ionlab")
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    vendor_id_value = str(payload.get("vendor_id") or "").strip()
    client_id = normalize_identifier(payload.get("client_id"))
    if not vendor_id_value or not client_id:
        raise ValueError("Informe vendedor e cliente.")
    day_value = day_by_day_date_value(str(payload.get("date") or "")[:10])
    if not is_day_by_day_business_day(day_value):
        raise ValueError("O Day by Day considera somente dias uteis, de segunda-feira a sexta-feira.")
    day_key = day_value.isoformat()
    client_payload = vendor_day_by_day_client_payload(company_id, vendor_id_value, client_id, day_key)
    raw_form = payload.get("form") or {}
    agrps = raw_form.get("interesse_agrp") or []
    if isinstance(agrps, str):
        agrps = [agrps]
    extra_contacts = []
    for contact in raw_form.get("contatos_adicionais") or []:
        if not isinstance(contact, dict):
            continue
        clean_contact = {
            "nome": normalize_record(contact.get("nome")),
            "telefone_ddd": normalize_record(contact.get("telefone_ddd")),
            "whatsapp": normalize_record(contact.get("whatsapp")),
            "emails_contato": normalize_record(contact.get("emails_contato")),
        }
        if any(clean_contact.values()):
            extra_contacts.append(clean_contact)
    form = {
        "cliente_revenda": normalize_record(raw_form.get("cliente_revenda")),
        "cliente_ainda_operacao": normalize_record(raw_form.get("cliente_ainda_operacao")),
        "empresa_baixada_cnpj": normalize_record(raw_form.get("empresa_baixada_cnpj")),
        "cliente_mudou_ramo": normalize_record(raw_form.get("cliente_mudou_ramo")),
        "observacoes": normalize_record(raw_form.get("observacoes")),
        "nome_contato": normalize_record(raw_form.get("nome_contato")),
        "telefone_ddd": normalize_record(raw_form.get("telefone_ddd")),
        "whatsapp": normalize_record(raw_form.get("whatsapp")),
        "emails_contato": normalize_record(raw_form.get("emails_contato")),
        "vende_licitacao": normalize_record(raw_form.get("vende_licitacao")),
        "motivo_sem_compra": normalize_record(raw_form.get("motivo_sem_compra")),
        "tem_ecommerce": normalize_record(raw_form.get("tem_ecommerce")),
        "tem_site": normalize_record(raw_form.get("tem_site")),
        "produtos_site_ecommerce": normalize_record(raw_form.get("produtos_site_ecommerce")),
        "relatou_problema_antes": normalize_record(raw_form.get("relatou_problema_antes")),
        "descritivo_problema_relatado": normalize_record(raw_form.get("descritivo_problema_relatado")),
        "pertence_grupo_economico": normalize_record(raw_form.get("pertence_grupo_economico")),
        "cliente_matriz_grupo_id": normalize_identifier(raw_form.get("cliente_matriz_grupo_id")),
        "cliente_matriz_grupo_nome": normalize_record(raw_form.get("cliente_matriz_grupo_nome")),
        "interesse_agrp": [normalize_record(item) for item in agrps if normalize_record(item)],
        "probabilidade_compra_futura": normalize_record(raw_form.get("probabilidade_compra_futura")),
        "quer_agendar_contato": normalize_record(raw_form.get("quer_agendar_contato")),
        "data_agendamento_contato": normalize_record(raw_form.get("data_agendamento_contato")),
        "descritivo_novo_contato": normalize_record(raw_form.get("descritivo_novo_contato")),
        "contatos_adicionais": extra_contacts,
    }
    validate_vendor_day_by_day_form(form, client_payload.get("status_key") or "")
    data = load_json(vendor_day_by_day_file(company_id), {"listas": {}, "atendimentos": {}, "contagens": {}})
    record_key = f"{day_key}|{vendor_id_value}|{client_id}"
    now = datetime.now().isoformat(timespec="seconds")
    data.setdefault("atendimentos", {})[record_key] = {
        "data": day_key,
        "vendor_id": vendor_id_value,
        "client_id": client_id,
        "status_key": client_payload.get("status_key"),
        "cliente": client_payload.get("cliente"),
        "form": form,
        "updated_at": now,
    }
    data.setdefault("historico", []).append({
        "data": day_key,
        "vendor_id": vendor_id_value,
        "client_id": client_id,
        "status_key": client_payload.get("status_key"),
        "status_label": (client_payload.get("cliente") or {}).get("status") or "",
        "cliente_nome": (client_payload.get("cliente") or {}).get("nome") or "",
        "cliente_codigo": (client_payload.get("cliente") or {}).get("codigo") or client_id,
        "cliente": client_payload.get("cliente"),
        "form": form,
        "resumo": vendor_day_by_day_contact_summary(form),
        "saved_at": now,
    })
    counter_key = f"{day_key}|{vendor_id_value}"
    counter = data.setdefault("contagens", {}).setdefault(counter_key, {
        "data": day_key,
        "vendor_id": vendor_id_value,
        "meta_diaria": DAY_BY_DAY_DAILY_TARGET,
        "eventos": [],
    })
    counter.setdefault("eventos", []).append({
        "saved_at": now,
        "client_id": client_id,
        "status_key": client_payload.get("status_key"),
        "cliente_nome": (client_payload.get("cliente") or {}).get("nome") or "",
    })
    counter["meta_diaria"] = DAY_BY_DAY_DAILY_TARGET
    counter["total_salvos"] = len(counter.get("eventos") or [])
    counter["updated_at"] = now
    try:
        save_json(vendor_day_by_day_file(company_id), data)
    except PermissionError as exc:
        raise ValueError(
            "Nao foi possivel salvar o atendimento porque o Windows bloqueou o arquivo do Day by Day. "
            "Feche outras janelas do CRM, backup ou sincronizacao que possam estar usando o arquivo e tente novamente."
        ) from exc
    response = {
        "message": "Atendimento salvo com sucesso.",
        "contagem": vendor_day_by_day_count_summary(data, vendor_id_value, day_key),
        **vendor_day_by_day_client_payload(company_id, vendor_id_value, client_id, day_key),
    }
    if day_by_day_is_no(form.get("cliente_revenda")):
        blocked = block_final_consumer_client(company_id, client_id)
        response["message"] = "Atendimento salvo e cliente bloqueado como Consumidor Final."
        response["blocked"] = blocked.get("blocked")
    elif day_by_day_is_no(form.get("cliente_ainda_operacao")):
        blocked = block_activity_closed_client(company_id, client_id)
        response["message"] = "Atendimento salvo e cliente bloqueado por Atividade Encerrada."
        response["blocked"] = blocked.get("blocked")
    return response


def vendor_day_by_day_records(company_id: str, vendor_id_value: str) -> list[dict]:
    data = load_json(vendor_day_by_day_file(company_id), {"historico": []})
    rows = []
    for record in data.get("historico") or []:
        if str(record.get("vendor_id") or "") != vendor_id_value:
            continue
        form = record.get("form") or {}
        rows.append({
            "data": record.get("data") or "",
            "data_label": display_date_br(record.get("data")),
            "saved_at": record.get("saved_at") or "",
            "client_id": normalize_identifier(record.get("client_id")),
            "cliente_codigo": record.get("cliente_codigo") or record.get("client_id") or "",
            "cliente_nome": record.get("cliente_nome") or "",
            "status": record.get("status_label") or record.get("status_key") or "",
            "resumo": record.get("resumo") or vendor_day_by_day_contact_summary(form),
            "data_agendamento_contato": form.get("data_agendamento_contato") or "",
            "data_agendamento_iso": (parse_br_date(form.get("data_agendamento_contato")) or date.min).isoformat() if form.get("data_agendamento_contato") else "",
            "descritivo_novo_contato": form.get("descritivo_novo_contato") or "",
            "probabilidade_compra_futura": form.get("probabilidade_compra_futura") or "",
            "form": form,
        })
    rows.sort(key=lambda item: (item.get("data") or "", item.get("saved_at") or ""), reverse=True)
    return rows


def vendor_day_by_day_daily_contacts_payload(company_id: str, vendor_id_value: str, params: dict | None = None):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    params = params or {}
    month_key = str((params.get("month") or [date.today().strftime("%Y-%m")])[0] or "").strip()
    if not re.match(r"^\d{4}-\d{2}$", month_key):
        month_key = date.today().strftime("%Y-%m")
    data = load_json(vendor_day_by_day_file(company_id), {"historico": []})
    grouped = {}
    for record in data.get("historico") or []:
        if str(record.get("vendor_id") or "") != vendor_id_value:
            continue
        day_key = str(record.get("data") or "")[:10]
        parsed = record_date_value(day_key)
        if not parsed:
            continue
        if not is_day_by_day_business_day(parsed):
            continue
        if parsed.strftime("%Y-%m") != month_key:
            continue
        status_key = str(record.get("status_key") or "").strip().lower()
        if status_key not in {"inactive", "never"}:
            continue
        row = grouped.setdefault(day_key, {
            "data": day_key,
            "data_label": display_date_br(day_key),
            "inativos": 0,
            "nunca_comprou": 0,
            "total": 0,
        })
        if status_key == "inactive":
            row["inativos"] += 1
        elif status_key == "never":
            row["nunca_comprou"] += 1
        row["total"] += 1
    rows = sorted(grouped.values(), key=lambda item: item["data"])
    totals = {
        "inativos": sum(row["inativos"] for row in rows),
        "nunca_comprou": sum(row["nunca_comprou"] for row in rows),
        "total": sum(row["total"] for row in rows),
    }
    return {
        "empresa": COMPANIES.get(company_id, company_id),
        "empresa_id": company_id,
        "vendor_id": vendor_id_value,
        "titulo": "Contatos diarios",
        "month": month_key,
        "rows": rows,
        "totals": totals,
    }


def month_business_days(year: int, month: int) -> list[date]:
    last_day = (date(year + 1, 1, 1) - timedelta(days=1)).day if month == 12 else (date(year, month + 1, 1) - timedelta(days=1)).day
    today_value = date.today()
    days = []
    for day_number in range(1, last_day + 1):
        current = date(year, month, day_number)
        if current > today_value:
            break
        if is_day_by_day_business_day(current):
            days.append(current)
    return days


def month_business_day_count(year: int, month: int) -> int:
    last_day = (date(year + 1, 1, 1) - timedelta(days=1)).day if month == 12 else (date(year, month + 1, 1) - timedelta(days=1)).day
    return sum(
        1
        for day_number in range(1, last_day + 1)
        if is_day_by_day_business_day(date(year, month, day_number))
    )


def current_user_vendor_context(token: str, company_id_value: str = "", vendor_id_value: str = "") -> tuple[str, dict, dict]:
    user = current_user_from_token(token)
    if not user:
        raise ValueError("Sessao expirada. Entre novamente no CRM.")

    if user.get("tipo") == "master" and company_id_value and vendor_id_value:
        if company_id_value not in COMPANIES:
            raise ValueError("Empresa invalida.")
        vendors = load_json(vendors_file(company_id_value), [])
        vendor = next(
            (
                item for item in vendors
                if item.get("status") == "Ativo" and str(item.get("id") or "") == str(vendor_id_value or "")
            ),
            None,
        )
        if not vendor:
            raise ValueError("Vendedor ativo nao encontrado.")
        return company_id_value, vendor, user

    for company_id in COMPANIES:
        allowed_ids = vendor_ids_for_user(company_id, user)
        if not allowed_ids:
            continue
        vendors = load_json(vendors_file(company_id), [])
        for vendor in vendors:
            if vendor.get("status") == "Ativo" and str(vendor.get("id") or "") in allowed_ids:
                return company_id, vendor, user
    raise ValueError("Nenhum vendedor ativo vinculado ao seu login para exibir indicadores.")


def vendor_month_sales_goal_value(company_id: str, vendor_id_value: str, year: int, month: int) -> float:
    try:
        all_goals = load_json(vendor_goals_file(company_id), {})
        stored = all_goals.get(vendor_goals_key(vendor_id_value, year), {})
        record = normalize_vendor_goal_record(stored, company_id, vendor_id_value, year)
        month_record = (record.get("months") or {}).get(str(month), {})
        objectives = month_record.get("objetivos") if isinstance(month_record.get("objetivos"), dict) else {}
        sales_goal = objectives.get("vendas_liquidas") if isinstance(objectives.get("vendas_liquidas"), dict) else {}
        value = number_value(sales_goal.get("meta"))
        if value > 0:
            return value

        payload = vendor_goals_payload(company_id, vendor_id_value, year)
        month_record = ((payload.get("goals") or {}).get("months") or {}).get(str(month), {})
        objectives = month_record.get("objetivos") if isinstance(month_record.get("objetivos"), dict) else {}
        sales_goal = objectives.get("vendas_liquidas") if isinstance(objectives.get("vendas_liquidas"), dict) else {}
        return number_value(sales_goal.get("meta"))
    except Exception:
        return 0.0


def vendor_home_month_goals(company_id: str, vendor_id_value: str, year: int, month: int) -> dict:
    goals = {
        "vendas_liquidas_mes": 0.0,
        "clientes_atendidos_mes": 0.0,
        "reativacoes_mes": 0.0,
        "contatos_diarios": 0.0,
    }
    try:
        payload = vendor_goals_payload(company_id, vendor_id_value, year)
        month_record = ((payload.get("goals") or {}).get("months") or {}).get(str(month), {})
        objectives = month_record.get("objetivos") if isinstance(month_record.get("objetivos"), dict) else {}
        goals["vendas_liquidas_mes"] = number_value((objectives.get("vendas_liquidas") or {}).get("meta"))
        goals["clientes_atendidos_mes"] = number_value((objectives.get("clientes_com_vendas") or {}).get("meta"))
        goals["reativacoes_mes"] = number_value((objectives.get("reativacao_inativos") or {}).get("meta"))
        goals["contatos_diarios"] = number_value((objectives.get("contatos_inativos_nunca") or {}).get("meta"))
    except Exception:
        goals["vendas_liquidas_mes"] = vendor_month_sales_goal_value(company_id, vendor_id_value, year, month)
        goals["contatos_diarios"] = DAY_BY_DAY_DAILY_TARGET
    return goals


def home_vendor_panel_payload(token: str, company_id_value: str = "", vendor_id_value: str = ""):
    company_id, vendor, user = current_user_vendor_context(token, company_id_value, vendor_id_value)
    vendor_id_value = str(vendor.get("id") or "")
    today_value = date.today()
    year = today_value.year
    month = today_value.month
    month_key = today_value.strftime("%Y-%m")
    days = month_business_days(year, month)
    business_days_in_month = month_business_day_count(year, month)
    month_goals = vendor_home_month_goals(company_id, vendor_id_value, year, month)
    monthly_sales_goal = number_value(month_goals.get("vendas_liquidas_mes"))
    required_daily_sales_average = monthly_sales_goal / business_days_in_month if business_days_in_month else 0.0
    day_rows = {
        current.isoformat(): {
            "data": current.isoformat(),
            "data_label": display_date_br(current.isoformat()),
            "vendas_liquidas": 0.0,
            "clientes_atendidos": 0,
            "reativacoes": 0,
            "contatos_inativos": 0,
            "contatos_nunca_comprou": 0,
        }
        for current in days
    }

    _vendor, _client_index, region_client_ids = vendor_region_client_context(company_id, vendor_id_value)
    region_client_ids = set(region_client_ids)
    sales_by_client_dates = defaultdict(list)
    daily_clients = defaultdict(set)

    for sale_company_id in vendor_sales_company_ids(company_id):
        for sale in load_json(sales_file(sale_company_id), []):
            if is_excluded_group_sale(sale_company_id, sale.get("CL_NOM")):
                continue
            client_id = normalize_identifier(sale.get("ID_CL"))
            if not client_id or client_id not in region_client_ids:
                continue
            sale_date = record_date_value(sale.get("NF_EMI"))
            if not sale_date:
                continue
            sales_by_client_dates[client_id].append(sale_date)
            if sale_date.year == year and sale_date.month == month and sale_date.isoformat() in day_rows:
                day_rows[sale_date.isoformat()]["vendas_liquidas"] += sale_net_revenue(sale)
                daily_clients[sale_date.isoformat()].add(client_id)

    for day_key, clients in daily_clients.items():
        day_rows[day_key]["clientes_atendidos"] = len(clients)

    for client_id, dates in sales_by_client_dates.items():
        sorted_dates = sorted(set(dates))
        first_current_year = next((item for item in sorted_dates if item.year == year), None)
        had_previous_year_purchase = any(item.year < year for item in sorted_dates)
        if (
            first_current_year
            and had_previous_year_purchase
            and first_current_year.month == month
            and first_current_year.isoformat() in day_rows
        ):
            day_rows[first_current_year.isoformat()]["reativacoes"] += 1

    contacts_payload = vendor_day_by_day_daily_contacts_payload(company_id, vendor_id_value, {"month": [month_key]})
    for row in contacts_payload.get("rows") or []:
        day_key = str(row.get("data") or "")
        if day_key in day_rows:
            day_rows[day_key]["contatos_inativos"] = int(row.get("inativos") or 0)
            day_rows[day_key]["contatos_nunca_comprou"] = int(row.get("nunca_comprou") or 0)

    rows = list(day_rows.values())
    for row in rows:
        row["vendas_liquidas"] = round(row["vendas_liquidas"], 2)
        row["contatos_total"] = row["contatos_inativos"] + row["contatos_nunca_comprou"]
        row["contatos_meta_atingida"] = (
            number_value(month_goals.get("contatos_diarios")) > 0
            and row["contatos_total"] >= number_value(month_goals.get("contatos_diarios"))
        )

    totals = {
        "vendas_liquidas": round(sum(row["vendas_liquidas"] for row in rows), 2),
        "clientes_atendidos": sum(row["clientes_atendidos"] for row in rows),
        "reativacoes": sum(row["reativacoes"] for row in rows),
        "contatos_inativos": sum(row["contatos_inativos"] for row in rows),
        "contatos_nunca_comprou": sum(row["contatos_nunca_comprou"] for row in rows),
    }
    totals["contatos_total"] = totals["contatos_inativos"] + totals["contatos_nunca_comprou"]

    return {
        "empresa": COMPANIES.get(company_id, company_id),
        "empresa_id": company_id,
        "vendedor": vendor_public_record(vendor),
        "usuario": user_public_record(user),
        "selecionavel": user.get("tipo") == "master",
        "vendedores": vendor_page_links_payload().get("rows", []) if user.get("tipo") == "master" else [],
        "month": month_key,
        "month_label": f"{today_value.month:02d}/{today_value.year}",
        "meta": {
            "vendas_liquidas_mes": round(monthly_sales_goal, 2),
            "media_diaria_necessaria": round(required_daily_sales_average, 2),
            "dias_uteis_mes": business_days_in_month,
            "clientes_atendidos_mes": round(number_value(month_goals.get("clientes_atendidos_mes")), 2),
            "reativacoes_mes": round(number_value(month_goals.get("reativacoes_mes")), 2),
            "contatos_diarios": round(number_value(month_goals.get("contatos_diarios")), 2),
        },
        "rows": rows,
        "totals": totals,
    }


def vendor_day_by_day_agenda_payload(company_id: str, vendor_id_value: str):
    rows = [
        row for row in vendor_day_by_day_records(company_id, vendor_id_value)
        if row.get("data_agendamento_contato")
    ]
    rows.sort(key=lambda item: item.get("data_agendamento_iso") or "9999-12-31")
    return {
        "empresa": COMPANIES.get(company_id, company_id),
        "empresa_id": company_id,
        "vendor_id": vendor_id_value,
        "columns": [
            {"key": "data_agendamento_contato", "label": "Data"},
            {"key": "cliente_codigo", "label": "Codigo"},
            {"key": "cliente_nome", "label": "Cliente"},
            {"key": "descritivo_novo_contato", "label": "O que sera tratado"},
            {"key": "resumo", "label": "Resumo anterior"},
        ],
        "rows": rows,
        "total": len(rows),
    }


def vendor_day_by_day_report_payload(company_id: str, vendor_id_value: str, params: dict):
    start_date = parse_br_date((params.get("start_date") or [""])[0])
    end_date = parse_br_date((params.get("end_date") or [""])[0])
    query = normalize_text((params.get("q") or [""])[0])
    rows = []
    for row in vendor_day_by_day_records(company_id, vendor_id_value):
        row_date = record_date_value(row.get("data"))
        if start_date and (not row_date or row_date < start_date):
            continue
        if end_date and (not row_date or row_date > end_date):
            continue
        if query and not search_text_matches((params.get("q") or [""])[0], f"{row.get('cliente_codigo')} {row.get('cliente_nome')}"):
            continue
        rows.append(row)
    return {
        "empresa": COMPANIES.get(company_id, company_id),
        "empresa_id": company_id,
        "vendor_id": vendor_id_value,
        "columns": [
            {"key": "data_label", "label": "Data"},
            {"key": "cliente_codigo", "label": "Codigo"},
            {"key": "cliente_nome", "label": "Cliente"},
            {"key": "status", "label": "Tipo"},
            {"key": "resumo", "label": "Resumo do contato"},
            {"key": "probabilidade_compra_futura", "label": "Probabilidade"},
            {"key": "data_agendamento_contato", "label": "Agenda"},
        ],
        "rows": rows,
        "total": len(rows),
    }


DAY_BY_DAY_FORM_LABELS = [
    ("cliente_revenda", "Cliente Revenda"),
    ("cliente_ainda_operacao", "Cliente ainda em operacao"),
    ("empresa_baixada_cnpj", "Empresa baixada no CNPJ"),
    ("cliente_mudou_ramo", "Cliente mudou ramo de atividade"),
    ("observacoes", "Observacoes"),
    ("nome_contato", "Nome do Contato"),
    ("telefone_ddd", "Telefone com DDD"),
    ("whatsapp", "WhatsApp"),
    ("emails_contato", "E-mails de contato"),
    ("vende_licitacao", "Cliente vende apenas em Licitacao"),
    ("motivo_sem_compra", "Motivo de nao ter mais compra"),
    ("tem_ecommerce", "Cliente tem ecommerce"),
    ("tem_site", "Cliente tem site"),
    ("produtos_site_ecommerce", "Cliente tem nossos produtos no site ou ecommerce"),
    ("relatou_problema_antes", "Cliente relatou algum problema ocorrido antes"),
    ("descritivo_problema_relatado", "Descritivo resumido do problema relatado"),
    ("pertence_grupo_economico", "Cliente pertence a um grupo Economico"),
    ("cliente_matriz_grupo_id", "Codigo do Cliente Matriz do grupo"),
    ("cliente_matriz_grupo_nome", "Cliente Matriz do grupo"),
    ("interesse_agrp", "Cliente relatou interesse em algum produto"),
    ("probabilidade_compra_futura", "Probabilidade de Compra Futura"),
    ("quer_agendar_contato", "Quer agendar uma data para entrar em contato?"),
    ("data_agendamento_contato", "Data do novo contato"),
    ("descritivo_novo_contato", "O que sera tratado no novo contato"),
]


def vendor_ids_for_user(company_id: str, user: dict) -> set[str]:
    vendors = load_json(vendors_file(company_id), [])
    if user and user.get("tipo") == "master":
        return {str(vendor.get("id") or "") for vendor in vendors if vendor.get("id")}
    login_key = normalize_text((user or {}).get("login"))
    user_id_key = normalize_text((user or {}).get("id"))
    allowed = set()
    for vendor in vendors:
        linked_user_id = normalize_text(vendor.get("usuario_vinculado_id"))
        vendor_login = normalize_text(vendor.get("login_acesso"))
        vendor_id = normalize_text(vendor.get("id"))
        if user_id_key and linked_user_id == user_id_key:
            allowed.add(str(vendor.get("id") or ""))
        elif login_key and vendor_login == login_key:
            allowed.add(str(vendor.get("id") or ""))
        elif user_id_key and vendor_id == user_id_key:
            allowed.add(str(vendor.get("id") or ""))
    return allowed


def follow_up_access_context(company_id: str, token: str) -> tuple[dict, set[str], bool]:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    user = current_user_from_token(token)
    if not user:
        raise ValueError("Sessao expirada. Entre novamente no CRM.")
    is_admin = user.get("tipo") == "master"
    allowed_vendor_ids = vendor_ids_for_user(company_id, user)
    if not is_admin and not allowed_vendor_ids:
        raise ValueError("Usuario sem vendedor vinculado para consultar Follow-UP.")
    return user, allowed_vendor_ids, is_admin


def dashboard_vendor_access_context(company_id: str, vendor_id_value: str, token: str):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    user = current_user_from_token(token)
    if not user:
        raise ValueError("Sessao expirada. Entre novamente no CRM.")
    if user.get("tipo") == "master":
        return
    allowed_vendor_ids = vendor_ids_for_user(company_id, user)
    if not allowed_vendor_ids:
        raise ValueError("Usuario sem vendedor vinculado para consultar o Dashboard.")
    if str(vendor_id_value or "").strip() not in allowed_vendor_ids:
        raise ValueError("Usuario sem permissao para consultar este vendedor no Dashboard.")


def follow_up_status_label(status_key: str, status_label: str = "") -> str:
    if status_label:
        return status_label
    if status_key == "inactive":
        return "Inativo"
    if status_key == "never":
        return "Nunca Comprou"
    return status_key or ""


def follow_up_form_answers(form: dict) -> list[dict]:
    answers = []
    for key, label in DAY_BY_DAY_FORM_LABELS:
        value = form.get(key)
        if isinstance(value, list):
            value = ", ".join(str(item) for item in value if str(item).strip())
        value = str(value or "").strip()
        if value:
            answers.append({"key": key, "label": label, "value": value})
    for index, contact in enumerate(form.get("contatos_adicionais") or [], start=1):
        if not isinstance(contact, dict):
            continue
        for key, label in (
            ("nome", "Nome do contato"),
            ("telefone_ddd", "Telefone com DDD"),
            ("whatsapp", "WhatsApp"),
            ("emails_contato", "E-mails de contato"),
        ):
            value = normalize_record(contact.get(key))
            if value:
                answers.append({
                    "key": f"contato_adicional_{index}_{key}",
                    "label": f"Contato adicional {index} - {label}",
                    "value": value,
                })
    return answers


def follow_up_record_row(company_id: str, record: dict, vendor_lookup: dict) -> dict:
    client = record.get("cliente") or {}
    form = record.get("form") or {}
    vendor_id_value = str(record.get("vendor_id") or "")
    vendor = vendor_lookup.get(vendor_id_value) or {}
    client_id = normalize_identifier(record.get("client_id") or client.get("id") or client.get("codigo"))
    uf = str(client.get("uf") or client.get("UF") or "").strip().upper()
    city = str(client.get("cidade") or client.get("CID") or "").strip()
    ddd = str(client.get("ddd") or client_ddd(client) or "").strip()
    saved_at = record.get("saved_at") or record.get("updated_at") or ""
    return {
        "empresa_id": company_id,
        "empresa": COMPANIES.get(company_id, company_id),
        "data": record.get("data") or "",
        "data_label": display_date_br(record.get("data")),
        "saved_at": saved_at,
        "vendor_id": vendor_id_value,
        "vendedor": vendor.get("nome_completo") or vendor_id_value,
        "client_id": client_id,
        "cliente_codigo": record.get("cliente_codigo") or client.get("codigo") or client_id,
        "cliente_nome": record.get("cliente_nome") or client.get("nome") or "",
        "uf": uf,
        "cidade": city,
        "ddd": ddd,
        "regiao": " / ".join(part for part in [uf, city, f"DDD {ddd}" if ddd else ""] if part),
        "status": follow_up_status_label(record.get("status_key") or "", record.get("status_label") or ""),
        "resumo": record.get("resumo") or vendor_day_by_day_contact_summary(form),
        "probabilidade_compra_futura": form.get("probabilidade_compra_futura") or "",
        "data_agendamento_contato": form.get("data_agendamento_contato") or "",
        "descritivo_novo_contato": form.get("descritivo_novo_contato") or "",
        "answers": follow_up_form_answers(form),
        "form": form,
    }


def follow_up_all_rows(company_id: str, allowed_vendor_ids: set[str], is_admin: bool) -> list[dict]:
    data = load_json(vendor_day_by_day_file(company_id), {"historico": []})
    vendors = load_json(vendors_file(company_id), [])
    vendor_lookup = {str(vendor.get("id") or ""): vendor for vendor in vendors}
    rows = []
    for record in data.get("historico") or []:
        vendor_id_value = str(record.get("vendor_id") or "")
        if not is_admin and vendor_id_value not in allowed_vendor_ids:
            continue
        rows.append(follow_up_record_row(company_id, record, vendor_lookup))
    rows.sort(key=lambda item: (item.get("data") or "", item.get("saved_at") or ""), reverse=True)
    return rows


def follow_up_options_payload(company_id: str, token: str):
    _user, allowed_vendor_ids, is_admin = follow_up_access_context(company_id, token)
    rows = follow_up_all_rows(company_id, allowed_vendor_ids, is_admin)
    vendors = load_json(vendors_file(company_id), [])
    vendor_options = [
        {"id": vendor.get("id"), "nome": vendor.get("nome_completo") or vendor.get("id")}
        for vendor in vendors
        if vendor.get("id") and (is_admin or vendor.get("id") in allowed_vendor_ids)
    ]
    return {
        "empresa": COMPANIES.get(company_id, company_id),
        "empresa_id": company_id,
        "is_admin": is_admin,
        "vendors": sorted(vendor_options, key=lambda item: normalize_text(item.get("nome"))),
        "ufs": sorted({row.get("uf") for row in rows if row.get("uf")}),
        "cities": sorted({row.get("cidade") for row in rows if row.get("cidade")}, key=normalize_text),
        "regions": sorted({row.get("regiao") for row in rows if row.get("regiao")}, key=normalize_text),
        "total": len(rows),
    }


def follow_up_payload(company_id: str, params: dict, token: str):
    _user, allowed_vendor_ids, is_admin = follow_up_access_context(company_id, token)
    selected_vendor = str((params.get("vendor_id") or [""])[0] or "").strip()
    if selected_vendor and selected_vendor not in allowed_vendor_ids:
        raise ValueError("Usuario sem permissao para consultar este vendedor.")
    rows = follow_up_all_rows(company_id, allowed_vendor_ids, is_admin)
    start_date = record_date_value((params.get("start_date") or [""])[0])
    end_date = record_date_value((params.get("end_date") or [""])[0])
    region = normalize_text((params.get("region") or [""])[0])
    uf = str((params.get("uf") or [""])[0] or "").strip().upper()
    city = normalize_text((params.get("city") or [""])[0])
    query = (params.get("q") or [""])[0]
    probability = str((params.get("probability") or [""])[0] or "").strip()
    scheduled = str((params.get("scheduled") or [""])[0] or "").strip()
    filtered = []
    for row in rows:
        row_date = record_date_value(row.get("data"))
        if selected_vendor and row.get("vendor_id") != selected_vendor:
            continue
        if start_date and (not row_date or row_date < start_date):
            continue
        if end_date and (not row_date or row_date > end_date):
            continue
        if region and normalize_text(row.get("regiao")) != region:
            continue
        if uf and row.get("uf") != uf:
            continue
        if city and normalize_text(row.get("cidade")) != city:
            continue
        if probability and str(row.get("probabilidade_compra_futura") or "") != probability:
            continue
        if scheduled == "yes" and not row.get("data_agendamento_contato"):
            continue
        if scheduled == "no" and row.get("data_agendamento_contato"):
            continue
        haystack = f"{row.get('cliente_codigo')} {row.get('cliente_nome')} {row.get('vendedor')} {row.get('resumo')}"
        if query and not search_text_matches(query, haystack):
            continue
        filtered.append(row)
    return {
        "empresa": COMPANIES.get(company_id, company_id),
        "empresa_id": company_id,
        "is_admin": is_admin,
        "columns": [
            {"key": "data_label", "label": "Data"},
            {"key": "vendedor", "label": "Vendedor"},
            {"key": "cliente_codigo", "label": "Codigo"},
            {"key": "cliente_nome", "label": "Cliente"},
            {"key": "uf", "label": "UF"},
            {"key": "cidade", "label": "Cidade"},
            {"key": "status", "label": "Tipo"},
            {"key": "resumo", "label": "Resumo"},
            {"key": "probabilidade_compra_futura", "label": "Prob."},
            {"key": "data_agendamento_contato", "label": "Agenda"},
        ],
        "rows": filtered[:500],
        "total": len(filtered),
        "total_registros": len(rows),
    }


def follow_up_export_xlsx(company_id: str, params: dict, token: str) -> bytes:
    payload = follow_up_payload(company_id, params, token)
    rows = payload.get("rows") or []
    base_columns = [
        ("data_label", "Data"),
        ("empresa", "Empresa"),
        ("vendedor", "Vendedor"),
        ("cliente_codigo", "Codigo"),
        ("cliente_nome", "Cliente"),
        ("uf", "UF"),
        ("cidade", "Cidade"),
        ("ddd", "DDD"),
        ("status", "Tipo"),
        ("resumo", "Resumo"),
        ("probabilidade_compra_futura", "Probabilidade"),
        ("data_agendamento_contato", "Agenda"),
        ("descritivo_novo_contato", "Tratativa do novo contato"),
    ]
    answer_columns = []
    seen_labels = set(label for _key, label in base_columns)
    for row in rows:
        for answer in row.get("answers") or []:
            label = normalize_record(answer.get("label"))
            if label and label not in seen_labels:
                seen_labels.add(label)
                answer_columns.append(label)
    export_rows = []
    for row in rows:
        item = {label: row.get(key) for key, label in base_columns}
        for answer in row.get("answers") or []:
            label = normalize_record(answer.get("label"))
            if label:
                item[label] = answer.get("value")
        export_rows.append(item)
    return export_records_xlsx(export_rows, [label for _key, label in base_columns] + answer_columns, "Follow-UP")


def vendor_region_management_payload(company_id: str, vendor_id_value: str):
    payload = vendor_region_clients_payload(company_id, vendor_id_value, {"status": ["all"]})
    rows = payload.get("rows", [])
    rules = active_region_rules(company_id)
    saved_rules = [
        rule
        for rule in rules
        if rule.get("vendor_id") == vendor_id_value
    ]
    exclusions = [
        rule
        for rule in saved_rules
        if str(rule.get("tipo") or "").startswith("exclusao_")
    ]
    return {
        **payload,
        "saved_rules": saved_rules,
        "exclusions": exclusions,
        "summary": {
            "ufs": len({row.get("UF") for row in rows if row.get("UF")}),
            "cities": len({f"{row.get('UF')}|{normalize_text(row.get('CID'))}" for row in rows if row.get("CID")}),
            "ddds": len({row.get("DDD") for row in rows if row.get("DDD")}),
            "clients": payload.get("total_carteira", len(rows)),
        },
    }


def save_vendor_region_exclusions_payload(payload: dict):
    company_id = str(payload.get("company") or "ionlab")
    vendor_id_value = str(payload.get("vendor_id") or "").strip()
    if not vendor_id_value:
        raise ValueError("Selecione um vendedor.")
    items = payload.get("items") or []
    if not isinstance(items, list) or not items:
        raise ValueError("Selecione ao menos um DDD, cidade ou cliente para excluir.")
    latest = None
    for item in items:
        exclusion_type = str(item.get("tipo") or "").strip().lower()
        if exclusion_type == "ddd":
            rule_type = "exclusao_ddd"
        elif exclusion_type == "cidade":
            rule_type = "exclusao_cidade"
        elif exclusion_type == "cliente":
            rule_type = "exclusao_cliente"
        else:
            raise ValueError("Tipo de exclusao invalido.")
        latest = save_region_rule_payload({
            "company": company_id,
            "tipo": rule_type,
            "vendor_id": vendor_id_value,
            "uf": item.get("uf"),
            "ddd": item.get("ddd"),
            "cidade": item.get("cidade"),
            "cliente_id": item.get("cliente_id"),
        })
    return {"message": "Exclusao salva com sucesso.", **vendor_region_management_payload(company_id, vendor_id_value)}


def vendor_page_payload(company_id: str, vendor_id_value: str):
    vendors, _inserted = sync_vendors_from_sales(company_id)
    vendor = next((item for item in vendors if item.get("id") == vendor_id_value), None)
    if not vendor:
        raise ValueError("Vendedor nao encontrado.")
    if vendor.get("status") != "Ativo":
        raise ValueError("Pagina disponivel somente para vendedores ativos.")
    return vendor_regions_payload(company_id, vendor_id_value)


def quote_vendor_prefix(vendor: dict) -> str:
    words = [word for word in re.split(r"\s+", str(vendor.get("nome_completo") or "").strip()) if word]
    if not words:
        return "OR"
    first = words[0][0]
    last = words[-1][0] if len(words) > 1 else words[0][-1]
    return normalize_text(f"{first}{last}")[:2] or "OR"


def next_quote_sequence(company_id: str, vendor_id_value: str) -> int:
    max_sequence = 0
    for quote in load_json(quotes_file(company_id), []):
        if quote.get("vendor_id") != vendor_id_value:
            continue
        max_sequence = max(max_sequence, int(number_value(quote.get("sequencia"))))
    return max_sequence + 1


def quote_number_for_vendor(company_id: str, vendor_id_value: str, sequence: int | None = None) -> str:
    vendor = vendor_record_by_id(company_id, vendor_id_value)
    sequence = sequence or next_quote_sequence(company_id, vendor_id_value)
    return f"{quote_vendor_prefix(vendor)}{sequence:05d}"


def quote_reference_candidates(reference: str) -> list[str]:
    reference = str(reference or "").strip()
    candidates = [reference]
    upper = reference.upper()
    for suffix in ("-P", "-1", "-UND"):
        if upper.endswith(suffix):
            candidates.append(reference[:-len(suffix)])
    candidates.extend([f"{reference}-P", f"{reference}-1", f"{reference}-UND"])
    normalized = []
    seen = set()
    for item in candidates:
        key = normalize_text(item)
        if key and key not in seen:
            seen.add(key)
            normalized.append(key)
    return normalized


def quote_price_map(company_id: str) -> dict:
    return cached_payload(
        ("quote_price_map", company_id),
        [price_table_file(company_id), costing_file(company_id), costing_fabricated_file(company_id)],
        lambda: build_quote_price_map(company_id),
    )


def build_quote_price_map(company_id: str) -> dict:
    prices = {}
    for record in price_table_records_with_cost(company_id, save=True):
        ref = record.get("PR_COD")
        for key in quote_reference_candidates(ref):
            prices.setdefault(key, record)
    return prices


def quote_product_map(company_id: str) -> dict:
    return cached_payload(
        ("quote_product_map", company_id),
        [products_file(company_id)],
        lambda: build_quote_product_map(company_id),
    )


def build_quote_product_map(company_id: str) -> dict:
    products = {}
    for record in load_json(products_file(company_id), []):
        ref = record.get("Referencia")
        for key in quote_reference_candidates(ref):
            products.setdefault(key, record)
    return products


def quote_stock_map(company_id: str) -> dict:
    return cached_payload(
        ("quote_stock_map", company_id),
        [stock_file(company_id)],
        lambda: build_quote_stock_map(company_id),
    )


def build_quote_stock_map(company_id: str) -> dict:
    stocks = {}
    for record in load_json(stock_file(company_id), []):
        ref = record.get("Referencia")
        for key in quote_reference_candidates(ref):
            stocks.setdefault(key, record)
    return stocks


def quote_transit_map(company_id: str) -> dict:
    return cached_payload(
        ("quote_transit_map", company_id),
        [in_transit_file(company_id)],
        lambda: build_quote_transit_map(company_id),
    )


def build_quote_transit_map(company_id: str) -> dict:
    transit = defaultdict(list)
    for record in load_json(in_transit_file(company_id), []):
        ref = record.get("Referencia")
        for key in quote_reference_candidates(ref):
            transit[key].append(record)
    for rows in transit.values():
        rows.sort(key=lambda item: table_sort_value(item, "Previsao"))
    return transit


def quote_ipi_rate(price_record: dict) -> float:
    raw = get_field(price_record or {}, "Aliquota IPI", "Aliq IPI", "Aliquota de IPI", "IPI", "ALIQ_IPI")
    rate = optional_number_value(raw) or 0.0
    return rate / 100 if rate > 1 else rate


def quote_discount_limit(price_record: dict) -> float:
    discount = optional_number_value((price_record or {}).get("PER_DESC")) or 0.0
    return max(0.0, discount)


def normalize_discount_percent(value, limit: float) -> float:
    discount = optional_number_value(value) or 0.0
    discount = max(0.0, discount)
    if discount > limit:
        raise ValueError(f"Desconto informado maior que o limite permitido de {br_decimal(limit)}%.")
    return discount


def quote_item_availability(company_id: str, reference: str, quantity: float) -> dict:
    key_candidates = quote_reference_candidates(reference)
    stock_by_ref = quote_stock_map(company_id)
    transit_by_ref = quote_transit_map(company_id)
    stock_record = next((stock_by_ref.get(key) for key in key_candidates if stock_by_ref.get(key)), {})
    stock_qty = number_value((stock_record or {}).get("Quantidade"))
    transit_rows = []
    for key in key_candidates:
        transit_rows.extend(transit_by_ref.get(key, []))
    transit_rows = sorted(transit_rows, key=lambda item: table_sort_value(item, "Previsao"))
    transit_date = next((row.get("Previsao") for row in transit_rows if row.get("Previsao")), "")
    transit_text = f"Em importacao - Previsao {format_date_br(transit_date)}" if transit_date else "Sob consulta"

    if stock_qty >= quantity and quantity > 0:
        text = "Imediato"
    elif stock_qty > 0:
        missing_qty = max(0.0, quantity - stock_qty)
        missing_text = f"{br_decimal(missing_qty)}"
        if transit_date:
            missing_text = f"{missing_text} - {transit_text}"
        else:
            missing_text = f"{missing_text} - Sob consulta"
        text = f"Parcial ({br_decimal(stock_qty)} em estoque). Falta: {missing_text}"
    elif transit_date:
        text = transit_text
    else:
        text = "Sob consulta"
    return {
        "texto": text,
        "estoque": round(stock_qty, 3),
        "previsao": transit_date,
    }


def format_date_br(value) -> str:
    parsed = record_date_value(value)
    if parsed:
        return parsed.strftime("%d/%m/%Y")
    return str(value or "")


def quote_client_search_payload(company_id: str, query: str, limit: int = 20):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    query_key = normalize_text(query)
    blocked = blocked_client_map(company_id)
    rows = []
    for client in load_json(clients_file(company_id), []):
        client_id = normalize_identifier(client.get("ID"))
        if not client_id or client_id in blocked:
            continue
        haystack = normalize_text(f"{client_id} {client.get('NOM')} {client.get('CGC')} {client.get('FAN')}")
        if query_key and not search_text_matches(query, haystack):
            continue
        rows.append({
            "codigo": client_id,
            "nome": client.get("NOM"),
            "documento": normalize_identifier(client.get("CGC")),
            "telefone": client.get("TEL") or "",
            "whatsapp": client.get("WhatsApp") or client.get("WHATSAPP") or client.get("WAT") or "",
            "email": client.get("MAI") or "",
            "comprador": client.get("Comprador") or client.get("COMPRADOR") or "",
            "cidade": client.get("CID") or "",
            "uf": client.get("UF") or "",
            "label": f"{client_id} - {client.get('NOM') or ''}",
        })
        if len(rows) >= limit:
            break
    return {"clients": rows}


def product_search_dependencies(company_id: str) -> list[Path]:
    return [
        products_file(company_id),
        price_table_file(company_id),
        costing_file(company_id),
        costing_fabricated_file(company_id),
        stock_file(company_id),
        in_transit_file(company_id),
    ]


def search_key_matches(query: str, haystack_key: str) -> bool:
    query_key = normalize_text(query)
    if not query_key:
        return True
    if query_key in haystack_key:
        return True
    tokens = normalize_search_tokens(query)
    return bool(tokens) and all(token in haystack_key for token in tokens)


def is_reference_like_product_query(query: str) -> bool:
    text = str(query or "").strip()
    if not text or re.search(r"\s", text):
        return False
    return bool(re.search(r"\d|[-_/\\.]", text))


def product_row_matches_query(row: dict, query: str, query_key: str, tokens: list[str], reference_like: bool) -> bool:
    reference_key = row.get("reference_key") or ""
    if reference_like:
        return (
            query_key in reference_key
            or bool(tokens) and all(token in reference_key for token in tokens)
        )
    searchable = row.get("searchable") or ""
    return (
        query_key in searchable
        or bool(tokens) and all(token in searchable for token in tokens)
    )


def product_search_rank(query: str, row: dict) -> tuple:
    query_key = normalize_text(query)
    reference_key = row.get("reference_key") or ""
    searchable = row.get("searchable") or ""
    if query_key and reference_key == query_key:
        return (0, reference_key)
    if query_key and reference_key.startswith(query_key):
        return (1, reference_key)
    if query_key and query_key in reference_key:
        return (2, reference_key)
    if query_key and searchable.startswith(query_key):
        return (3, reference_key)
    return (4, reference_key)


def product_transit_labels(transit_by_ref: dict, candidates: list[str]) -> list[str]:
    transit_rows = []
    for key in candidates:
        transit_rows.extend(transit_by_ref.get(key, []))
    transit_dates = []
    for row in sorted(transit_rows, key=lambda item: table_sort_value(item, "Previsao")):
        raw_date = row.get("Previsao")
        if not raw_date:
            continue
        formatted = format_date_br(raw_date)
        quantity = optional_number_value(row.get("QTY"))
        label = formatted
        if quantity is not None:
            label = f"{formatted} - {br_decimal(quantity)} un."
        if label not in transit_dates:
            transit_dates.append(label)
    return transit_dates


def product_search_index(company_id: str) -> list[dict]:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    return shared_cached_payload(
        ("product_search_index", company_id),
        product_search_dependencies(company_id),
        lambda: build_product_search_index(company_id),
    )


def build_product_search_index(company_id: str) -> list[dict]:
    price_by_ref = quote_price_map(company_id)
    product_by_ref = quote_product_map(company_id)
    stock_by_ref = quote_stock_map(company_id)
    transit_by_ref = quote_transit_map(company_id)
    rows = []
    seen = set()

    def first_record(candidates: list[str], records_by_ref: dict) -> dict:
        return next((records_by_ref.get(key) for key in candidates if records_by_ref.get(key)), {})

    def add_row(reference: str, product: dict | None = None, price: dict | None = None, stock: dict | None = None):
        reference = str(reference or "").strip()
        reference_key = normalize_text(reference)
        if not reference or not reference_key or reference_key in seen:
            return
        candidates = quote_reference_candidates(reference)
        product_record = product or first_record(candidates, product_by_ref)
        price_record = price or first_record(candidates, price_by_ref)
        stock_record = stock or first_record(candidates, stock_by_ref)
        short_description = (
            (product_record or {}).get("Descricao resumida")
            or (product_record or {}).get("Descricao")
            or (price_record or {}).get("PR_DES")
            or (stock_record or {}).get("Descricao")
            or ""
        )
        long_description = (
            (product_record or {}).get("Descricao completa")
            or (product_record or {}).get("Descricao")
            or (price_record or {}).get("PR_DES")
            or (stock_record or {}).get("Descricao")
            or ""
        )
        brand = (product_record or {}).get("Marca") or ""
        searchable = normalize_text(
            f"{reference} {short_description} {long_description} {brand} "
            f"{(stock_record or {}).get('AGRP')} {(price_record or {}).get('PR_DES')}"
        )
        seen.add(reference_key)
        rows.append({
            "reference": reference,
            "reference_key": reference_key,
            "searchable": searchable,
            "product": product_record or {},
            "price": price_record or {},
            "stock": stock_record or {},
            "transit_dates": product_transit_labels(transit_by_ref, candidates),
            "descricao_resumida": short_description,
            "descricao_detalhada": long_description,
            "marca": brand,
        })

    for product in load_json(products_file(company_id), []):
        add_row(product.get("Referencia"), product=product)
    for price in price_by_ref.values():
        add_row((price or {}).get("PR_COD"), price=price)
    for stock in stock_by_ref.values():
        add_row((stock or {}).get("Referencia"), stock=stock)

    rows.sort(key=lambda row: row.get("reference_key") or "")
    return rows


def product_index_matches(company_id: str, query: str, limit: int = 20) -> list[dict]:
    query_key = normalize_text(query)
    if not query_key:
        return []
    tokens = normalize_search_tokens(query)
    reference_like = is_reference_like_product_query(query)
    matches = [
        row for row in product_search_index(company_id)
        if product_row_matches_query(row, query, query_key, tokens, reference_like)
    ]
    matches.sort(key=lambda row: product_search_rank(query, row))
    return matches[:limit]


def quote_product_search_payload(company_id: str, query: str, limit: int = 20):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    rows = []
    for item in product_index_matches(company_id, query, limit):
        price = item.get("price") or {}
        unit_price = optional_number_value(price.get("VAL_VEND"))
        ipi_rate = quote_ipi_rate(price)
        discount_limit = quote_discount_limit(price)
        reference = item.get("reference") or ""
        description = item.get("descricao_resumida") or ""
        rows.append({
            "referencia": reference,
            "descricao_resumida": description,
            "descricao_detalhada": item.get("descricao_detalhada") or description,
            "marca": item.get("marca") or "",
            "valor_unitario": round(unit_price, 2) if unit_price is not None else None,
            "aliq_ipi": round(ipi_rate * 100, 4),
            "desconto_maximo": round(discount_limit, 4),
            "preco_encontrado": unit_price is not None,
            "mensagem": "" if unit_price is not None else "Item nao cadastrado na tabela de precos",
            "label": f"{reference} - {description}",
        })
    return {"rows": rows}


def quick_consult_products_payload(company_id: str, query: str, limit: int = 20):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    query_key = normalize_text(query)
    if len(query_key) < 2:
        return {"rows": []}

    rows = []
    for item in product_index_matches(company_id, query, limit):
        price = item.get("price") or {}
        stock = item.get("stock") or {}
        ipi_rate = quote_ipi_rate(price)
        unit_price = optional_number_value(price.get("VAL_VEND"))
        rows.append({
            "referencia": item.get("reference") or "",
            "descricao_resumida": item.get("descricao_resumida") or "",
            "valor_unitario": round(unit_price, 2) if unit_price is not None else None,
            "preco_encontrado": unit_price is not None,
            "aliquota_ipi": round(ipi_rate * 100, 4),
            "saldo_estoque": round(number_value(stock.get("Quantidade")), 3),
            "previsoes_importacao": item.get("transit_dates") or [],
            "marca": item.get("marca") or "",
        })
    return {"rows": rows}


def quote_item_preview_payload(company_id: str, reference: str, quantity_value, discount_value=0):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    reference = str(reference or "").strip()
    quantity = optional_number_value(quantity_value) or 0.0
    if not reference:
        raise ValueError("Informe a referencia.")
    if quantity <= 0:
        raise ValueError("Informe uma quantidade maior que zero.")
    price_by_ref = quote_price_map(company_id)
    product_by_ref = quote_product_map(company_id)
    price = next((price_by_ref.get(key) for key in quote_reference_candidates(reference) if price_by_ref.get(key)), None)
    if not price or optional_number_value(price.get("VAL_VEND")) is None:
        raise ValueError("Item nao cadastrado na tabela de precos.")
    product = next((product_by_ref.get(key) for key in quote_reference_candidates(reference) if product_by_ref.get(key)), {})
    unit_price = optional_number_value(price.get("VAL_VEND")) or 0.0
    ipi_rate = quote_ipi_rate(price)
    discount_limit = quote_discount_limit(price)
    discount_percent = normalize_discount_percent(discount_value, discount_limit)
    discounted_unit = unit_price * (1 - discount_percent / 100)
    ipi_value = discounted_unit * ipi_rate
    availability = quote_item_availability(company_id, reference, quantity)
    return {
        "referencia": reference,
        "descricao_resumida": product.get("Descricao resumida") or price.get("PR_DES") or "",
        "descricao_detalhada": product.get("Descricao completa") or product.get("Descricao") or price.get("PR_DES") or "",
        "marca": product.get("Marca") or "",
        "quantidade": round(quantity, 4),
        "valor_unitario": round(unit_price, 2),
        "desconto": round(discount_percent, 4),
        "desconto_maximo": round(discount_limit, 4),
        "aliq_ipi": round(ipi_rate * 100, 4),
        "valor_ipi": round(ipi_value, 2),
        "total_item": round((discounted_unit + ipi_value) * quantity, 2),
        "disponibilidade": availability.get("texto"),
    }


def quote_context_payload(company_id: str, vendor_id_value: str):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    vendor = vendor_record_by_id(company_id, vendor_id_value)
    return {
        "empresa": COMPANIES[company_id],
        "company_id": company_id,
        "vendedor": vendor,
        "proximo_orcamento": quote_number_for_vendor(company_id, vendor_id_value),
        "sequencia": next_quote_sequence(company_id, vendor_id_value),
    }


def update_quote_client_record(company_id: str, client_payload: dict):
    client_id = normalize_identifier(client_payload.get("codigo"))
    if not client_id:
        return
    clients = load_json(clients_file(company_id), [])
    updated = False
    for client in clients:
        if normalize_identifier(client.get("ID")) != client_id:
            continue
        field_map = [
            ("telefone", "TEL"),
            ("whatsapp", "WHATSAPP"),
            ("email", "MAI"),
            ("comprador", "COMPRADOR"),
        ]
        for source, target in field_map:
            next_value = str(client_payload.get(source) or "").strip()
            if next_value and not str(client.get(target) or "").strip():
                client[target] = next_value
                updated = True
        if updated:
            client["_atualizado_orcamento_em"] = datetime.now().isoformat(timespec="seconds")
        break
    if updated:
        save_json(clients_file(company_id), clients)


def next_revision_number(number: str) -> str:
    text = str(number or "").strip()
    match = re.match(r"^(.*?)-R(\d*)$", text)
    if not match:
        return f"{text}-R"
    base, suffix = match.groups()
    next_index = int(suffix or "0") + 1
    return f"{base}-R{next_index}"


def build_commercial_document(company_id: str, vendor_id_value: str, payload: dict, document_type: str, existing: dict | None = None) -> dict:
    company_id = str(payload.get("company") or "").strip()
    vendor_id_value = str(payload.get("vendor_id") or "").strip()
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    vendor = vendor_record_by_id(company_id, vendor_id_value)
    client_payload = payload.get("cliente") or {}
    client_id = normalize_identifier(client_payload.get("codigo"))
    if not client_id:
        raise ValueError("Selecione o cliente.")
    if client_id in blocked_client_map(company_id):
        blocked = blocked_client_map(company_id)[client_id]
        raise ValueError(f"Cliente bloqueado: {blocked.get('motivo') or 'sem motivo informado'}.")

    items_payload = payload.get("itens") or []
    if not items_payload:
        raise ValueError("Inclua ao menos um item.")

    price_by_ref = quote_price_map(company_id)
    product_by_ref = quote_product_map(company_id)
    quote_items = []
    total_items = 0.0
    total_ipi = 0.0
    for index, raw_item in enumerate(items_payload, start=1):
        reference = str(raw_item.get("referencia") or "").strip()
        quantity = optional_number_value(raw_item.get("quantidade")) or 0.0
        if not reference or quantity <= 0:
            raise ValueError("Todos os itens precisam de referencia e quantidade maior que zero.")
        price = next((price_by_ref.get(key) for key in quote_reference_candidates(reference) if price_by_ref.get(key)), None)
        if not price or optional_number_value(price.get("VAL_VEND")) is None:
            raise ValueError(f"Item {reference} nao cadastrado na tabela de precos.")
        product = next((product_by_ref.get(key) for key in quote_reference_candidates(reference) if product_by_ref.get(key)), {})
        unit_price = optional_number_value(price.get("VAL_VEND")) or 0.0
        ipi_rate = quote_ipi_rate(price)
        discount_limit = quote_discount_limit(price)
        discount_percent = normalize_discount_percent(raw_item.get("desconto"), discount_limit)
        discounted_unit = unit_price * (1 - discount_percent / 100)
        ipi_value = discounted_unit * ipi_rate
        item_base_total = discounted_unit * quantity
        item_ipi_total = ipi_value * quantity
        availability = quote_item_availability(company_id, reference, quantity)
        quote_items.append({
            "item": index,
            "referencia": reference,
            "descricao_resumida": product.get("Descricao resumida") or raw_item.get("descricao_resumida") or price.get("PR_DES") or "",
            "descricao_detalhada": product.get("Descricao completa") or raw_item.get("descricao_detalhada") or product.get("Descricao") or price.get("PR_DES") or "",
            "marca": product.get("Marca") or raw_item.get("marca") or "",
            "quantidade": round(quantity, 4),
            "valor_unitario": round(unit_price, 2),
            "desconto": round(discount_percent, 4),
            "desconto_maximo": round(discount_limit, 4),
            "aliq_ipi": round(ipi_rate * 100, 4),
            "valor_ipi": round(ipi_value, 2),
            "total_item": round(item_base_total + item_ipi_total, 2),
            "disponibilidade": availability.get("texto"),
        })
        total_items += item_base_total
        total_ipi += item_ipi_total

    sequence = existing.get("sequencia") if existing else next_quote_sequence(company_id, vendor_id_value)
    if document_type == "pedido" and not existing:
        base_number = str(payload.get("numero") or "").strip() or quote_number_for_vendor(company_id, vendor_id_value, sequence)
        number = base_number if base_number.upper().startswith("PED-") else f"PED-{base_number}"
    elif existing:
        number = next_revision_number(existing.get("numero"))
    else:
        number = quote_number_for_vendor(company_id, vendor_id_value, sequence)

    now = datetime.now().isoformat(timespec="seconds")
    document = {
        "id": existing.get("id") if existing else f"{company_id}-{vendor_id_value}-{document_type}-{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
        "numero": number,
        "sequencia": sequence,
        "tipo": document_type,
        "company": company_id,
        "empresa": COMPANIES[company_id],
        "vendor_id": vendor_id_value,
        "vendedor": vendor_public_record(vendor),
        "tipo_descricao": payload.get("tipo_descricao") if payload.get("tipo_descricao") in ("resumida", "detalhada") else "resumida",
        "cliente": {
            "codigo": client_id,
            "nome": str(client_payload.get("nome") or "").strip(),
            "telefone": str(client_payload.get("telefone") or "").strip(),
            "whatsapp": str(client_payload.get("whatsapp") or "").strip(),
            "email": str(client_payload.get("email") or "").strip(),
            "comprador": str(client_payload.get("comprador") or "").strip(),
            "documento": normalize_identifier(client_payload.get("documento")),
            "cidade": str(client_payload.get("cidade") or "").strip(),
            "uf": str(client_payload.get("uf") or "").strip(),
        },
        "itens": quote_items,
        "resumo": {
            "total_itens": round(total_items, 2),
            "total_ipi": round(total_ipi, 2),
            "total_orcamento": round(total_items + total_ipi, 2),
        },
        "criado_em": existing.get("criado_em") if existing else now,
        "atualizado_em": now,
    }
    if existing:
        history = list(existing.get("historico_revisoes") or [])
        previous = copy.deepcopy(existing)
        previous.pop("historico_revisoes", None)
        history.append(previous)
        document["historico_revisoes"] = history
        document["revisado_em"] = now
    return document


def save_quote_payload(payload: dict):
    company_id = str(payload.get("company") or "").strip()
    quote_id = str(payload.get("id") or payload.get("quote_id") or "").strip()
    quotes = load_json(quotes_file(company_id), [])
    existing = None
    if quote_id:
        existing = next((item for item in quotes if item.get("id") == quote_id), None)
        if not existing:
            raise ValueError("Orcamento nao encontrado para revisao.")
    quote = build_commercial_document(company_id, str(payload.get("vendor_id") or ""), payload, "orcamento", existing)
    if existing:
        quotes = [quote if item.get("id") == quote_id else item for item in quotes]
    else:
        quotes.append(quote)
    save_json(quotes_file(company_id), quotes)
    update_quote_client_record(company_id, quote["cliente"])
    return quote


def save_order_payload(payload: dict):
    company_id = str(payload.get("company") or "").strip()
    order_id = str(payload.get("id") or payload.get("order_id") or "").strip()
    orders = load_json(orders_file(company_id), [])
    existing = None
    if order_id:
        existing = next((item for item in orders if item.get("id") == order_id), None)
        if not existing:
            raise ValueError("Pedido nao encontrado para revisao.")
    order = build_commercial_document(company_id, str(payload.get("vendor_id") or ""), payload, "pedido", existing)
    if payload.get("quote_id"):
        order["quote_id"] = payload.get("quote_id")
        order["orcamento_origem"] = payload.get("orcamento_origem") or payload.get("numero")
    if existing:
        orders = [order if item.get("id") == order_id else item for item in orders]
    else:
        orders.append(order)
    save_json(orders_file(company_id), orders)
    update_quote_client_record(company_id, order["cliente"])
    return order


def quote_by_id(company_id: str, quote_id: str) -> dict:
    for quote in load_json(quotes_file(company_id), []):
        if quote.get("id") == quote_id:
            return quote
    raise ValueError("Orcamento nao encontrado.")


def order_by_id(company_id: str, order_id: str) -> dict:
    for order in load_json(orders_file(company_id), []):
        if order.get("id") == order_id:
            return order
    raise ValueError("Pedido nao encontrado.")


def commercial_document_search_payload(company_id: str, vendor_id_value: str, query: str, kind: str, limit: int = 20, page: int = 1):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    source = orders_file(company_id) if kind == "pedido" else quotes_file(company_id)
    query_key = normalize_text(query)
    rows = []
    for document in load_json(source, []):
        if vendor_id_value and document.get("vendor_id") != vendor_id_value:
            continue
        client = document.get("cliente") or {}
        haystack = normalize_text(f"{document.get('numero')} {client.get('nome')} {client.get('documento')} {client.get('codigo')}")
        if query_key and not search_text_matches(query, haystack):
            continue
        rows.append({
            "id": document.get("id"),
            "numero": document.get("numero"),
            "cliente": client.get("nome"),
            "cliente_documento": client.get("documento"),
            "total": (document.get("resumo") or {}).get("total_orcamento"),
            "atualizado_em": document.get("atualizado_em") or document.get("criado_em"),
            "documento": document,
            "label": f"{document.get('numero')} - {client.get('nome') or ''}",
        })
    rows.sort(key=lambda item: (str(item.get("atualizado_em") or ""), normalize_text(item.get("numero"))), reverse=True)
    total = len(rows)
    page_size = max(1, min(int(optional_number_value(limit) or 20), 100))
    total_pages = max(1, math.ceil(total / page_size))
    page_number = max(1, min(int(optional_number_value(page) or 1), total_pages))
    start = (page_number - 1) * page_size
    paged_rows = rows[start:start + page_size]
    return {
        "rows": paged_rows,
        "kind": kind,
        "page": page_number,
        "page_size": page_size,
        "total": total,
        "total_pages": total_pages,
    }


def pdf_clean_text(value) -> str:
    text = clean_text_encoding(str(value or ""))
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return text.replace("\u00b0", "o").replace("\u2013", "-").replace("\u2014", "-").replace("\u2022", "-")


def pdf_string(value) -> bytes:
    raw = pdf_clean_text(value).encode("latin-1", "replace")
    raw = raw.replace(b"\\", b"\\\\").replace(b"(", b"\\(").replace(b")", b"\\)")
    return b"(" + raw + b")"


def pdf_money(value) -> str:
    text = f"{number_value(value):,.2f}"
    text = text.replace(",", "_").replace(".", ",").replace("_", ".")
    return f"R$ {text}"


def pdf_text_op(x: float, y: float, text, size: int = 9, font: str = "F1") -> bytes:
    return b"BT /" + font.encode("ascii") + b" " + str(size).encode("ascii") + b" Tf " + f"{x:.2f} {y:.2f}".encode("ascii") + b" Td " + pdf_string(text) + b" Tj ET\n"


def pdf_text_width(text, size: int = 9) -> float:
    return len(pdf_clean_text(text)) * size * 0.48


def pdf_text_right_op(x_right: float, y: float, text, size: int = 9, font: str = "F1") -> bytes:
    return pdf_text_op(x_right - pdf_text_width(text, size), y, text, size, font)


def pdf_line_op(x1: float, y1: float, x2: float, y2: float, width: float = 0.5) -> bytes:
    return f"{width:.2f} w {x1:.2f} {y1:.2f} m {x2:.2f} {y2:.2f} l S\n".encode("ascii")


def pdf_rect_op(x: float, y: float, width: float, height: float, line_width: float = 0.5) -> bytes:
    return f"{line_width:.2f} w {x:.2f} {y:.2f} {width:.2f} {height:.2f} re S\n".encode("ascii")


def pdf_wrap(text, max_chars: int) -> list[str]:
    words = pdf_clean_text(text).split()
    lines = []
    current = ""
    for word in words:
        next_text = f"{current} {word}".strip()
        if len(next_text) > max_chars and current:
            lines.append(current)
            current = word
        else:
            current = next_text
    if current:
        lines.append(current)
    return lines or [""]


def build_pdf_document(page_streams: list[bytes], logo_bytes: bytes | None = None, logo_size: tuple[int, int] | None = None) -> bytes:
    objects: list[bytes] = []

    def add_object(payload: bytes) -> int:
        objects.append(payload)
        return len(objects)

    catalog_id = add_object(b"")
    pages_id = add_object(b"")
    font_id = add_object(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica /Encoding /WinAnsiEncoding >>")
    bold_font_id = add_object(b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica-Bold /Encoding /WinAnsiEncoding >>")
    logo_id = None
    if logo_bytes and logo_size:
        logo_id = add_object(
            b"<< /Type /XObject /Subtype /Image /Width "
            + str(logo_size[0]).encode("ascii")
            + b" /Height "
            + str(logo_size[1]).encode("ascii")
            + b" /ColorSpace /DeviceRGB /BitsPerComponent 8 /Filter /DCTDecode /Length "
            + str(len(logo_bytes)).encode("ascii")
            + b" >>\nstream\n"
            + logo_bytes
            + b"\nendstream"
        )

    page_ids = []
    for stream in page_streams:
        content_id = add_object(
            b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream"
        )
        xobject = b""
        if logo_id:
            xobject = b" /XObject << /Logo " + str(logo_id).encode("ascii") + b" 0 R >>"
        page_id = add_object(
            b"<< /Type /Page /Parent "
            + str(pages_id).encode("ascii")
            + b" 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 "
            + str(font_id).encode("ascii")
            + b" 0 R /F2 "
            + str(bold_font_id).encode("ascii")
            + b" 0 R >>"
            + xobject
            + b" >> /Contents "
            + str(content_id).encode("ascii")
            + b" 0 R >>"
        )
        page_ids.append(page_id)

    objects[catalog_id - 1] = b"<< /Type /Catalog /Pages " + str(pages_id).encode("ascii") + b" 0 R >>"
    kids = b" ".join(str(page_id).encode("ascii") + b" 0 R" for page_id in page_ids)
    objects[pages_id - 1] = b"<< /Type /Pages /Kids [" + kids + b"] /Count " + str(len(page_ids)).encode("ascii") + b" >>"

    output = BytesIO()
    output.write(b"%PDF-1.4\n%\xe2\xe3\xcf\xd3\n")
    offsets = []
    for index, payload in enumerate(objects, start=1):
        offsets.append(output.tell())
        output.write(f"{index} 0 obj\n".encode("ascii"))
        output.write(payload)
        output.write(b"\nendobj\n")
    xref = output.tell()
    output.write(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    output.write(b"0000000000 65535 f \n")
    for offset in offsets:
        output.write(f"{offset:010d} 00000 n \n".encode("ascii"))
    output.write(
        b"trailer\n<< /Size "
        + str(len(objects) + 1).encode("ascii")
        + b" /Root "
        + str(catalog_id).encode("ascii")
        + b" 0 R >>\nstartxref\n"
        + str(xref).encode("ascii")
        + b"\n%%EOF"
    )
    return output.getvalue()


def quote_pdf_bytes(quote: dict) -> bytes:
    logo_path = STATIC_DIR / "assets" / "ion_logo.jpg"
    logo_bytes = logo_path.read_bytes() if logo_path.exists() else None
    logo_size = None
    if logo_bytes and Image is not None:
        try:
            with Image.open(BytesIO(logo_bytes)) as image:
                logo_size = image.size
        except Exception:
            logo_bytes = None
    elif logo_bytes:
        logo_bytes = None

    show_detailed_description = quote.get("tipo_descricao") == "detalhada"
    pages = []
    commands = []
    y = 800

    def new_page():
        nonlocal commands, y
        if commands:
            pages.append(b"".join(commands))
        commands = []
        if logo_bytes:
            commands.append(b"q 56 0 0 56 40 762 cm /Logo Do Q\n")
        commands.append(pdf_text_op(110, 800, "Ionlab Equipamentos Para Laboratorios e Hospitais Ltda", 11, "F2"))
        commands.append(pdf_text_op(110, 784, "Cnpj: 11.916.966/0001-90 | IE: 10252814-77", 9))
        commands.append(pdf_text_op(110, 770, "Rua Martins Deda, 453, Bairro Chapada, Araucaria - Parana", 9))
        title = "Pedido" if quote.get("tipo") == "pedido" else "Orcamento"
        commands.append(pdf_text_op(420, 800, f"{title} No {quote.get('numero')}", 12, "F2"))
        commands.append(pdf_text_op(420, 784, f"Data: {format_date_br(quote.get('criado_em'))}", 9))
        commands.append(pdf_line_op(40, 750, 555, 750, 0.8))
        y = 730

    def ensure_space(height):
        nonlocal y
        if y - height < 55:
            new_page()

    new_page()
    client = quote.get("cliente") or {}
    commands.append(pdf_text_op(40, y, "Dados do cliente", 11, "F2")); y -= 16
    commands.append(pdf_text_op(40, y, f"Codigo: {client.get('codigo')} | Cliente: {client.get('nome')}", 9)); y -= 14
    commands.append(pdf_text_op(40, y, f"Telefone: {client.get('telefone')} | WhatsApp: {client.get('whatsapp')} | Email: {client.get('email')}", 9)); y -= 14
    commands.append(pdf_text_op(40, y, f"Comprador: {client.get('comprador')} | Cidade/UF: {client.get('cidade')} - {client.get('uf')}", 9)); y -= 24

    def draw_items_header():
        nonlocal y
        commands.append(pdf_text_op(40, y, "Itens do orcamento", 11, "F2")); y -= 18
        commands.append(pdf_line_op(40, y + 5, 555, y + 5, 0.5))
        headers = [
            (42, "Item"),
            (66, "Referencia"),
            (132, "Descricao"),
            (340, "Qtd"),
            (395, "V.Unit."),
            (435, "IPI"),
            (500, "V IPI"),
            (555, "Total"),
        ]
        for x, label in headers:
            if label in {"Qtd", "V.Unit.", "IPI", "V IPI", "Total"}:
                commands.append(pdf_text_right_op(x, y, label, 8, "F2"))
            else:
                commands.append(pdf_text_op(x, y, label, 8, "F2"))
        y -= 12
        commands.append(pdf_line_op(40, y + 4, 555, y + 4, 0.5))
        y -= 10

    def write_detail_box(lines: list[str], availability_lines: list[str], title: str = "Descritivo Detalhado"):
        nonlocal y
        remaining = list(lines) or [""]
        availability = list(availability_lines)
        first_box = True
        while remaining:
            if y < 145:
                new_page()
                draw_items_header()
            available_height = max(60, y - 65)
            line_height = 8.5
            max_lines = max(3, int((available_height - 32) / line_height))
            if len(remaining) <= max_lines and availability:
                footer_height = 15 + len(availability) * 9
                max_lines = max(3, int((available_height - 32 - footer_height) / line_height))
            chunk = remaining[:max_lines]
            remaining = remaining[max_lines:]
            include_footer = not remaining and bool(availability)
            footer_height = 15 + len(availability) * 9 if include_footer else 0
            box_height = 28 + len(chunk) * line_height + footer_height
            box_top = y
            box_bottom = box_top - box_height
            commands.append(pdf_rect_op(40, box_bottom, 515, box_height, 0.7))
            header = title if first_box else f"{title} - continuacao"
            commands.append(pdf_text_op(50, box_top - 13, header, 8, "F2"))
            text_y = box_top - 27
            for line in chunk:
                commands.append(pdf_text_op(52, text_y, line, 7))
                text_y -= line_height
            if include_footer:
                commands.append(pdf_line_op(50, text_y - 2, 545, text_y - 2, 0.25))
                text_y -= 13
                for line in availability:
                    commands.append(pdf_text_op(52, text_y, line, 7, "F2"))
                    text_y -= 9
            y = box_bottom - 10
            first_box = False

    draw_items_header()

    for item in quote.get("itens", []):
        summary_lines = pdf_wrap(item.get("descricao_resumida") or item.get("descricao_detalhada"), 32)[:3]
        detail_lines = pdf_wrap(item.get("descricao_detalhada") or item.get("descricao_resumida"), 145)
        availability_lines = pdf_wrap(f"Disponibilidade: {item.get('disponibilidade')}", 130)
        row_height = 20 + len(summary_lines) * 10
        ensure_space(row_height)
        row_start = y
        commands.append(pdf_text_op(42, row_start, item.get("item"), 8))
        commands.append(pdf_text_op(66, row_start, pdf_clean_text(item.get("referencia"))[:14], 7))
        commands.append(pdf_text_right_op(340, row_start, br_decimal(item.get("quantidade")), 7))
        commands.append(pdf_text_right_op(395, row_start, pdf_money(item.get("valor_unitario")), 7))
        commands.append(pdf_text_right_op(435, row_start, f"{br_decimal(item.get('aliq_ipi'))}%", 7))
        item_ipi_total = number_value(item.get("valor_ipi")) * number_value(item.get("quantidade"))
        commands.append(pdf_text_right_op(500, row_start, pdf_money(item_ipi_total), 7))
        commands.append(pdf_text_right_op(555, row_start, pdf_money(item.get("total_item")), 7, "F2"))
        desc_y = row_start
        for line in summary_lines:
            commands.append(pdf_text_op(132, desc_y, line, 8))
            desc_y -= 10
        y = desc_y - 5
        if show_detailed_description:
            write_detail_box(detail_lines, availability_lines)
        else:
            for line in pdf_wrap(f"Disponibilidade: {item.get('disponibilidade')}", 58):
                commands.append(pdf_text_op(132, y, line, 8))
                y -= 10
            y -= 4
        commands.append(pdf_line_op(40, y, 555, y, 0.25))
        y -= 12

    ensure_space(90)
    resumo = quote.get("resumo") or {}
    y -= 8
    commands.append(pdf_text_op(365, y, "Resumo", 11, "F2")); y -= 16
    commands.append(pdf_text_op(365, y, f"Total dos itens: {pdf_money(resumo.get('total_itens'))}", 10)); y -= 15
    commands.append(pdf_text_op(365, y, f"Total de IPI: {pdf_money(resumo.get('total_ipi'))}", 10)); y -= 15
    commands.append(pdf_text_op(365, y, f"Total do Orcamento: {pdf_money(resumo.get('total_orcamento'))}", 12, "F2"))

    def write_footer_section(title: str, lines: list[str]):
        nonlocal y
        wrapped = []
        for line in lines:
            wrapped.extend(pdf_wrap(line, 128))
        height = 18 + len(wrapped) * 8.5
        if y - height < 55:
            new_page()
        commands.append(pdf_text_op(40, y, title, 9, "F2"))
        y -= 12
        for line in wrapped:
            commands.append(pdf_text_op(48, y, line, 7))
            y -= 8.5
        y -= 5

    vendor = quote.get("vendedor") or {}
    try:
        current_vendor = vendor_record_by_id(str(quote.get("company") or ""), str(quote.get("vendor_id") or ""))
        vendor = {**vendor, **vendor_public_record(current_vendor)}
    except Exception:
        pass
    vendor_name = (
        vendor.get("nome_completo")
        or vendor.get("nome")
        or vendor.get("Nome Completo")
        or "Nao informado"
    )
    vendor_phone = (
        vendor.get("telefone")
        or vendor.get("Telefone")
        or vendor.get("TEL")
        or vendor.get("telefone_ddd")
        or "Nao informado"
    )
    vendor_whatsapp = (
        vendor.get("whatsapp")
        or vendor.get("WhatsApp")
        or vendor.get("WHATSAPP")
        or vendor.get("WattSapp")
        or "Nao informado"
    )

    y -= 30
    write_footer_section("Informacoes comerciais", [
        "Validade da proposta: 10 dias.",
        "Prazo de entrega especificado em cada item. Para itens sob consulta, verifique com vendedor.",
        "Condicoes de pagamento dessa proposta: A vista. Para condicoes de pagamento a prazo, consulte vendedor e/ou setor financeiro para avaliacao.",
        "Frete: Sob negociacao.",
        "Impostos: Inclusos no orcamento. IPI destacado no orcamento.",
        "Regime tributario da empresa: Empresa com apuracao Lucro Presumido.",
    ])
    write_footer_section("Garantia", [
        "Prazo de garantia dos equipamentos: conforme condicoes comerciais, certificado do equipamento e/ou manual do fabricante.",
        "Condicoes que invalidam a garantia: mau uso, instalacao inadequada, violacao do equipamento, manutencao por terceiros nao autorizados, danos eletricos, quedas, contato com agentes quimicos indevidos ou uso fora das especificacoes tecnicas.",
    ])
    write_footer_section("Observacoes legais", [
        "\"Os precos poderao sofrer alteracao apos o vencimento da proposta.\"",
        "\"O pedido sera considerado confirmado apos aprovacao formal e/ou recebimento do pagamento acordado.\"",
        "\"Equipamentos fabricados sob encomenda nao poderao ser cancelados apos o inicio da producao.\"",
    ])
    write_footer_section("Contato", [
        f"Vendedor: {vendor_name}.",
        f"Telefone da empresa/vendedor: {vendor_phone}. WhatsApp: {vendor_whatsapp}.",
        "Consulte nosso site para mais detalhes: www.ionlab.com.br.",
        "Siga-nos nas redes sociais: @ionlaboficial.",
    ])
    commands.append(pdf_text_op(40, 35, "Documento gerado pelo CRM Ionlab.", 8))
    pages.append(b"".join(commands))
    return build_pdf_document(pages, logo_bytes, logo_size)


VENDOR_GOAL_OBJECTIVES = [
    ("vendas_liquidas", "Vendas Liquidas sua regiao", "currency"),
    ("reativacao_inativos", "Reativacao Inativos", "number"),
    ("clientes_com_vendas", "Quantidade Minima de Clientes com Vendas no Mes", "number"),
    ("contatos_inativos_nunca", "Contatos diarios com Clientes Inativos / Nunca Comprou", "number"),
    ("novos_clientes", "Novos Clientes", "currency"),
    ("sem_giro", "Vendas de Itens/Equipamentos Sem Giro", "percent"),
]

DAY_BY_DAY_DAILY_TARGET = 50


def default_vendor_goal_month():
    default_objectives = {
        "vendas_liquidas": {"meta": 0.0, "peso": 60.0},
        "reativacao_inativos": {"meta": 0.0, "peso": 10.0},
        "clientes_com_vendas": {"meta": 0.0, "peso": 20.0},
        "contatos_inativos_nunca": {"meta": 50.0, "peso": 10.0},
        "novos_clientes": {"meta": 50.0, "peso": 0.0},
        "sem_giro": {"meta": 0.0, "peso": 0.0},
    }
    return {
        "percentual_aumento": 0.0,
        "quantidade_minima_clientes": 0.0,
        "contatos_diarios_inativos": 50.0,
        "valor_novo_cliente": 50.0,
        "objetivos": {
            key: dict(default_objectives.get(key, {"meta": 0.0, "peso": 0.0}))
            for key, _label, _kind in VENDOR_GOAL_OBJECTIVES
        },
    }


def default_vendor_goal_record(company_id: str, vendor_id_value: str, year: int):
    return {
        "company": company_id,
        "vendor_id": vendor_id_value,
        "year": year,
        "comissao_percentual": 1.0,
        "months": {str(month): default_vendor_goal_month() for month in range(1, 13)},
    }


def vendor_goals_key(vendor_id_value: str, year: int) -> str:
    return f"{vendor_id_value}|{year}"


def vendor_goal_option_record(vendor: dict) -> dict:
    return {
        "id": vendor.get("id"),
        "nome_completo": vendor.get("nome_completo"),
        "status": vendor.get("status") or "Inativo",
    }


def vendor_sales_goal_base(company_id: str, vendor_id_value: str, target_year=None):
    year = int(optional_number_value(target_year) or CURRENT_YEAR)
    empty = {
        "media_ultimos_2_anos": 0.0,
        "media_ano_corrente": 0.0,
        "media_base": 0.0,
        "anos_historicos": [],
        "meses_fechados": [],
        "criterio": "media_top_3_meses_ultimos_2_anos_fechados",
        "meses_base": [],
    }
    try:
        payload = vendor_regions_payload(company_id, vendor_id_value)
    except Exception:
        return empty

    grouped = (payload.get("agrupado") or [{}])[0]
    closed_months = [int(month) for month in payload.get("meses_fechados", [])]
    last_two_years = [base_year for base_year in (year - 2, year - 1) if base_year >= START_YEAR]
    months_by_year = grouped.get("vendas_meses_anos") or {}
    monthly_values = []
    for base_year in last_two_years:
        monthly_values.extend([
            {
                "year": base_year,
                "month": month,
                "value": number_value((months_by_year.get(str(base_year)) or {}).get(str(month))),
            }
            for month in range(1, 13)
        ])
    selected_months = sorted(monthly_values, key=lambda item: item["value"], reverse=True)[:3]

    base_average = (
        sum(item["value"] for item in selected_months) / len(selected_months)
        if selected_months else 0.0
    )
    return {
        "media_ultimos_2_anos": round(base_average, 2),
        "media_ano_corrente": 0.0,
        "media_base": round(base_average, 2),
        "anos_historicos": last_two_years,
        "meses_fechados": closed_months,
        "criterio": "media_top_3_meses_ultimos_2_anos_fechados",
        "meses_base": [
            {"ano": item["year"], "mes": item["month"], "valor": round(item["value"], 2)}
            for item in selected_months
        ],
    }


def vendor_client_sales_goal_base(company_id: str, vendor_id_value: str, target_year=None):
    year = int(optional_number_value(target_year) or CURRENT_YEAR)
    empty = {
        "criterio": "maior_quantidade_clientes_mes_ultimos_2_anos_fechados",
        "anos_historicos": [],
        "maior_mes": None,
        "quantidade_sugerida": 0,
        "meses_base": [],
    }
    try:
        _vendor, client_index, region_client_ids = vendor_region_client_context(company_id, vendor_id_value)
    except Exception:
        return empty

    last_two_years = [base_year for base_year in (year - 2, year - 1) if base_year >= START_YEAR]
    if not last_two_years:
        return empty

    clients_by_month = {
        (base_year, month): set()
        for base_year in last_two_years
        for month in range(1, 13)
    }
    for sale_company_id in vendor_sales_company_ids(company_id):
        for sale in load_json(sales_file(sale_company_id), []):
            if is_excluded_group_sale(sale_company_id, sale.get("CL_NOM")):
                continue
            sale_year = record_year(sale.get("NF_EMI"))
            sale_month = record_month(sale.get("NF_EMI"))
            if sale_year not in last_two_years or not sale_month or not (1 <= sale_month <= 12):
                continue
            client_id = normalize_identifier(sale.get("ID_CL"))
            if client_id not in region_client_ids:
                continue
            client = client_index.get(client_id)
            group_key = economic_client_key(client_id, client, sale)
            clients_by_month[(sale_year, sale_month)].add(group_key)

    month_rows = [
        {"ano": base_year, "mes": month, "clientes": len(client_ids)}
        for (base_year, month), client_ids in clients_by_month.items()
    ]
    best = max(month_rows, key=lambda item: (item["clientes"], item["ano"], item["mes"]), default=None)
    return {
        "criterio": "maior_quantidade_clientes_mes_ultimos_2_anos_fechados",
        "anos_historicos": last_two_years,
        "maior_mes": best,
        "quantidade_sugerida": int(best.get("clientes") or 0) if best else 0,
        "meses_base": sorted(month_rows, key=lambda item: (item["ano"], item["mes"])),
    }


def vendor_items_key(year: int, month: int) -> str:
    return f"{int(year):04d}-{int(month):02d}"


def vendor_monthly_items_payload(company_id: str, year_value=None, month_value=None):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    year = int(optional_number_value(year_value) or CURRENT_YEAR)
    month = int(optional_number_value(month_value) or date.today().month)
    key = vendor_items_key(year, month)
    data = load_json(vendor_monthly_items_file(company_id), {})
    items = data.get(key, []) if isinstance(data.get(key), list) else []
    return {
        "empresa": COMPANIES[company_id],
        "empresa_id": company_id,
        "year": year,
        "month": month,
        "items": items,
    }


def product_info_maps(company_id: str):
    stock_by_ref = {
        normalize_text(record.get("Referencia")): record
        for record in stock_records_with_last_exit(company_id)
        if normalize_text(record.get("Referencia"))
    }
    costing_by_ref = {
        normalize_text(record.get("Referencia")): record
        for record in load_json(costing_file(company_id), [])
        if normalize_text(record.get("Referencia"))
    }
    last_sale_by_ref = {}
    for sale in load_json(sales_file(company_id), []):
        ref_key = normalize_text(sale.get("PR_COD"))
        sale_date = record_date_value(sale.get("NF_EMI"))
        if not ref_key or not sale_date:
            continue
        current = last_sale_by_ref.get(ref_key)
        if current is None or sale_date > current["date"]:
            qty = number_value(sale.get("PR_QTD"))
            current_price = sale_net_revenue(sale) / qty if qty else 0.0
            last_sale_by_ref[ref_key] = {"date": sale_date, "unit_price": current_price}
    return stock_by_ref, costing_by_ref, last_sale_by_ref


def product_search_payload(company_id: str, query: str, limit: int = 20):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    query_key = normalize_text(query)
    stock_by_ref, costing_by_ref, last_sale_by_ref = product_info_maps(company_id)
    refs = set(stock_by_ref) | set(costing_by_ref)
    for sale in load_json(sales_file(company_id), []):
        ref_key = normalize_text(sale.get("PR_COD"))
        if ref_key:
            refs.add(ref_key)

    rows = []
    seen = set()
    for ref_key in refs:
        stock = stock_by_ref.get(ref_key, {})
        costing = costing_by_ref.get(ref_key, {})
        reference = str(stock.get("Referencia") or costing.get("Referencia") or "").strip()
        description = str(stock.get("Descricao") or costing.get("Descricao item") or "").strip()
        if not reference:
            continue
        searchable = normalize_text(f"{reference} {description}")
        if query_key and not search_text_matches(query, searchable):
            continue
        if ref_key in seen:
            continue
        seen.add(ref_key)
        last_sale = last_sale_by_ref.get(ref_key, {})
        rows.append({
            "referencia": reference,
            "descricao": description,
            "estoque": round(number_value(stock.get("Quantidade")), 3),
            "ultima_venda": last_sale.get("date").isoformat() if last_sale.get("date") else "",
            "valor_unitario_tabela": round(number_value(last_sale.get("unit_price")) or number_value(stock.get("V.Unitario")), 2),
        })
    rows.sort(key=lambda item: normalize_text(f"{item['referencia']} {item['descricao']}"))
    return {"rows": rows[:max(1, min(limit, 50))]}


def save_vendor_monthly_items_payload(payload: dict):
    company_id = payload.get("company", "")
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    year = int(optional_number_value(payload.get("year")) or CURRENT_YEAR)
    month = int(optional_number_value(payload.get("month")) or date.today().month)
    incoming_items = payload.get("items") if isinstance(payload.get("items"), list) else []
    items = []
    seen = set()
    for item in incoming_items:
        ref_key = normalize_text(item.get("referencia"))
        if not ref_key or ref_key in seen:
            continue
        seen.add(ref_key)
        items.append({
            "referencia": str(item.get("referencia") or "").strip(),
            "descricao": str(item.get("descricao") or "").strip(),
            "estoque": round(number_value(item.get("estoque")), 3),
            "ultima_venda": str(item.get("ultima_venda") or ""),
            "valor_unitario_tabela": round(number_value(item.get("valor_unitario_tabela")), 2),
        })
    data = load_json(vendor_monthly_items_file(company_id), {})
    data[vendor_items_key(year, month)] = items
    save_json(vendor_monthly_items_file(company_id), data)
    return {"message": "Itens/Equipamentos do mes salvos com sucesso.", "items": items, "total": len(items)}


def vendor_goal_achievements(company_id: str, vendor_id_value: str, year: int, goal_record: dict):
    try:
        vendor, client_index, region_client_ids = vendor_region_client_context(company_id, vendor_id_value)
    except Exception:
        return {str(month): {} for month in range(1, 13)}

    achievements = {str(month): {key: 0.0 for key, _label, _kind in VENDOR_GOAL_OBJECTIVES} for month in range(1, 13)}
    _stock_by_ref, _costing_by_ref, last_sale_by_ref = product_info_maps(company_id)
    sales = []
    other_region_sales = []
    vendor_name_key = normalize_text(vendor.get("nome_completo"))
    for sale_company_id in vendor_sales_company_ids(company_id):
        for sale in load_json(sales_file(sale_company_id), []):
            if is_excluded_group_sale(sale_company_id, sale.get("CL_NOM")):
                continue
            client_id = normalize_identifier(sale.get("ID_CL"))
            sale_date = record_date_value(sale.get("NF_EMI"))
            sale_year = record_year(sale.get("NF_EMI"))
            sale_month = record_month(sale.get("NF_EMI"))
            if not sale_date or not sale_year or not sale_month:
                continue
            if client_id in region_client_ids:
                sales.append((sale_date, client_id, sale_year, sale_month, sale_company_id, sale))
                continue
            if vendor_name_key and normalize_text(sale.get("VN_NOM")) == vendor_name_key:
                other_region_sales.append((sale_date, client_id, sale_year, sale_month, sale_company_id, sale))

    first_purchase_by_economic_group = {}
    previous_purchase_by_client = {}
    reactivated_by_month = {str(month): set() for month in range(1, 13)}
    active_clients_by_month = {str(month): set() for month in range(1, 13)}
    new_clients_by_month = {str(month): set() for month in range(1, 13)}
    product_previous_sale = {}
    slow_item_percentages = {
        sale_company_id: load_json(slow_items_file(sale_company_id), {})
        for sale_company_id in vendor_sales_company_ids(company_id)
        if sale_company_id in COMPANIES
    }
    sem_giro_commission_by_month = {str(month): 0.0 for month in range(1, 13)}

    for _sale_date, _client_id, sale_year, sale_month, _sale_company_id, sale in other_region_sales:
        if sale_year == year and 1 <= sale_month <= 12:
            achievements[str(sale_month)]["vendas_liquidas_outras_regioes"] = (
                achievements[str(sale_month)].get("vendas_liquidas_outras_regioes", 0.0)
                + sale_net_revenue(sale)
            )

    for sale_date, client_id, sale_year, sale_month, sale_company_id, sale in sorted(sales, key=lambda item: (item[0], item[1])):
        client = client_index.get(client_id)
        group_key = economic_client_key(client_id, client, sale)
        ref_key = normalize_text(sale.get("PR_COD"))
        first_purchase_by_economic_group.setdefault(group_key, sale_date)
        if sale_year != year or not 1 <= sale_month <= 12:
            previous_purchase_by_client[client_id] = sale_date
            if ref_key:
                product_previous_sale[ref_key] = sale_date
            continue

        month_key = str(sale_month)
        net_revenue = sale_net_revenue(sale)
        achievements[month_key]["vendas_liquidas"] += net_revenue
        active_clients_by_month[month_key].add(group_key)
        previous_purchase = previous_purchase_by_client.get(client_id)
        if previous_purchase and (sale_date - previous_purchase).days > 180:
            reactivated_by_month[month_key].add(client_id)
        if first_purchase_by_economic_group.get(group_key) == sale_date and not is_excluded_group_sale(sale_company_id, sale.get("CL_NOM")):
            new_clients_by_month[month_key].add(group_key)

        previous_product_sale = product_previous_sale.get(ref_key)
        if ref_key and previous_product_sale and previous_product_sale < sale_date and (sale_date - previous_product_sale).days > 365:
            achievements[month_key]["sem_giro"] += net_revenue
            item_config = slow_item_percentages.get(sale_company_id, {}).get(ref_key, {})
            sem_giro_commission_by_month[month_key] += net_revenue * max(0.0, number_value(item_config.get("percentual"))) / 100

        previous_purchase_by_client[client_id] = sale_date
        if ref_key:
            product_previous_sale[ref_key] = sale_date

    for month in range(1, 13):
        month_key = str(month)
        month_record = goal_record.get("months", {}).get(month_key, {})
        objectives = month_record.get("objetivos") if isinstance(month_record.get("objetivos"), dict) else {}
        new_client_goal = objectives.get("novos_clientes") if isinstance(objectives.get("novos_clientes"), dict) else {}
        new_client_value = optional_number_value(new_client_goal.get("meta")) or optional_number_value(month_record.get("valor_novo_cliente")) or 0.0
        achievements[month_key]["reativacao_inativos"] = len(reactivated_by_month[month_key])
        achievements[month_key]["clientes_com_vendas"] = len(active_clients_by_month[month_key])
        achievements[month_key]["contatos_inativos_nunca"] = vendor_day_by_day_monthly_count(company_id, vendor_id_value, year, month)
        achievements[month_key]["novos_clientes"] = len(new_clients_by_month[month_key]) * new_client_value
        achievements[month_key]["sem_giro_comissao"] = sem_giro_commission_by_month[month_key]
        achievements[month_key]["vendas_liquidas_total"] = (
            number_value(achievements[month_key].get("vendas_liquidas"))
            + number_value(achievements[month_key].get("vendas_liquidas_outras_regioes"))
        )
        for key in achievements[month_key]:
            achievements[month_key][key] = round(number_value(achievements[month_key][key]), 2)
    return achievements


def normalize_vendor_goal_record(record: dict, company_id: str, vendor_id_value: str, year: int):
    normalized = default_vendor_goal_record(company_id, vendor_id_value, year)
    normalized["comissao_percentual"] = optional_number_value(record.get("comissao_percentual")) or 1.0
    months = record.get("months") if isinstance(record.get("months"), dict) else {}
    for month in range(1, 13):
        month_key = str(month)
        source_month = months.get(month_key) if isinstance(months.get(month_key), dict) else {}
        target_month = normalized["months"][month_key]
        target_month["percentual_aumento"] = optional_number_value(source_month.get("percentual_aumento")) or 0.0
        target_month["quantidade_minima_clientes"] = optional_number_value(source_month.get("quantidade_minima_clientes")) or 0.0
        target_month["contatos_diarios_inativos"] = optional_number_value(source_month.get("contatos_diarios_inativos")) or target_month["contatos_diarios_inativos"]
        target_month["valor_novo_cliente"] = optional_number_value(source_month.get("valor_novo_cliente")) or target_month["valor_novo_cliente"]
        source_objectives = source_month.get("objetivos") if isinstance(source_month.get("objetivos"), dict) else {}
        for key, _label, _kind in VENDOR_GOAL_OBJECTIVES:
            source_objective = source_objectives.get(key) if isinstance(source_objectives.get(key), dict) else {}
            default_objective = target_month["objetivos"].get(key, {"meta": 0.0, "peso": 0.0})
            source_meta = optional_number_value(source_objective.get("meta"))
            source_weight = optional_number_value(source_objective.get("peso"))
            default_meta = number_value(default_objective.get("meta"))
            default_weight = number_value(default_objective.get("peso"))
            if source_meta == 0 and default_meta > 0:
                source_meta = default_meta
            if source_weight == 0 and default_weight > 0:
                source_weight = default_weight
            target_month["objetivos"][key] = {
                "meta": source_meta if source_meta is not None else default_meta,
                "peso": source_weight if source_weight is not None else default_weight,
            }
        target_month["valor_novo_cliente"] = number_value(target_month["objetivos"]["novos_clientes"].get("meta"))
        target_month["contatos_diarios_inativos"] = number_value(target_month["objetivos"]["contatos_inativos_nunca"].get("meta")) or target_month["contatos_diarios_inativos"]
        target_month["objetivos"]["sem_giro"]["peso"] = 0.0
    return normalized


def vendor_goals_payload(company_id: str, vendor_id_value: str = "", year_value=None):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    year = int(optional_number_value(year_value) or CURRENT_YEAR)
    vendors, _inserted = sync_vendors_from_sales(company_id)
    active_vendors = [vendor for vendor in vendors if vendor.get("status") == "Ativo"]
    goal_vendors = [vendor_goal_option_record(vendor) for vendor in active_vendors]
    if not vendor_id_value and goal_vendors:
        vendor_id_value = goal_vendors[0].get("id")
    vendor = next((item for item in active_vendors if item.get("id") == vendor_id_value), None)
    if vendor_id_value and not vendor:
        raise ValueError("Vendedor ativo nao encontrado.")

    all_goals = load_json(vendor_goals_file(company_id), {})
    stored = all_goals.get(vendor_goals_key(vendor_id_value, year), {}) if vendor_id_value else {}
    record = normalize_vendor_goal_record(stored, company_id, vendor_id_value, year) if vendor_id_value else default_vendor_goal_record(company_id, "", year)
    sales_base = vendor_sales_goal_base(company_id, vendor_id_value, year) if vendor_id_value else vendor_sales_goal_base(company_id, "", year)
    client_sales_base = vendor_client_sales_goal_base(company_id, vendor_id_value, year) if vendor_id_value else vendor_client_sales_goal_base(company_id, "", year)
    for month in range(1, 13):
        month_record = record["months"][str(month)]
        increase = optional_number_value(month_record.get("percentual_aumento")) or 0.0
        suggested_goal = round(sales_base["media_base"] * (1 + increase / 100), 2)
        month_record["objetivos"]["vendas_liquidas"]["meta"] = suggested_goal
        month_record["meta_vendas_liquidas_sugerida"] = suggested_goal
        suggested_clients = int(client_sales_base.get("quantidade_sugerida") or 0)
        current_clients_goal = number_value(month_record["objetivos"]["clientes_com_vendas"].get("meta"))
        if current_clients_goal <= 0 and suggested_clients > 0:
            month_record["objetivos"]["clientes_com_vendas"]["meta"] = suggested_clients
        month_record["meta_clientes_com_vendas_sugerida"] = suggested_clients
    achievements = vendor_goal_achievements(company_id, vendor_id_value, year, record) if vendor_id_value else {}
    for month in range(1, 13):
        month_key = str(month)
        month_record = record["months"][month_key]
        for key, _label, _kind in VENDOR_GOAL_OBJECTIVES:
            achieved = number_value((achievements.get(month_key) or {}).get(key))
            meta = number_value(month_record["objetivos"][key].get("meta"))
            if key == "vendas_liquidas":
                achieved_total = number_value((achievements.get(month_key) or {}).get("vendas_liquidas_total"))
                achieved_other_regions = number_value((achievements.get(month_key) or {}).get("vendas_liquidas_outras_regioes"))
                percent_total = (achieved_total / meta * 100) if meta else 0.0
                month_record["objetivos"][key]["valor_atingido"] = round(achieved, 2)
                month_record["objetivos"][key]["valor_atingido_outras_regioes"] = round(achieved_other_regions, 2)
                month_record["objetivos"][key]["valor_atingido_total"] = round(achieved_total, 2)
                month_record["objetivos"][key]["percentual_atingido"] = round(max(0.0, min((achieved / meta * 100) if meta else 0.0, 999.0)), 2)
                month_record["objetivos"][key]["percentual_atingido_total"] = round(max(0.0, min(percent_total, 999.0)), 2)
                continue
            if key in {"novos_clientes", "sem_giro"}:
                percent = 100.0 if achieved > 0 else 0.0
            else:
                percent = (achieved / meta * 100) if meta else 0.0
            month_record["objetivos"][key]["valor_atingido"] = round(achieved, 2)
            month_record["objetivos"][key]["percentual_atingido"] = round(max(0.0, min(percent, 999.0)), 2)
            if key == "sem_giro":
                month_record["objetivos"][key]["comissao_valor"] = round(number_value((achievements.get(month_key) or {}).get("sem_giro_comissao")), 2)

    return {
        "empresa": COMPANIES[company_id],
        "empresa_id": company_id,
        "year": year,
        "years": list(range(START_YEAR, CURRENT_YEAR + 2)),
        "vendors": goal_vendors,
        "vendedor": vendor_goal_option_record(vendor) if vendor else None,
        "objectives": [{"key": key, "label": label, "kind": kind} for key, label, kind in VENDOR_GOAL_OBJECTIVES],
        "sales_base": sales_base,
        "client_sales_base": client_sales_base,
        "goals": record,
    }


def save_vendor_goals_payload(payload: dict):
    company_id = payload.get("company", "")
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    actor_role = payload.get("actor_role", "vendedor")
    if not vendor_can_manage(actor_role):
        raise ValueError("Somente Usuario Master ou Vendedor Lider/Supervisor pode alterar metas.")
    vendor_id_value = payload.get("vendor_id", "")
    year = int(optional_number_value(payload.get("year")) or CURRENT_YEAR)
    vendors, _inserted = sync_vendors_from_sales(company_id)
    if not any(vendor.get("id") == vendor_id_value for vendor in vendors):
        raise ValueError("Vendedor nao encontrado.")

    record = normalize_vendor_goal_record({
        "comissao_percentual": payload.get("comissao_percentual"),
        "months": payload.get("months") if isinstance(payload.get("months"), dict) else {},
    }, company_id, vendor_id_value, year)
    record["updated_at"] = datetime.now().isoformat(timespec="seconds")
    all_goals = load_json(vendor_goals_file(company_id), {})
    all_goals[vendor_goals_key(vendor_id_value, year)] = record
    save_json(vendor_goals_file(company_id), all_goals)
    return {"message": "Metas e objetivos salvos com sucesso.", "goals": record}


def record_key(company_id: str, record: dict) -> str:
    id_nf = record.get("ID_NF")
    id_prnf = record.get("ID_PRNF")
    if id_nf not in (None, "") and id_prnf not in (None, ""):
        return f"{company_id}|ID_NF:{id_nf}|ID_PRNF:{id_prnf}"

    parts = [
        company_id,
        str(record.get("NF_SER") or ""),
        str(record.get("NF_NUM") or ""),
        str(record.get("NF_EMI") or ""),
        str(record.get("PR_COD") or ""),
        str(record.get("PR_DES") or ""),
    ]
    return "|".join(parts)


def client_key(company_id: str, record: dict) -> str:
    cgc = normalize_identifier(record.get("CGC"))
    if cgc:
        return f"{company_id}|CGC:{cgc}"

    client_id = normalize_identifier(record.get("ID"))
    if client_id:
        return f"{company_id}|ID:{client_id}"

    parts = [
        company_id,
        str(record.get("NOM") or "").strip().upper(),
        str(record.get("UF") or "").strip().upper(),
        normalize_identifier(record.get("CEP")),
    ]
    return "|".join(parts)


def vendor_key(company_id: str, name: str) -> str:
    return f"{company_id}|VENDEDOR:{normalize_text(name)}"


def vendor_id(company_id: str, name: str) -> str:
    key = normalize_text(name).lower() or "vendedor"
    return f"{company_id}-{key[:48]}"


def user_label_by_id(user_id: str) -> str:
    user_id = str(user_id or "").strip()
    if not user_id:
        return ""
    user = next((item for item in load_users() if str(item.get("id") or "") == user_id), None)
    if not user:
        return user_id
    name = user.get("nome") or user.get("login") or user_id
    login = user.get("login") or ""
    return f"{user_id} - {name}" + (f" ({login})" if login and login != name else "")


def is_client_record(record: dict) -> bool:
    category = str(record.get("CAT") or "").strip().casefold()
    account_type = str(record.get("CC_DES") or "").strip().casefold()
    return category == "clientes" and account_type == "cliente"


def vendor_public_record(record: dict) -> dict:
    allowed = [
        company_id
        for company_id in record.get("empresas_acesso", [])
        if company_id in COMPANIES
    ]
    payload = dict(record)
    payload.pop("senha_hash", None)
    payload.pop("senha_acesso", None)
    payload.setdefault("tipo_usuario", "Vendedor")
    payload.setdefault("clientes_atendidos", {"ufs": [], "cidades": [], "clientes_especificos": []})
    payload["empresas_acesso"] = allowed
    payload["empresas_acesso_nomes"] = ", ".join(COMPANIES[company_id] for company_id in allowed)
    payload["usuario_vinculado_id"] = str(record.get("usuario_vinculado_id") or "").strip()
    payload["usuario_vinculado_nome"] = user_label_by_id(payload["usuario_vinculado_id"])
    payload["acesso_configurado"] = "Configurado" if record.get("senha_hash") or record.get("senha_acesso") else "Pendente"
    page_company = record.get("_empresa_id") or (allowed[0] if allowed else "")
    payload["pagina_vendedor"] = f"/vendedor/{page_company}/{record.get('id')}"
    return payload


def sync_vendors_from_sales(company_id: str):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    vendors = load_json(vendors_file(company_id), [])
    existing_keys = {
        item.get("_dedupe_key") or vendor_key(company_id, item.get("nome_completo", ""))
        for item in vendors
    }
    now = datetime.now().isoformat(timespec="seconds")
    inserted = 0

    for sale in load_json(sales_file(company_id), []):
        name = str(sale.get("VN_NOM") or "").strip()
        if not name:
            continue

        key = vendor_key(company_id, name)
        if key in existing_keys:
            continue

        vendors.append({
            "id": vendor_id(company_id, name),
            "nome_completo": name,
            "telefone": "",
            "whatsapp": "",
            "email": "",
            "foto_vendedor": "",
            "login_acesso": "",
            "usuario_vinculado_id": "",
            "tipo_usuario": "Vendedor",
            "status": "Inativo",
            "empresas_acesso": [company_id],
            "clientes_atendidos": {"ufs": [], "cidades": [], "clientes_especificos": []},
            "origem": "Notas fiscais",
            "_empresa_id": company_id,
            "_empresa_nome": COMPANIES[company_id],
            "_dedupe_key": key,
            "_criado_em": now,
            "_atualizado_em": now,
        })
        existing_keys.add(key)
        inserted += 1

    if inserted:
        vendors.sort(key=lambda item: normalize_text(item.get("nome_completo")))
        save_json(vendors_file(company_id), vendors)

    return vendors, inserted


def valid_clients_for_assignment(company_id: str):
    blocked = blocked_client_map(company_id)
    rows = []
    for client in load_json(clients_file(company_id), []):
        client_id = normalize_identifier(client.get("ID"))
        if not client_id or client_id in blocked or is_group_company_client(client):
            continue
        rows.append(client)
    return rows


def vendor_assignment_options_payload(company_id: str):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    vendors, _inserted = sync_vendors_from_sales(company_id)
    assigned_ufs = {}
    assigned_cities = []
    for vendor in vendors:
        for uf in (vendor.get("clientes_atendidos") or {}).get("ufs", []):
            uf_key = str(uf or "").strip().upper()
            if not uf_key:
                continue
            assigned_ufs.setdefault(uf_key, []).append(vendor.get("nome_completo") or "")
        for city in (vendor.get("clientes_atendidos") or {}).get("cidades", []):
            uf_key = str(city.get("uf") or "").strip().upper()
            city_name = str(city.get("cidade") or "").strip()
            if not uf_key or not city_name:
                continue
            assigned_cities.append({
                "vendor_id": vendor.get("id"),
                "vendor_name": vendor.get("nome_completo") or "",
                "uf": uf_key,
                "cidade": city_name,
                "ddds": city.get("ddds") or [],
            })

    cities = {}
    for client in valid_clients_for_assignment(company_id):
        uf = str(client.get("UF") or "").strip().upper()
        city = str(client.get("CID") or "").strip()
        if not uf or not city:
            continue
        key = f"{uf}|{normalize_text(city)}"
        entry = cities.setdefault(key, {"uf": uf, "cidade": city, "ddds": set(), "clientes": 0})
        ddd = client_ddd(client)
        if ddd:
            entry["ddds"].add(ddd)
        entry["clientes"] += 1

    city_rows = []
    for entry in cities.values():
        ddds = sorted(entry["ddds"])
        city_rows.append({
            "uf": entry["uf"],
            "cidade": entry["cidade"],
            "ddds": ddds,
            "clientes": entry["clientes"],
            "exige_ddd": len(ddds) > 1,
        })

    city_rows.sort(key=lambda item: (item["uf"], normalize_text(item["cidade"])))
    return {
        "ufs": BRAZIL_UFS,
        "assigned_ufs": [
            {"uf": uf, "vendors": sorted(name for name in names if name)}
            for uf, names in sorted(assigned_ufs.items())
        ],
        "assigned_cities": assigned_cities,
        "cities": city_rows,
    }


def vendor_client_search_payload(company_id: str, query: str, limit: int = 20):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    normalized_query = normalize_text(query)
    matches = []
    for client in valid_clients_for_assignment(company_id):
        client_id = normalize_identifier(client.get("ID"))
        client_name = str(client.get("NOM") or "")
        if normalized_query and not search_text_matches(query, f"{client_id} {client_name}"):
            continue
        matches.append({
            "id": client_id,
            "name": client_name,
            "uf": client.get("UF"),
            "city": client.get("CID"),
            "ddd": client_ddd(client),
            "label": f"{client_id} - {client_name}",
        })
        if len(matches) >= limit:
            break
    return {"clients": matches}


def region_client_label(client: dict) -> str:
    return f"{normalize_identifier(client.get('ID'))} - {client.get('NOM') or ''}".strip()


def region_options_payload(company_id: str):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    vendors, _inserted = sync_vendors_from_sales(company_id)
    clients = valid_clients_for_assignment(company_id)
    cities = {}
    ddds_by_uf = {}
    client_rows = []
    clients_by_region = {}
    for client in clients:
        client_id = normalize_identifier(client.get("ID"))
        client_name = client.get("NOM") or ""
        uf = str(client.get("UF") or "").strip().upper()
        city = str(client.get("CID") or "").strip()
        ddd = client_ddd(client)
        if uf:
            if ddd:
                ddds_by_uf.setdefault(uf, set()).add(ddd)
            if city:
                key = f"{uf}|{normalize_text(city)}"
                cities.setdefault(key, {"uf": uf, "cidade": city, "ddds": set(), "clientes": 0})
                if ddd:
                    cities[key]["ddds"].add(ddd)
                cities[key]["clientes"] += 1
                region_key = f"{uf}|{normalize_text(city)}|{ddd}"
                clients_by_region.setdefault(region_key, []).append({
                    "id": client_id,
                    "nome": client_name,
                    "uf": uf,
                    "cidade": city,
                    "ddd": ddd,
                    "label": region_client_label(client),
                })
        client_rows.append({
            "id": client_id,
            "nome": client_name,
            "uf": uf,
            "cidade": city,
            "ddd": ddd,
            "label": region_client_label(client),
        })
    city_rows = [
        {
            "uf": row["uf"],
            "cidade": row["cidade"],
            "ddds": sorted(row["ddds"]) + ([""] if clients_by_region.get(f"{row['uf']}|{normalize_text(row['cidade'])}|") else []),
            "clientes_por_ddd": {
                ddd: len(clients_by_region.get(f"{row['uf']}|{normalize_text(row['cidade'])}|{ddd}", []))
                for ddd in sorted(row["ddds"]) + ([""] if clients_by_region.get(f"{row['uf']}|{normalize_text(row['cidade'])}|") else [])
            },
            "clientes": row["clientes"],
        }
        for row in cities.values()
    ]
    city_rows.sort(key=lambda item: (item["uf"], normalize_text(item["cidade"])))
    rules = load_json(region_assignments_file(company_id), [])
    vendor_names = {vendor.get("id"): vendor.get("nome_completo") or "" for vendor in vendors}
    for rule in rules:
        rule["vendor_name"] = vendor_names.get(rule.get("vendor_id"), "")
    return {
        "empresa": COMPANIES[company_id],
        "empresa_id": company_id,
        "vendors": [
            {
                "id": vendor.get("id"),
                "nome_completo": vendor.get("nome_completo") or "",
                "status": vendor.get("status") or "Inativo",
            }
            for vendor in vendors
            if vendor.get("status") == "Ativo"
        ],
        "ufs": BRAZIL_UFS,
        "ddds_by_uf": {uf: sorted(values) for uf, values in sorted(ddds_by_uf.items())},
        "cities": city_rows,
        "clients_by_region": {
            key: sorted(rows, key=lambda item: normalize_text(item["label"]))
            for key, rows in sorted(clients_by_region.items())
        },
        "clients": sorted(client_rows, key=lambda item: normalize_text(item["label"]))[:2000],
        "rules": sorted(rules, key=lambda item: (item.get("uf") or "", item.get("tipo") or "", normalize_text(item.get("label") or ""))),
    }


def normalize_region_rule(payload: dict, company_id: str) -> dict:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    rule_type = str(payload.get("tipo") or "").strip().lower()
    if rule_type not in REGION_PRIORITY:
        raise ValueError("Selecione o tipo de regiao.")
    vendor_id_value = str(payload.get("vendor_id") or "").strip()
    vendors, _inserted = sync_vendors_from_sales(company_id)
    vendor = next((item for item in vendors if item.get("id") == vendor_id_value), None)
    if not vendor:
        raise ValueError("Selecione um vendedor valido.")
    uf = str(payload.get("uf") or "").strip().upper()
    ddd = re.sub(r"\D", "", str(payload.get("ddd") or ""))[:2]
    city = str(payload.get("cidade") or "").strip()
    client_id = normalize_identifier(payload.get("cliente_id"))
    label = ""
    if rule_type in ("uf", "ddd", "cidade", "exclusao_ddd", "exclusao_cidade") and uf not in BRAZIL_UFS:
        raise ValueError("Selecione uma UF valida.")
    if rule_type in ("ddd", "exclusao_ddd") and not ddd:
        raise ValueError("Selecione o DDD.")
    if rule_type in ("cidade", "exclusao_cidade") and not city:
        raise ValueError("Selecione a cidade.")
    if rule_type in ("cliente", "exclusao_cliente"):
        clients = {normalize_identifier(client.get("ID")): client for client in valid_clients_for_assignment(company_id)}
        client = clients.get(client_id)
        if not client:
            raise ValueError("Selecione um cliente valido.")
        uf = str(client.get("UF") or "").strip().upper()
        city = str(client.get("CID") or "").strip()
        ddd = client_ddd(client)
        label = region_client_label(client)
    elif rule_type == "exclusao_cidade":
        label = f"Excluir {city} - {uf}{f' / DDD {ddd}' if ddd else ''}"
    elif rule_type == "exclusao_ddd":
        label = f"Excluir {uf} / DDD {ddd}"
    elif rule_type == "cidade":
        label = f"{city} - {uf}"
    elif rule_type == "ddd":
        label = f"{uf} / DDD {ddd}"
    else:
        label = uf
    return {
        "id": str(payload.get("id") or "").strip() or f"{rule_type}-{vendor_id_value}-{datetime.now().strftime('%Y%m%d%H%M%S%f')}",
        "tipo": rule_type,
        "vendor_id": vendor_id_value,
        "uf": uf,
        "ddd": ddd,
        "cidade": city,
        "cliente_id": client_id,
        "label": label,
        "status": "Ativo",
    }


def region_rule_identity(rule: dict) -> tuple:
    return (
        str(rule.get("tipo") or "").strip().lower(),
        str(rule.get("vendor_id") or "").strip(),
        str(rule.get("uf") or "").strip().upper(),
        re.sub(r"\D", "", str(rule.get("ddd") or ""))[:2],
        normalize_text(rule.get("cidade")),
        normalize_identifier(rule.get("cliente_id")),
    )


def region_rule_scope_identity(rule: dict) -> tuple:
    return (
        str(rule.get("tipo") or "").strip().lower(),
        str(rule.get("uf") or "").strip().upper(),
        re.sub(r"\D", "", str(rule.get("ddd") or ""))[:2],
        normalize_text(rule.get("cidade")),
        normalize_identifier(rule.get("cliente_id")),
    )


def save_region_rule_payload(payload: dict):
    company_id = str(payload.get("company") or "ionlab")
    rule = normalize_region_rule(payload, company_id)
    rules = load_json(region_assignments_file(company_id), [])
    now = datetime.now().isoformat(timespec="seconds")
    rule_type = str(rule.get("tipo") or "")
    rule_identity = region_rule_identity(rule)
    rule_scope_identity = region_rule_scope_identity(rule)
    existing = next(
        (
            item
            for item in rules
            if item.get("id") == rule["id"]
            or (
                region_rule_identity(item) == rule_identity
                if rule_type.startswith("exclusao_")
                else region_rule_scope_identity(item) == rule_scope_identity
            )
        ),
        None,
    )
    if existing:
        existing.update(rule)
        existing["_atualizado_em"] = now
    else:
        rule["_criado_em"] = now
        rule["_atualizado_em"] = now
        rules.append(rule)
    save_json(region_assignments_file(company_id), rules)
    return {"message": "Regiao salva com sucesso.", **region_options_payload(company_id)}


def delete_region_rule_payload(company_id: str, rule_id: str):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    rules = load_json(region_assignments_file(company_id), [])
    next_rules = [rule for rule in rules if rule.get("id") != rule_id]
    save_json(region_assignments_file(company_id), next_rules)
    return {"message": "Regiao removida com sucesso.", **region_options_payload(company_id)}


def normalize_vendor_assignments(payload: dict, company_id: str) -> dict:
    valid_client_ids = {
        normalize_identifier(client.get("ID")): client
        for client in valid_clients_for_assignment(company_id)
    }
    city_options = vendor_assignment_options_payload(company_id)["cities"]
    city_option_map = {
        f"{item['uf']}|{normalize_text(item['cidade'])}": item
        for item in city_options
    }

    raw = payload.get("clientes_atendidos") or {}
    ufs = sorted({
        str(uf or "").strip().upper()
        for uf in raw.get("ufs", [])
        if str(uf or "").strip().upper() in BRAZIL_UFS
    })

    cities = []
    for city in raw.get("cidades", []):
        uf = str(city.get("uf") or "").strip().upper()
        if ufs and uf not in ufs:
            continue
        city_name = str(city.get("cidade") or "").strip()
        key = f"{uf}|{normalize_text(city_name)}"
        option = city_option_map.get(key)
        if not option:
            continue
        available_ddds = option.get("ddds") or []
        selected_ddds = sorted({
            re.sub(r"\D", "", str(ddd or ""))[:2]
            for ddd in city.get("ddds", [])
            if re.sub(r"\D", "", str(ddd or ""))[:2]
        })
        selected_ddds = [ddd for ddd in selected_ddds if ddd in available_ddds]
        if len(available_ddds) == 1 and not selected_ddds:
            selected_ddds = available_ddds
        cities.append({
            "uf": option["uf"],
            "cidade": option["cidade"],
            "ddds": selected_ddds,
        })

    specific_clients = []
    seen_clients = set()
    for item in raw.get("clientes_especificos", []):
        client_id = normalize_identifier(item.get("id") if isinstance(item, dict) else item)
        client = valid_client_ids.get(client_id)
        if not client or client_id in seen_clients:
            continue
        seen_clients.add(client_id)
        specific_clients.append({
            "id": client_id,
            "name": client.get("NOM"),
            "uf": client.get("UF"),
            "city": client.get("CID"),
            "ddd": client_ddd(client),
        })

    return {
        "ufs": ufs,
        "cidades": cities,
        "clientes_especificos": specific_clients,
    }


def vendor_city_assignment_keys(assignments: dict) -> set:
    keys = set()
    for city in assignments.get("cidades", []):
        uf = str(city.get("uf") or "").strip().upper()
        city_name = normalize_text(city.get("cidade"))
        ddds = city.get("ddds") or [""]
        for ddd in ddds:
            keys.add(f"{uf}|{city_name}|{ddd}")
    return keys


def validate_vendor_assignments(vendors: list, current_id: str, assignments: dict):
    new_keys = vendor_city_assignment_keys(assignments)
    if not new_keys:
        return
    for vendor in vendors:
        if vendor.get("id") == current_id:
            continue
        existing_keys = vendor_city_assignment_keys(vendor.get("clientes_atendidos") or {})
        conflict = new_keys & existing_keys
        if conflict:
            sample = sorted(conflict)[0].split("|")
            ddd_text = f" e DDD {sample[2]}" if sample[2] else ""
            raise ValueError(
                f"Ja existe vendedor atendendo {sample[1]}-{sample[0]}{ddd_text}. "
                "A mesma cidade e DDD nao pode ficar com mais de um vendedor."
            )


def vendors_payload(company_id: str, query: str = ""):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    vendors, inserted = sync_vendors_from_sales(company_id)
    matches = [
        vendor_public_record(vendor)
        for vendor in vendors
        if value_matches(vendor_public_record(vendor), query)
    ]

    return {
        "empresa": COMPANIES[company_id],
        "empresa_id": company_id,
        "novos_desde_notas": inserted,
        "total_registros": len(vendors),
        "total_filtrado": len(matches),
        "ativos": sum(1 for item in vendors if item.get("status") == "Ativo"),
        "inativos": sum(1 for item in vendors if item.get("status") != "Ativo"),
        "columns": [{"key": key, "label": label} for key, label in VENDOR_TABLE_COLUMNS],
        "rows": matches[:500],
        "companies": [{"id": key, "name": value} for key, value in COMPANIES.items()],
    }


def vendor_page_links_payload():
    links = []
    for company_id, company_name in COMPANIES.items():
        try:
            vendors, _inserted = sync_vendors_from_sales(company_id)
        except Exception:
            continue
        for vendor in vendors:
            if vendor.get("status") != "Ativo":
                continue
            public = vendor_public_record(vendor)
            page = public.get("pagina_vendedor") or f"/vendedor/{company_id}/{vendor.get('id')}"
            links.append({
                "empresa": company_name,
                "empresa_id": company_id,
                "id": vendor.get("id"),
                "nome_completo": vendor.get("nome_completo") or "Vendedor",
                "pagina_vendedor": page,
                "usuario_vinculado_id": public.get("usuario_vinculado_id", ""),
                "usuario_vinculado_nome": public.get("usuario_vinculado_nome", ""),
                "login_acesso": public.get("login_acesso", ""),
                "tipo_usuario": public.get("tipo_usuario", "Vendedor"),
            })
    links.sort(key=lambda item: (normalize_text(item.get("empresa")), normalize_text(item.get("nome_completo"))))
    return {
        "total": len(links),
        "rows": links,
    }


def save_vendor_payload(payload: dict):
    company_id = str(payload.get("company") or "")
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    actor_role = str(payload.get("actor_role") or "vendedor")
    if not vendor_can_manage(actor_role):
        raise ValueError("Somente Usuario Master ou Vendedor Lider/Supervisor pode incluir ou alterar vendedores.")

    name = str(payload.get("nome_completo") or "").strip()
    if not name:
        raise ValueError("Informe o nome completo do vendedor.")

    status = "Ativo" if payload.get("status") == "Ativo" else "Inativo"
    allowed = [
        company_id
        for company_id in payload.get("empresas_acesso", [])
        if company_id in COMPANIES
    ] or [company_id]
    user_type = (
        "Vendedor Lider/Supervisor"
        if payload.get("tipo_usuario") == "Vendedor Lider/Supervisor"
        else "Vendedor"
    )

    vendors, _inserted = sync_vendors_from_sales(company_id)
    record_id = str(payload.get("id") or "").strip()
    key = vendor_key(company_id, name)
    now = datetime.now().isoformat(timespec="seconds")
    target = None

    if record_id:
        target = next((item for item in vendors if item.get("id") == record_id), None)
    if target is None:
        target = next((item for item in vendors if (item.get("_dedupe_key") or vendor_key(company_id, item.get("nome_completo", ""))) == key), None)

    if target is None:
        target = {
            "id": vendor_id(company_id, name),
            "_empresa_id": company_id,
            "_empresa_nome": COMPANIES[company_id],
            "_criado_em": now,
        }
        vendors.append(target)

    assignments = normalize_vendor_assignments(payload, company_id)
    validate_vendor_assignments(vendors, target.get("id"), assignments)

    target.update({
        "nome_completo": name,
        "telefone": str(payload.get("telefone") or "").strip(),
        "whatsapp": str(payload.get("whatsapp") or "").strip(),
        "email": str(payload.get("email") or "").strip(),
        "foto_vendedor": str(payload.get("foto_vendedor") or target.get("foto_vendedor") or "").strip(),
        "login_acesso": str(payload.get("login_acesso") or "").strip(),
        "usuario_vinculado_id": str(payload.get("usuario_vinculado_id") or "").strip(),
        "tipo_usuario": user_type,
        "status": status,
        "empresas_acesso": allowed,
        "clientes_atendidos": assignments,
        "origem": target.get("origem") or "Cadastro manual",
        "_dedupe_key": key,
        "_atualizado_em": now,
    })
    password = str(payload.get("senha_acesso") or "").strip()
    if password:
        target["senha_hash"] = hashlib.sha256(password.encode("utf-8")).hexdigest()

    vendors.sort(key=lambda item: normalize_text(item.get("nome_completo")))
    try:
        save_json(vendors_file(company_id), vendors)
    except PermissionError as exc:
        raise ValueError(
            f"Nao foi possivel salvar o vendedor no banco. Feche planilhas, backups ou outro CRM que esteja usando o arquivo {vendors_file(company_id)} e tente novamente."
        ) from exc
    return {
        "empresa": COMPANIES[company_id],
        "message": "Vendedor salvo com sucesso.",
        "vendor": vendor_public_record(target),
    }


ACCESS_LEVELS = ("edit", "view", "blocked")
ACCESS_LEVEL_NAMES = {
    "edit": "Incluir/Alterar",
    "view": "Visualizar",
    "blocked": "Bloqueado",
}
USER_PERMISSION_GROUPS = [
    {
        "id": "dashboard",
        "label": "Pagina Dashboard",
        "items": [
            ("dashboard.uf", "Clientes por UF"),
            ("dashboard.activity", "Clientes Ativos/Inativos por UF"),
            ("dashboard.vendor-regions", "Vendedores X Regioes Atendidas"),
            ("dashboard.annual-sales", "Vendas Anuais"),
            ("dashboard.monthly-evolution", "Evolucao Mensal"),
            ("dashboard.family-sales", "Faturamento por Familia"),
        ],
    },
    {"id": "prospect", "label": "Prospect", "items": [("prospect", "Prospect")]},
    {
        "id": "cadastros",
        "label": "Cadastros",
        "items": [
            ("cadastros.sales", "Notas fiscais de vendas"),
            ("cadastros.clients", "Clientes"),
            ("cadastros.stock", "Estoque"),
            ("cadastros.in-transit", "Em Transito"),
            ("cadastros.slow-items", "Itens/Equipamentos Sem Giro"),
            ("cadastros.prices-import", "Tabela de Precos"),
            ("cadastros.products", "Cadastro de Produtos"),
            ("cadastros.costing", "Custeio Importados"),
            ("cadastros.costing-fabricated", "Custeio Fabricado"),
            ("cadastros.vendors", "Vendedores"),
            ("cadastros.regions", "Regioes"),
        ],
    },
    {"id": "sales-table", "label": "Vendas", "items": [("sales-table", "Vendas")]},
    {"id": "clients-table", "label": "Clientes", "items": [("clients-table", "Clientes")]},
    {"id": "stock-table", "label": "Estoque", "items": [("stock-table", "Estoque")]},
    {"id": "in-transit-table", "label": "Em Transito", "items": [("in-transit-table", "Em Transito")]},
    {"id": "prices-table", "label": "Tabela de Precos", "items": [("prices-table", "Tabela de Precos")]},
    {"id": "products-table", "label": "Cadastro de Produtos", "items": [("products-table", "Cadastro de Produtos")]},
    {"id": "mercado-livre", "label": "Gestao Mercado Livre", "items": [("mercado-livre", "Gestao Mercado Livre")]},
    {"id": "costing-table", "label": "Custeio", "items": [("costing-table", "Custeio")]},
    {"id": "follow-up", "label": "Follow-UP Atendimentos", "items": [("follow-up", "Follow-UP Atendimentos")]},
    {"id": "blocked-clients", "label": "Bloqueios", "items": [("blocked-clients", "Bloqueios")]},
    {"id": "users", "label": "Usuarios", "items": [("users", "Cadastro de usuarios")]},
]
USER_PERMISSION_KEYS = [
    key
    for group in USER_PERMISSION_GROUPS
    for key, _label in group["items"]
]


def password_hash(password: str) -> str:
    return hashlib.sha256(str(password or "").encode("utf-8")).hexdigest()


def normalize_permissions(raw_permissions: dict | None, default_level: str = "blocked") -> dict:
    raw_permissions = raw_permissions if isinstance(raw_permissions, dict) else {}
    normalized = {}
    for key in USER_PERMISSION_KEYS:
        level = str(raw_permissions.get(key) or default_level)
        normalized[key] = level if level in ACCESS_LEVELS else default_level
    return normalized


def default_master_user() -> dict:
    now = datetime.now().isoformat(timespec="seconds")
    return {
        "id": "master",
        "nome": "Usuario Master",
        "login": "master",
        "email": "",
        "senha_hash": password_hash("master"),
        "status": "Ativo",
        "tipo": "master",
        "trocar_senha_primeiro_acesso": False,
        "permissions": normalize_permissions({}, "edit"),
        "_criado_em": now,
        "_atualizado_em": now,
    }


def operational_data_exists() -> bool:
    if not DATA_DIR.exists():
        return False
    protected_names = {
        "clientes.json",
        "vendas.json",
        "vendedores.json",
        "day_by_day_vendedores.json",
        "clientes_bloqueados.json",
        "orcamentos.json",
        "pedidos.json",
    }
    for path in DATA_DIR.rglob("*.json"):
        if path.name not in protected_names:
            continue
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue
        if isinstance(payload, (list, dict)) and len(payload) > 0:
            return True
    return False


def load_users() -> list[dict]:
    users = load_json(users_file(), [])
    if not users:
        if operational_data_exists():
            raise RuntimeError(
                "Arquivo de usuarios ausente ou vazio em uma base com dados. "
                "A atualizacao nao deve substituir a pasta data. Restaure data/usuarios.json do backup do site quente."
            )
        users = [default_master_user()]
        save_json(users_file(), users)
    changed = False
    for user in users:
        default_level = "edit" if user.get("tipo") == "master" else "blocked"
        permissions = normalize_permissions(user.get("permissions"), default_level)
        if permissions != user.get("permissions"):
            user["permissions"] = permissions
            changed = True
    if changed:
        save_json(users_file(), users)
    return users


def user_public_record(user: dict) -> dict:
    return {
        "id": user.get("id"),
        "nome": user.get("nome"),
        "login": user.get("login"),
        "email": user.get("email") or "",
        "status": user.get("status") or "Ativo",
        "tipo": user.get("tipo") or "usuario",
        "trocar_senha_primeiro_acesso": bool(user.get("trocar_senha_primeiro_acesso")),
        "permissions": normalize_permissions(
            user.get("permissions"),
            "edit" if user.get("tipo") == "master" else "blocked",
        ),
        "_atualizado_em": user.get("_atualizado_em"),
    }


def user_catalog_payload():
    return {
        "levels": [{"id": key, "label": ACCESS_LEVEL_NAMES[key]} for key in ACCESS_LEVELS],
        "groups": [
            {
                "id": group["id"],
                "label": group["label"],
                "items": [{"key": key, "label": label} for key, label in group["items"]],
            }
            for group in USER_PERMISSION_GROUPS
        ],
    }


def users_payload(query: str = ""):
    rows = [user_public_record(user) for user in load_users()]
    matches = [row for row in rows if value_matches(row, query)]
    return {
        "total_registros": len(rows),
        "total_filtrado": len(matches),
        "columns": [
            {"key": "nome", "label": "Nome"},
            {"key": "login", "label": "Login"},
            {"key": "email", "label": "E-mail"},
            {"key": "status", "label": "Status"},
            {"key": "tipo", "label": "Tipo"},
        ],
        "rows": sorted(matches, key=lambda item: normalize_text(item.get("nome"))),
        **user_catalog_payload(),
    }


def current_user_from_token(token: str) -> dict | None:
    token = str(token or "").strip()
    if not token:
        return None
    sessions = load_json(auth_sessions_file(), {})
    session = sessions.get(token) if isinstance(sessions, dict) else None
    if not session:
        session = RUNTIME_AUTH_SESSIONS.get(token)
    if not session:
        return None
    user_id = session.get("user_id")
    user = next((item for item in load_users() if item.get("id") == user_id), None)
    if not user or user.get("status") != "Ativo":
        return None
    return user


def auth_login_payload(payload: dict):
    login = str(normalize_record(payload.get("login")) or "").strip()
    password = str(payload.get("password") or "")
    user = next((item for item in load_users() if normalize_text(item.get("login")) == normalize_text(login)), None)
    if not user or user.get("status") != "Ativo" or user.get("senha_hash") != password_hash(password):
        raise ValueError("Login ou senha invalidos.")

    token = hashlib.sha256(f"{user.get('id')}|{datetime.now().isoformat()}|{password}".encode("utf-8")).hexdigest()
    sessions = load_json(auth_sessions_file(), {})
    if not isinstance(sessions, dict):
        sessions = {}
    sessions[token] = {
        "user_id": user.get("id"),
        "login": user.get("login"),
        "created_at": datetime.now().isoformat(timespec="seconds"),
    }
    RUNTIME_AUTH_SESSIONS[token] = sessions[token]
    try:
        save_json(auth_sessions_file(), sessions)
    except PermissionError:
        pass
    return {"token": token, "user": user_public_record(user), "force_password_change": bool(user.get("trocar_senha_primeiro_acesso")), **user_catalog_payload()}


def auth_session_payload(token: str):
    user = current_user_from_token(token)
    if not user:
        return {"authenticated": False, **user_catalog_payload()}
    return {"authenticated": True, "user": user_public_record(user), "force_password_change": bool(user.get("trocar_senha_primeiro_acesso")), **user_catalog_payload()}


def auth_logout_payload(token: str):
    sessions = load_json(auth_sessions_file(), {})
    if isinstance(sessions, dict) and token in sessions:
        sessions.pop(token, None)
        try:
            save_json(auth_sessions_file(), sessions)
        except PermissionError:
            pass
    RUNTIME_AUTH_SESSIONS.pop(token, None)
    return {"message": "Sessao encerrada."}


def change_password_payload(payload: dict, token: str):
    user = current_user_from_token(token)
    if not user:
        raise ValueError("Sessao expirada. Entre novamente.")
    new_password = str(payload.get("nova_senha") or "").strip()
    confirm_password = str(payload.get("confirmar_senha") or "").strip()
    if len(new_password) < 6:
        raise ValueError("A nova senha deve ter pelo menos 6 caracteres.")
    if new_password != confirm_password:
        raise ValueError("A confirmacao da senha nao confere.")
    if user.get("senha_hash") == password_hash(new_password):
        raise ValueError("A nova senha deve ser diferente da senha atual.")
    users = load_users()
    now = datetime.now().isoformat(timespec="seconds")
    for item in users:
        if item.get("id") == user.get("id"):
            item["senha_hash"] = password_hash(new_password)
            item["trocar_senha_primeiro_acesso"] = False
            item["_atualizado_em"] = now
            user = item
            break
    save_json(users_file(), users)
    return {"message": "Senha alterada com sucesso.", "user": user_public_record(user), **user_catalog_payload()}


def require_master_user(token: str):
    user = current_user_from_token(token)
    if not user or user.get("tipo") != "master":
        raise ValueError("Somente Usuario Master pode alterar usuarios.")
    return user


def save_user_payload(payload: dict, token: str):
    require_master_user(token)
    name = str(normalize_record(payload.get("nome")) or "").strip()
    login = str(normalize_record(payload.get("login")) or "").strip()
    if not name:
        raise ValueError("Informe o nome do usuario.")
    if not login:
        raise ValueError("Informe o login do usuario.")

    users = load_users()
    record_id = str(normalize_record(payload.get("id")) or "").strip()
    key = normalize_text(login)
    existing_login = next((item for item in users if normalize_text(item.get("login")) == key and item.get("id") != record_id), None)
    if existing_login:
        raise ValueError("Ja existe usuario com este login.")

    target = next((item for item in users if item.get("id") == record_id), None)
    now = datetime.now().isoformat(timespec="seconds")
    if target is None:
        base_id = normalize_text(login).lower() or f"usuario-{len(users) + 1}"
        candidate = base_id
        suffix = 2
        existing_ids = {item.get("id") for item in users}
        while candidate in existing_ids:
            candidate = f"{base_id}-{suffix}"
            suffix += 1
        target = {"id": candidate, "_criado_em": now}
        users.append(target)

    user_type = "master" if payload.get("tipo") == "master" else "usuario"
    target.update({
        "nome": name,
        "login": login,
        "email": str(normalize_record(payload.get("email")) or "").strip(),
        "status": "Ativo" if payload.get("status") == "Ativo" else "Inativo",
        "tipo": user_type,
        "trocar_senha_primeiro_acesso": bool(payload.get("trocar_senha_primeiro_acesso")),
        "permissions": normalize_permissions(payload.get("permissions"), "edit" if user_type == "master" else "blocked"),
        "_atualizado_em": now,
    })
    password = str(payload.get("senha") or "").strip()
    if password:
        target["senha_hash"] = password_hash(password)
    elif not target.get("senha_hash"):
        raise ValueError("Informe uma senha para o novo usuario.")

    users.sort(key=lambda item: normalize_text(item.get("nome")))
    try:
        save_json(users_file(), users)
    except PermissionError as exc:
        raise ValueError(
            "Nao foi possivel salvar o usuario porque o servidor local esta sem permissao para gravar "
            "data/usuarios.json. Feche a janela do CRM e abra novamente pelo arquivo abrir_crm_atualizado.bat."
        ) from exc
    return {"message": "Usuario salvo com sucesso.", "user": user_public_record(target), **user_catalog_payload()}


def default_block_reasons():
    now = datetime.now().isoformat(timespec="seconds")
    return [
        {"id": "financeiro", "motivo": "Financeiro", "_criado_em": now},
        {"id": "solicitacao-diretoria", "motivo": "Solicitação Diretoria", "_criado_em": now},
        {"id": "empresa-encerrada", "motivo": "Empresa Encerrada", "_criado_em": now},
        {"id": "cliente-bloqueado-consumidor-final", "motivo": "Cliente Bloqueado - Consumidor Final", "_criado_em": now},
        {"id": "cliente-com-atividade-encerrada", "motivo": "Cliente com Atividade Encerrada", "_criado_em": now},
    ]


def load_block_reasons():
    reasons = load_json(block_reasons_file(), [])
    if not reasons:
        reasons = default_block_reasons()
        save_json(block_reasons_file(), reasons)
    required = default_block_reasons()
    existing_keys = {reason_id(item.get("motivo")) for item in reasons}
    changed = False
    for reason in required:
        key = reason_id(reason.get("motivo"))
        if key not in existing_keys:
            reasons.append(reason)
            existing_keys.add(key)
            changed = True
    if changed:
        save_json(block_reasons_file(), sorted(reasons, key=lambda item: normalize_text(item.get("motivo"))))
    return reasons


def reason_id(reason: str) -> str:
    normalized = normalize_text(reason).lower() or "motivo"
    return normalized[:60]


def block_reasons_payload():
    reasons = load_block_reasons()
    return {
        "total_registros": len(reasons),
        "rows": sorted(reasons, key=lambda item: normalize_text(item.get("motivo"))),
    }


def save_block_reason_payload(payload: dict):
    reason = str(payload.get("motivo") or "").strip()
    if not reason:
        raise ValueError("Informe o motivo de bloqueio.")

    reasons = load_block_reasons()
    key = reason_id(reason)
    existing = next((item for item in reasons if item.get("id") == key or normalize_text(item.get("motivo")) == normalize_text(reason)), None)
    if existing is None:
        existing = {
            "id": key,
            "_criado_em": datetime.now().isoformat(timespec="seconds"),
        }
        reasons.append(existing)

    existing["motivo"] = reason
    save_json(block_reasons_file(), sorted(reasons, key=lambda item: normalize_text(item.get("motivo"))))
    return {"message": "Motivo salvo com sucesso.", "reason": existing}


def find_client_by_code(company_id: str, code: str):
    normalized_code = normalize_identifier(code)
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    if not normalized_code:
        raise ValueError("Informe o codigo do cliente.")

    for client in load_json(clients_file(company_id), []):
        if normalize_identifier(client.get("ID")) == normalized_code:
            return client
    return None


def client_lookup_payload(company_id: str, code: str):
    client = find_client_by_code(company_id, code)
    if not client:
        raise ValueError("Cliente nao encontrado nesta empresa.")
    return {
        "codigo_cliente": normalize_identifier(client.get("ID")),
        "nome_cliente": client.get("NOM"),
        "cidade": client.get("CID"),
        "uf": client.get("UF"),
    }


def blocked_client_key(company_id: str, client_code: str) -> str:
    return f"{company_id}|CLIENTE:{normalize_identifier(client_code)}"


def blocked_client_map(company_id: str) -> dict:
    if company_id not in COMPANIES:
        return {}

    blocked = {}
    for record in load_json(blocked_clients_file(company_id), []):
        client_code = normalize_identifier(record.get("codigo_cliente"))
        if not client_code:
            continue
        blocked[client_code] = {
            "codigo_cliente": client_code,
            "nome_cliente": record.get("nome_cliente"),
            "motivo": record.get("motivo"),
            "bloqueado_em": record.get("bloqueado_em"),
            "empresa": COMPANIES.get(company_id, company_id),
        }
    return blocked


def client_block_status_payload(company_id: str, code: str):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    client_code = normalize_identifier(code)
    blocked = blocked_client_map(company_id).get(client_code)
    if blocked:
        return {
            "blocked": True,
            "message": f"Cliente bloqueado. Motivo: {blocked.get('motivo') or 'Nao informado'}.",
            "details": blocked,
        }

    return {
        "blocked": False,
        "message": "Cliente liberado.",
        "details": None,
    }


def blocked_clients_payload(company_id: str, query: str = ""):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    records = load_json(blocked_clients_file(company_id), [])
    rows = []
    for record in records:
        row = dict(record)
        row["empresa_nome"] = COMPANIES.get(company_id, company_id)
        rows.append(row)

    matches = [row for row in rows if value_matches(row, query)]
    return {
        "empresa": COMPANIES[company_id],
        "empresa_id": company_id,
        "total_registros": len(records),
        "total_filtrado": len(matches),
        "columns": [{"key": key, "label": label} for key, label in BLOCKED_CLIENT_COLUMNS],
        "rows": matches[:500],
    }


def save_blocked_client_payload(payload: dict):
    company_id = str(payload.get("company") or "")
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    code = normalize_identifier(payload.get("codigo_cliente"))
    if not code:
        raise ValueError("Informe o codigo do cliente.")

    reason = str(payload.get("motivo") or "").strip()
    valid_reasons = {normalize_text(item.get("motivo")): item.get("motivo") for item in load_block_reasons()}
    if normalize_text(reason) not in valid_reasons:
        raise ValueError("Selecione um motivo valido.")

    client = find_client_by_code(company_id, code)
    if not client:
        raise ValueError("Cliente nao encontrado nesta empresa.")

    records = load_json(blocked_clients_file(company_id), [])
    key = blocked_client_key(company_id, code)
    existing = next((item for item in records if item.get("_dedupe_key") == key), None)
    now = datetime.now().isoformat(timespec="seconds")
    if existing is None:
        existing = {
            "_dedupe_key": key,
            "_empresa_id": company_id,
            "_criado_em": now,
        }
        records.append(existing)

    existing.update({
        "codigo_cliente": code,
        "nome_cliente": client.get("NOM"),
        "motivo": valid_reasons[normalize_text(reason)],
        "bloqueado_em": now,
    })
    records.sort(key=lambda item: normalize_text(item.get("nome_cliente")))
    save_json(blocked_clients_file(company_id), records)
    return {
        "empresa": COMPANIES[company_id],
        "message": "Cliente bloqueado com sucesso.",
        "blocked": existing,
    }


FINAL_CONSUMER_BLOCK_REASON = "Cliente Bloqueado - Consumidor Final"
ACTIVITY_CLOSED_BLOCK_REASON = "Cliente com Atividade Encerrada"


def block_final_consumer_client(company_id: str, client_code: str) -> dict:
    return save_blocked_client_payload({
        "company": company_id,
        "codigo_cliente": client_code,
        "motivo": FINAL_CONSUMER_BLOCK_REASON,
    })


def block_activity_closed_client(company_id: str, client_code: str) -> dict:
    return save_blocked_client_payload({
        "company": company_id,
        "codigo_cliente": client_code,
        "motivo": ACTIVITY_CLOSED_BLOCK_REASON,
    })


def parse_multipart(headers, body: bytes):
    content_type = headers.get("Content-Type", "")
    raw = b"Content-Type: " + content_type.encode("utf-8") + b"\r\n\r\n" + body
    message = BytesParser(policy=policy.default).parsebytes(raw)
    fields = {}
    files = {}

    for part in message.iter_parts():
        disposition = part.get("Content-Disposition", "")
        name_match = re.search(r'name="([^"]+)"', disposition)
        if not name_match:
            continue

        name = name_match.group(1)
        filename_match = re.search(r'filename="([^"]*)"', disposition)
        payload = part.get_payload(decode=True) or b""

        if filename_match:
            files[name] = {
                "filename": filename_match.group(1),
                "content": payload,
                "content_type": part.get_content_type(),
            }
        else:
            fields[name] = payload.decode(part.get_content_charset() or "utf-8").strip()

    return fields, files


def ods_cell_text(cell, namespaces):
    value_type = cell.attrib.get(f"{{{namespaces['office']}}}value-type")
    if value_type in ("float", "currency", "percentage"):
        value = cell.attrib.get(f"{{{namespaces['office']}}}value")
        if value not in (None, ""):
            return value
    if value_type == "date":
        value = cell.attrib.get(f"{{{namespaces['office']}}}date-value")
        if value not in (None, ""):
            return value
    parts = []
    for element in cell.findall(".//text:p", namespaces):
        parts.append("".join(element.itertext()))
    return "\n".join(part for part in parts if part is not None)


def read_ods_rows(file_bytes: bytes):
    namespaces = {
        "office": "urn:oasis:names:tc:opendocument:xmlns:office:1.0",
        "table": "urn:oasis:names:tc:opendocument:xmlns:table:1.0",
        "text": "urn:oasis:names:tc:opendocument:xmlns:text:1.0",
    }
    with zipfile.ZipFile(BytesIO(file_bytes)) as ods:
        root = ET.fromstring(ods.read("content.xml"))
    table = root.find(".//table:table", namespaces)
    if table is None:
        return []
    rows = []
    for row in table.findall("table:table-row", namespaces):
        row_repeat = int(row.attrib.get(f"{{{namespaces['table']}}}number-rows-repeated", "1"))
        values = []
        for cell in row.findall("table:table-cell", namespaces):
            column_repeat = int(cell.attrib.get(f"{{{namespaces['table']}}}number-columns-repeated", "1"))
            value = ods_cell_text(cell, namespaces)
            values.extend([value] * min(column_repeat, 200))
        while values and values[-1] in (None, ""):
            values.pop()
        if values:
            rows.extend([values] * min(row_repeat, 1000))
    return rows


def rows_to_records(rows):
    if not rows:
        return [], ""
    header_index = 0
    for index, row in enumerate(rows[:20]):
        normalized = {normalize_text(value) for value in row}
        if "produto" in normalized or "referencia" in normalized or "descricao" in normalized:
            header_index = index
            break
    def clean_header(value):
        header = str(value or "").strip()
        normalized = normalize_text(header)
        if normalized == "TP" or (normalized.startswith("TIPO") and normalized.endswith("TP")) or "TOTALIZACAO GERAL" in normalized:
            return "Tipo"
        if normalized in ("END. FISICO", "END FISICO"):
            return "End. Fisico"
        return header

    headers = [clean_header(value) for value in rows[header_index]]
    records = []
    for row in rows[header_index + 1:]:
        if not any(str(value or "").strip() for value in row):
            continue
        record = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            record[header] = row[index] if index < len(row) else None
        records.append(record)
    return records, f"Aba 1"


def read_spreadsheet_records(file_name: str, file_bytes: bytes):
    suffix = Path(file_name or "").suffix.lower()
    if suffix == ".ods" or (file_bytes[:2] == b"PK" and b"content.xml" in file_bytes[:20000]):
        return rows_to_records(read_ods_rows(file_bytes))
    if suffix in (".csv", ".txt", ".tsv"):
        text = file_bytes.decode("utf-8-sig", errors="replace")
        delimiter = "\t" if suffix == ".tsv" else None
        sample = text[:4096]
        if delimiter is None:
            try:
                delimiter = csv.Sniffer().sniff(sample, delimiters=";,|\t,").delimiter if sample.strip() else ","
            except csv.Error:
                delimiter = ";"
        rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
        return rows_to_records(rows)

    try:
        workbook = pd.ExcelFile(BytesIO(file_bytes))
        sheet_name = workbook.sheet_names[0]
        frame = pd.read_excel(workbook, sheet_name=sheet_name)
        frame = frame.where(pd.notnull(frame), None)
        return frame.to_dict(orient="records"), sheet_name
    except Exception:
        text = file_bytes.decode("utf-8-sig", errors="replace")
        sample = text[:4096]
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=";,|\t,").delimiter if sample.strip() else ";"
        except csv.Error:
            delimiter = ";"
        rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
        return rows_to_records(rows)


def unique_headers(raw_headers: list) -> list[str]:
    headers = []
    seen = {}
    for index, header in enumerate(raw_headers):
        label = normalize_record(header)
        label = str(label or "").strip() or f"Coluna {index + 1}"
        count = seen.get(label, 0) + 1
        seen[label] = count
        headers.append(label if count == 1 else f"{label} ({count})")
    return headers


def read_mercado_livre_ads_records(file_name: str, file_bytes: bytes):
    workbook = pd.ExcelFile(BytesIO(file_bytes))
    sheet_name = next((name for name in workbook.sheet_names if "ANUNCI" in normalize_text(name)), workbook.sheet_names[0])
    frame = pd.read_excel(workbook, sheet_name=sheet_name, header=None, dtype=object)
    frame = frame.where(pd.notnull(frame), None)
    rows = frame.values.tolist()

    header_index = None
    for index, row in enumerate(rows):
        row_key = " ".join(normalize_text(cell) for cell in row[:8] if normalize_text(cell))
        if "AGRUPADORDEVARIACOES" in row_key and "CODIGODOANUNCIO" in row_key:
            header_index = index
            break
    if header_index is None:
        raise ValueError("Nao encontrei o cabecalho com Agrupador de variacoes e Codigo do anuncio.")

    headers = unique_headers(rows[header_index])
    records = []
    skipped = 0
    for row in rows[header_index + 1:]:
        normalized_values = [normalize_record(value) for value in row]
        if not any(str(value or "").strip() for value in normalized_values):
            continue
        first_cells_key = " ".join(normalize_text(value) for value in normalized_values[:8] if normalize_text(value))
        if "AGRUPADORDEVARIACOES" in first_cells_key or "OBRIGATORIO" in first_cells_key:
            skipped += 1
            continue
        record = {
            header: normalized_values[index] if index < len(normalized_values) else None
            for index, header in enumerate(headers)
        }
        records.append(record)
    return records, sheet_name, header_index + 1, skipped


def read_mercado_livre_technical_sheet_records(file_name: str, file_bytes: bytes):
    workbook = pd.ExcelFile(BytesIO(file_bytes))
    records = []
    skipped = 0
    used_sheets = []
    first_header_row = None
    for sheet_name in workbook.sheet_names:
        if normalize_text(sheet_name) in ("AJUDA", "HIDDEN"):
            continue
        frame = pd.read_excel(workbook, sheet_name=sheet_name, header=None, dtype=object)
        frame = frame.where(pd.notnull(frame), None)
        rows = frame.values.tolist()

        header_index = None
        for index, row in enumerate(rows[:12]):
            row_key = " ".join(normalize_text(cell) for cell in row[:10] if normalize_text(cell))
            if "AGRUPADORDEVARIACOES" in row_key and "CODIGODOANUNCIO" in row_key and "SKU" in row_key:
                header_index = index
                break
        if header_index is None:
            continue

        headers = unique_headers(rows[header_index])
        used_sheets.append(sheet_name)
        if first_header_row is None:
            first_header_row = header_index + 1
        for row in rows[header_index + 1:]:
            normalized_values = [normalize_record(value) for value in row]
            if not any(str(value or "").strip() for value in normalized_values):
                continue
            first_cells_key = " ".join(normalize_text(value) for value in normalized_values[:10] if normalize_text(value))
            if "AGRUPADORDEVARIACOES" in first_cells_key or "OBRIGATORIO" in first_cells_key:
                skipped += 1
                continue
            record = {
                header: normalized_values[index] if index < len(normalized_values) else None
                for index, header in enumerate(headers)
            }
            record["_categoria_ml"] = sheet_name
            records.append(record)

    if not records:
        raise ValueError("Nao encontrei abas de Ficha Tecnica com Agrupador de variacoes, Codigo do anuncio e SKU.")
    sheet_label = "Ficha Tecnica" if len(used_sheets) > 1 else used_sheets[0]
    return records, sheet_label, first_header_row or 0, skipped


def read_mercado_livre_sales_records(file_name: str, file_bytes: bytes):
    workbook = pd.ExcelFile(BytesIO(file_bytes))
    sheet_name = next((name for name in workbook.sheet_names if "VENDA" in normalize_text(name)), workbook.sheet_names[0])
    frame = pd.read_excel(workbook, sheet_name=sheet_name, header=None, dtype=object)
    frame = frame.where(pd.notnull(frame), None)
    rows = frame.values.tolist()

    header_index = None
    for index, row in enumerate(rows):
        row_key = " ".join(normalize_text(cell) for cell in row[:30] if normalize_text(cell))
        compact_key = row_key.replace(" ", "")
        if "DATADAVENDA" in compact_key and "SKU" in compact_key and "ANUNCIO" in compact_key:
            header_index = index
            break
        if "NODEVENDA" in compact_key and "SKU" in compact_key:
            header_index = index
            break
    if header_index is None:
        raise ValueError("Nao encontrei o cabecalho com N. de venda, Data da venda e SKU.")

    headers = unique_headers(rows[header_index])
    records = []
    skipped = 0
    for row in rows[header_index + 1:]:
        normalized_values = [normalize_record(value) for value in row]
        if not any(str(value or "").strip() for value in normalized_values):
            continue
        record = {
            header: normalized_values[index] if index < len(normalized_values) else None
            for index, header in enumerate(headers)
        }
        if not mercado_livre_sales_key(record):
            skipped += 1
            continue
        records.append(record)
    return records, sheet_name, header_index + 1, skipped


def mercado_livre_record_value(record: dict, *labels: str):
    targets = {normalize_text(label) for label in labels}
    for key, value in record.items():
        if normalize_text(key) in targets:
            return value
    return None


def mercado_livre_sales_value(record: dict, *labels: str):
    direct = mercado_livre_record_value(record, *labels)
    if direct not in (None, ""):
        return direct
    targets = {normalize_text(label) for label in labels}
    for key, value in record.items():
        key_norm = normalize_text(key)
        if any(target and target in key_norm for target in targets):
            return value
    return None


def mercado_livre_sales_key(record: dict) -> str:
    sale_number = normalize_text(mercado_livre_sales_value(record, "N. de venda", "N.� de venda", "N.º de venda"))
    sku = normalize_text(mercado_livre_sales_value(record, "SKU"))
    ad_code = normalize_text(mercado_livre_sales_value(record, "# de anuncio", "# de anúncio", "# de an�ncio"))
    return "|".join(part for part in (sale_number, sku, ad_code) if part)


def mercado_livre_sales_payload(company_id: str, query: str = "", limit: int = 300):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    records = load_json(mercado_livre_sales_file(company_id), [])
    columns = []
    for record in records:
        for key in record:
            if not str(key).startswith("_") and key not in columns:
                columns.append(key)
        if columns:
            break
    matches = [record for record in records if value_matches(record, query)]
    limit = max(1, min(int(limit or 300), 1000))
    return {
        "empresa": COMPANIES[company_id],
        "empresa_id": company_id,
        "columns": [{"key": key, "label": key} for key in columns],
        "rows": [{key: record.get(key) for key in columns} for record in matches[:limit]],
        "total_filtrado": len(matches),
        "total_registros": len(records),
        "limite": limit,
    }


MERCADO_LIVRE_MONTHS_PT = {
    "JANEIRO": 1,
    "FEVEREIRO": 2,
    "MARCO": 3,
    "MARCO": 3,
    "ABRIL": 4,
    "MAIO": 5,
    "JUNHO": 6,
    "JULHO": 7,
    "AGOSTO": 8,
    "SETEMBRO": 9,
    "OUTUBRO": 10,
    "NOVEMBRO": 11,
    "DEZEMBRO": 12,
}


def parse_mercado_livre_sale_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    text = str(normalize_record(value) or "").strip()
    if not text:
        return None

    numeric_match = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})", text)
    if numeric_match:
        day = int(numeric_match.group(1))
        month = int(numeric_match.group(2))
        year = int(numeric_match.group(3))
        if year < 100:
            year += 2000
        try:
            return date(year, month, day)
        except ValueError:
            return None

    normalized = normalize_text(text)
    text_match = re.search(r"(\d{1,2})DE([A-Z]+)DE(\d{4})", normalized)
    if text_match:
        day = int(text_match.group(1))
        month = MERCADO_LIVRE_MONTHS_PT.get(text_match.group(2))
        year = int(text_match.group(3))
        if month:
            try:
                return date(year, month, day)
            except ValueError:
                return None
    return None


def mercado_livre_sales_net_value(record: dict) -> float:
    columns = [
        ("Receita por produtos (BRL)", "Receita por produtos"),
        ("Taxa de parcelamento equivalente ao acrescimo", "Taxa de parcelamento equivalente ao acr\u00e9scimo"),
        ("Tarifa de venda e impostos (BRL)", "Tarifa de venda e impostos"),
        ("Tarifas de envio (BRL)", "Tarifas de envio"),
        ("Cancelamentos e reembolsos (BRL)", "Cancelamentos e reembolsos"),
    ]
    total = 0.0
    for labels in columns:
        total += optional_number_value(mercado_livre_sales_value(record, *labels)) or 0.0
    return total


def mercado_livre_sales_dashboard_payload(company_id: str = "agrupado", year_filter: str = "todos"):
    if company_id == "agrupado":
        company_ids = list(MERCADO_LIVRE_COMPANIES)
        company_name = "Agrupado"
    elif company_id in MERCADO_LIVRE_COMPANIES:
        company_ids = [company_id]
        company_name = COMPANIES[company_id]
    else:
        raise ValueError("Empresa invalida para Mercado Livre.")

    year_filter = normalize_text(year_filter or "todos").lower()
    dependencies = [mercado_livre_sales_file(current_company_id) for current_company_id in company_ids]
    cache_key = ("mercado_livre_sales_dashboard", tuple(company_ids), year_filter)

    def build():
        monthly = defaultdict(float)
        years = set()
        rows_count = 0
        ignored_without_date = 0
        for current_company_id in company_ids:
            for record in load_json(mercado_livre_sales_file(current_company_id), []):
                sold_at = parse_mercado_livre_sale_date(mercado_livre_sales_value(record, "Data da venda"))
                if not sold_at:
                    ignored_without_date += 1
                    continue
                if sold_at.year < 2025:
                    continue
                years.add(sold_at.year)
                rows_count += 1
                monthly[(sold_at.year, sold_at.month)] += mercado_livre_sales_net_value(record)

        if not years:
            current_year = datetime.now().year
            years.add(max(2025, current_year))
        sorted_years = sorted(years)
        selected_year = year_filter
        month_labels = ["Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]

        points = []
        if selected_year in ("todos", "all"):
            for year in sorted_years:
                for month in range(1, 13):
                    points.append({
                        "label": f"{month_labels[month - 1]}/{year}",
                        "year": year,
                        "month": month,
                        "value": round(monthly.get((year, month), 0.0), 2),
                    })
        else:
            try:
                year = int(selected_year)
            except ValueError:
                year = sorted_years[-1]
                selected_year = str(year)
            if year not in years:
                years.add(year)
                sorted_years = sorted(years)
            for month in range(1, 13):
                points.append({
                    "label": month_labels[month - 1],
                    "year": year,
                    "month": month,
                    "value": round(monthly.get((year, month), 0.0), 2),
                })

        total = round(sum(point["value"] for point in points), 2)
        return {
            "empresa": company_name,
            "empresa_id": company_id,
            "year": selected_year,
            "years": sorted_years,
            "chart": {
                "title": "Vendas",
                "subtitle": "Valor liquido por mes",
                "points": points,
                "total": total,
            },
            "linhas_consideradas": rows_count,
            "linhas_sem_data": ignored_without_date,
        }

    return cached_payload(cache_key, dependencies, build)


def mercado_livre_record_identity_parts(record: dict) -> tuple[str, str]:
    ad_code = normalize_text(mercado_livre_record_value(record, "Codigo do anuncio", "Código do anúncio", "C�digo do an�ncio"))
    product_number = normalize_text(mercado_livre_record_value(record, "Numero do produto", "Número do produto", "N�mero do produto"))
    return ad_code, product_number


def mercado_livre_record_key(record: dict, ad_code_counts: Counter) -> str:
    ad_code, product_number = mercado_livre_record_identity_parts(record)
    if not ad_code:
        return ""
    if ad_code_counts.get(ad_code, 0) > 1 and product_number:
        return f"{ad_code}|{product_number}"
    return ad_code


MERCADO_LIVRE_SUMMARY_COLUMNS = [
    "C\u00f3digo do an\u00fancio",
    "N\u00famero do produto",
    "SKU",
    "T\u00edtulo",
    "Estoque no dep\u00f3sito",
    "Estoque Ionlab",
    "Pre\u00e7o",
    "Forma de entrega",
    "Tipo de an\u00fancio",
    "Tarifa de venda",
    "Estado",
    "Pre\u00e7o de venda Ionlab",
    "MKPC",
    "MKPPV",
]


def mercado_livre_summary_field(record: dict, label: str):
    aliases = {
        "C\u00f3digo do an\u00fancio": ("Codigo do anuncio", "Código do anúncio", "C�digo do an�ncio"),
        "N\u00famero do produto": ("Numero do produto", "Número do produto", "N�mero do produto"),
        "SKU": ("SKU",),
        "T\u00edtulo": ("Titulo", "Título", "T�tulo"),
        "Estoque no dep\u00f3sito": ("Estoque no deposito", "Estoque no depósito", "Estoque no dep�sito"),
        "Pre\u00e7o": ("Preco", "Preço", "Pre�o"),
        "Forma de entrega": ("Forma de entrega",),
        "Tipo de an\u00fancio": ("Tipo de anuncio", "Tipo de anúncio", "Tipo de an�ncio"),
        "Tarifa de venda": ("Tarifa de venda",),
        "Estado": ("Estado",),
    }
    return mercado_livre_record_value(record, *aliases.get(label, (label,)))


def rounded_number(value: float | None, digits: int = 2):
    if value is None:
        return ""
    return round(float(value), digits)


def mercado_livre_kit_quantity(text: str) -> int:
    normalized = normalize_text(text)
    match = re.search(r"\bKIT\s*(\d+)", normalized)
    if not match:
        return 1
    try:
        return max(1, int(match.group(1)))
    except ValueError:
        return 1


def mercado_livre_ionlab_reference_maps():
    stock_by_sku = {}
    for record in load_json(stock_file("ionlab"), []):
        key = normalize_text(record.get("Referencia"))
        if not key:
            continue
        quantity = optional_number_value(record.get("Quantidade")) or 0
        stock_by_sku[key] = stock_by_sku.get(key, 0) + quantity

    price_by_sku = {}
    for record in load_json(price_table_file("ionlab"), []):
        key = normalize_text(record.get("PR_COD"))
        price = optional_number_value(record.get("VAL_VEND"))
        if not key or price is None:
            continue
        current = price_by_sku.get(key)
        price_by_sku[key] = price if current is None else max(current, price)

    cost_by_sku = {
        key: info.get("custo")
        for key, info in costing_info_by_reference("ionlab").items()
        if info.get("custo") is not None
    }
    return stock_by_sku, price_by_sku, cost_by_sku


def mercado_livre_import_by_type(company_id: str, import_type: str) -> dict | None:
    imports = mercado_livre_load_imports(company_id)
    return next((item for item in imports if item.get("tipo", "anuncios") == import_type), None)


def mercado_livre_summary_dependencies(company_id: str) -> list[Path]:
    ml_dir = mercado_livre_dir(company_id)
    mercado_files = sorted(ml_dir.glob("*.json")) if ml_dir.exists() else []
    return [
        *mercado_files,
        stock_file("ionlab"),
        price_table_file("ionlab"),
        costing_file("ionlab"),
        costing_fabricated_file("ionlab"),
    ]


def mercado_livre_summary_rows(company_id: str) -> list[dict]:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    return cached_payload(
        ("mercado_livre_summary_rows", company_id),
        mercado_livre_summary_dependencies(company_id),
        lambda: build_mercado_livre_summary_rows(company_id),
    )


def build_mercado_livre_summary_rows(company_id: str) -> list[dict]:
    source_records = []
    for import_type in ("ficha_tecnica", "anuncios"):
        import_meta = mercado_livre_import_by_type(company_id, import_type)
        if not import_meta:
            continue
        records = load_json(mercado_livre_import_file(company_id, import_meta.get("id")), [])
        for record in records:
            source_records.append((import_type, record))

    ad_code_counts = Counter(
        ad_code
        for ad_code, _product_number in (
            mercado_livre_record_identity_parts(record)
            for _source, record in source_records
        )
        if ad_code
    )
    stock_by_sku, price_by_sku, cost_by_sku = mercado_livre_ionlab_reference_maps()
    summary_by_key = {}
    summary_rows = []
    for source, record in source_records:
        key = mercado_livre_record_key(record, ad_code_counts)
        if not key:
            continue
        row = summary_by_key.get(key)
        if row is None:
            row = {column: "" for column in MERCADO_LIVRE_SUMMARY_COLUMNS}
            summary_by_key[key] = row
            summary_rows.append(row)
        for column in MERCADO_LIVRE_SUMMARY_COLUMNS:
            if column in ("Estoque Ionlab", "Pre\u00e7o de venda Ionlab", "MKPC", "MKPPV"):
                continue
            value = mercado_livre_summary_field(record, column)
            if value not in (None, "") and (source == "anuncios" or row.get(column) in (None, "")):
                row[column] = normalize_record(value)

    for row in summary_rows:
        sku_key = normalize_text(row.get("SKU"))
        ionlab_stock = stock_by_sku.get(sku_key)
        ionlab_price = price_by_sku.get(sku_key)
        ionlab_cost = cost_by_sku.get(sku_key)
        ml_price = optional_number_value(row.get("Pre\u00e7o"))
        kit_quantity = mercado_livre_kit_quantity(row.get("T\u00edtulo") or "")
        kit_cost = ionlab_cost * kit_quantity if ionlab_cost is not None else None
        kit_ionlab_price = ionlab_price * kit_quantity if ionlab_price is not None else None

        row["Estoque Ionlab"] = round(ionlab_stock, 4) if ionlab_stock is not None else ""
        row["Pre\u00e7o"] = rounded_number(ml_price, 2) if ml_price is not None else ""
        row["Pre\u00e7o de venda Ionlab"] = rounded_number(ionlab_price, 2) if ionlab_price is not None else ""
        row["MKPC"] = rounded_number(((ml_price - kit_cost) / kit_cost) + 1, 2) if ml_price and kit_cost else ""
        row["MKPPV"] = rounded_number(((ml_price - kit_ionlab_price) / ml_price) + 1, 2) if ml_price and kit_ionlab_price else ""

    return sorted(summary_rows, key=lambda item: (
        1 if not normalize_text(item.get("SKU")) else 0,
        normalize_text(item.get("SKU")),
        normalize_text(item.get("C\u00f3digo do an\u00fancio")),
    ))


def mercado_livre_summary_payload(company_id: str, query: str = "", limit: int = 300):
    rows = mercado_livre_summary_rows(company_id)
    matches = [record for record in rows if value_matches(record, query)]
    limit = max(1, min(int(limit or 300), 1000))
    return {
        "empresa": COMPANIES[company_id],
        "empresa_id": company_id,
        "imports": mercado_livre_load_imports(company_id),
        "import_id": "resumo",
        "import_name": "Resumo",
        "tipo": "resumo",
        "tipo_nome": "Resumo",
        "columns": [{"key": key, "label": key} for key in MERCADO_LIVRE_SUMMARY_COLUMNS],
        "rows": matches[:limit],
        "total_filtrado": len(matches),
        "total_registros": len(rows),
        "limite": limit,
    }


def mercado_livre_summary_export_xlsx(company_id: str) -> bytes:
    rows = mercado_livre_summary_rows(company_id)
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe = pd.DataFrame(rows, columns=MERCADO_LIVRE_SUMMARY_COLUMNS)
        dataframe.to_excel(writer, index=False, sheet_name="Resumo")
        worksheet = writer.sheets["Resumo"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        from openpyxl.styles import PatternFill

        inactive_fill = PatternFill(fill_type="solid", fgColor="FFF7CC")
        money_columns = {"Pre\u00e7o", "Pre\u00e7o de venda Ionlab"}
        stock_columns = {"Estoque no dep\u00f3sito", "Estoque Ionlab"}
        markup_columns = {"MKPC", "MKPPV"}
        column_indexes = {cell.value: cell.column for cell in worksheet[1]}

        for column_name in money_columns:
            column_index = column_indexes.get(column_name)
            if column_index:
                for cell in worksheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2, max_row=worksheet.max_row):
                    for item in cell:
                        item.number_format = '#,##0.00'
        for column_name in stock_columns:
            column_index = column_indexes.get(column_name)
            if column_index:
                for cell in worksheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2, max_row=worksheet.max_row):
                    for item in cell:
                        item.number_format = '#,##0.###'
        for column_name in markup_columns:
            column_index = column_indexes.get(column_name)
            if column_index:
                for cell in worksheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2, max_row=worksheet.max_row):
                    for item in cell:
                        item.number_format = '#,##0.00'

        state_index = column_indexes.get("Estado")
        if state_index:
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                state_value = row[state_index - 1].value
                if normalize_text(state_value) == "INATIVO":
                    for cell in row:
                        cell.fill = inactive_fill

        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 14), 55)
    return output.getvalue()


MERCADO_LIVRE_GENERAL_COLUMNS = [
    "SKU",
    "T\u00edtulo",
    "Estoque no dep\u00f3sito Onix",
    "Estoque no dep\u00f3sito Vitralab",
    "Estoque no dep\u00f3sito Nativalab",
    "Estoque Ionlab",
    "Pre\u00e7o Onix",
    "Pre\u00e7o Vitralab",
    "Pre\u00e7o Nativalab",
    "Estado Onix",
    "Estado Vitralab",
    "Estado Nativalab",
    "Pre\u00e7o de venda Ionlab",
    "Observa\u00e7\u00f5es",
]


MERCADO_LIVRE_GENERAL_COMPANIES = [
    ("onix", "Onix"),
    ("vitralab", "Vitralab"),
    ("nativalab", "Nativalab"),
]


def mercado_livre_general_rows() -> list[dict]:
    dependencies = []
    for company_id, _company_label in MERCADO_LIVRE_GENERAL_COMPANIES:
        dependencies.extend(mercado_livre_summary_dependencies(company_id))
    return cached_payload(
        ("mercado_livre_general_rows",),
        dependencies,
        build_mercado_livre_general_rows,
    )


def build_mercado_livre_general_rows() -> list[dict]:
    rows_by_sku = {}
    for company_id, company_label in MERCADO_LIVRE_GENERAL_COMPANIES:
        for summary_row in mercado_livre_summary_rows(company_id):
            sku = normalize_record(summary_row.get("SKU"))
            sku_key = normalize_text(sku)
            if not sku_key:
                continue
            row = rows_by_sku.get(sku_key)
            if row is None:
                row = {column: "" for column in MERCADO_LIVRE_GENERAL_COLUMNS}
                row["SKU"] = sku
                rows_by_sku[sku_key] = row

            if not row.get("T\u00edtulo") and summary_row.get("T\u00edtulo"):
                row["T\u00edtulo"] = summary_row.get("T\u00edtulo")
            if row.get("Estoque Ionlab") in ("", None) and summary_row.get("Estoque Ionlab") not in ("", None):
                row["Estoque Ionlab"] = summary_row.get("Estoque Ionlab")
            if row.get("Pre\u00e7o de venda Ionlab") in ("", None) and summary_row.get("Pre\u00e7o de venda Ionlab") not in ("", None):
                row["Pre\u00e7o de venda Ionlab"] = summary_row.get("Pre\u00e7o de venda Ionlab")

            row[f"Estoque no dep\u00f3sito {company_label}"] = summary_row.get("Estoque no dep\u00f3sito") if summary_row.get("Estoque no dep\u00f3sito") not in (None, "") else ""
            row[f"Pre\u00e7o {company_label}"] = summary_row.get("Pre\u00e7o") if summary_row.get("Pre\u00e7o") not in (None, "") else ""
            row[f"Estado {company_label}"] = summary_row.get("Estado") if summary_row.get("Estado") not in (None, "") else ""

    for row in rows_by_sku.values():
        announced_stock = sum(
            optional_number_value(row.get(column)) or 0
            for column in (
                "Estoque no dep\u00f3sito Onix",
                "Estoque no dep\u00f3sito Vitralab",
                "Estoque no dep\u00f3sito Nativalab",
            )
        )
        ionlab_stock = optional_number_value(row.get("Estoque Ionlab")) or 0
        if announced_stock > ionlab_stock:
            row["Observa\u00e7\u00f5es"] = "Estoque anunciado maior que o estoque disponivel"

    return sorted(rows_by_sku.values(), key=lambda item: normalize_text(item.get("SKU")))


def mercado_livre_general_payload(query: str = "", limit: int = 300):
    rows = mercado_livre_general_rows()
    matches = [record for record in rows if value_matches(record, query)]
    limit = max(1, min(int(limit or 300), 1000))
    return {
        "tipo": "gestao_geral",
        "tipo_nome": "Gestao Geral",
        "columns": [{"key": key, "label": key} for key in MERCADO_LIVRE_GENERAL_COLUMNS],
        "rows": matches[:limit],
        "total_filtrado": len(matches),
        "total_registros": len(rows),
        "limite": limit,
    }


def mercado_livre_general_export_xlsx() -> bytes:
    rows = mercado_livre_general_rows()
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe = pd.DataFrame(rows, columns=MERCADO_LIVRE_GENERAL_COLUMNS)
        dataframe.to_excel(writer, index=False, sheet_name="Gestao Geral")
        worksheet = writer.sheets["Gestao Geral"]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        money_columns = {"Pre\u00e7o Onix", "Pre\u00e7o Vitralab", "Pre\u00e7o Nativalab", "Pre\u00e7o de venda Ionlab"}
        stock_columns = {"Estoque no dep\u00f3sito Onix", "Estoque no dep\u00f3sito Vitralab", "Estoque no dep\u00f3sito Nativalab", "Estoque Ionlab"}
        column_indexes = {cell.value: cell.column for cell in worksheet[1]}
        for column_name in money_columns:
            column_index = column_indexes.get(column_name)
            if column_index:
                for column_cells in worksheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2, max_row=worksheet.max_row):
                    for cell in column_cells:
                        cell.number_format = '#,##0.00'
        for column_name in stock_columns:
            column_index = column_indexes.get(column_name)
            if column_index:
                for column_cells in worksheet.iter_cols(min_col=column_index, max_col=column_index, min_row=2, max_row=worksheet.max_row):
                    for cell in column_cells:
                        cell.number_format = '#,##0.###'
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 14), 55)
    return output.getvalue()


def stock_record_key(record: dict) -> str:
    return normalize_text(
        record.get("Produto")
        or record.get("Referencia")
        or record.get("Referência")
        or record.get("PR_COD")
        or record.get("Codigo")
        or record.get("Código")
    )


def import_stock(company_id: str, file_name: str, file_bytes: bytes):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    raw_records, sheet_name = read_spreadsheet_records(file_name, file_bytes)
    stock_by_reference = {}
    skipped = 0
    for raw in raw_records:
        reference = normalize_record(
            raw.get("Produto")
            or raw.get("Referencia")
            or raw.get("Referência")
            or raw.get("PR_COD")
            or raw.get("Codigo")
            or raw.get("Código")
        )
        reference = str(reference or "").strip()
        if not reference:
            skipped += 1
            continue
        key = normalize_text(reference)
        quantity = optional_number_value(raw.get("Quantidade") or raw.get("Qtd") or raw.get("Saldo")) or 0.0
        existing = stock_by_reference.get(key)
        if not existing:
            existing = {
                "Tipo": normalize_record(raw.get("Tipo") or raw.get("TP")),
                "C.Fiscal": normalize_record(raw.get("C.Fiscal") or raw.get("NCM")),
                "Id.Estoque": normalize_record(raw.get("Id.Estoque")),
                "Local": normalize_record(raw.get("Local")),
                "Referencia": reference,
                "Descricao": normalize_record(raw.get("Descricao") or raw.get("Descrição")),
                "AGRP": normalize_record(raw.get("AGRP")),
                "Origem": normalize_record(raw.get("Origem")),
                "Modelo/Serie": normalize_record(raw.get("Modelo/Serie") or raw.get("Modelo/Série")),
                "Unidade": normalize_record(raw.get("Unidade")),
                "Quantidade": 0.0,
                "V.Unitario": None,
                "Total": None,
                "End. Fisico": normalize_record(raw.get("End. Físico") or raw.get("End. Fisico")),
                "_empresa_id": company_id,
                "_empresa_nome": COMPANIES[company_id],
                "_atualizado_em": datetime.now().isoformat(timespec="seconds"),
            }
            stock_by_reference[key] = existing
        existing["Quantidade"] = round(number_value(existing.get("Quantidade")) + quantity, 4)

    records = sorted(stock_by_reference.values(), key=lambda item: normalize_text(item.get("Referencia")))
    save_json(stock_file(company_id), records)

    summary = {
        "empresa": COMPANIES[company_id],
        "arquivo": file_name,
        "aba": sheet_name,
        "linhas_lidas": len(raw_records),
        "referencias_atualizadas": len(records),
        "referencias_excluidas": "Base substituida pela ultima importacao",
        "linhas_ignoradas": skipped,
        "importado_em": datetime.now().isoformat(timespec="seconds"),
    }
    logs = load_json(stock_import_log_file(company_id), [])
    logs.append(summary)
    save_json(stock_import_log_file(company_id), logs)
    return summary


def get_field_by_prefix(record: dict, *prefixes):
    normalized_prefixes = [normalize_text(prefix) for prefix in prefixes if prefix]
    for key, value in record.items():
        normalized_key = normalize_text(key)
        if any(normalized_key.startswith(prefix) for prefix in normalized_prefixes):
            return value
    return None


def normalize_transit_date(value):
    parsed = record_date_value(value)
    if parsed:
        return parsed.isoformat()
    value = normalize_record(value)
    return str(value or "").strip()


def import_in_transit(company_id: str, file_name: str, file_bytes: bytes):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    raw_records, sheet_name = read_spreadsheet_records(file_name, file_bytes)
    records_by_key = {}
    skipped = 0
    now = datetime.now().isoformat(timespec="seconds")

    for raw in raw_records:
        reference = normalize_record(get_field_by_prefix(raw, "Referencia", "Refer", "PR_COD", "Codigo"))
        reference = str(reference or "").strip()
        if not reference or normalize_text(reference) in {"REFERENCIA", "REFER"}:
            skipped += 1
            continue
        process = normalize_record(get_field_by_prefix(raw, "Processo"))
        key = "|".join([normalize_text(reference), normalize_text(process)])
        record = {
            "Referencia": reference,
            "Descricao": product_sentence_case(get_field_by_prefix(raw, "Descricao", "Descr")),
            "QTY": optional_number_value(get_field_by_prefix(raw, "QTY", "Quantidade", "Qtd")) or 0.0,
            "Processo": normalize_record(process),
            "Previsao": normalize_transit_date(get_field_by_prefix(raw, "Previsao", "Previs")),
            "_empresa_id": company_id,
            "_empresa_nome": COMPANIES[company_id],
            "_atualizado_em": now,
        }
        records_by_key[key] = record

    records = sorted(records_by_key.values(), key=lambda item: (
        table_sort_value(item, "Previsao"),
        normalize_text(item.get("Referencia")),
        normalize_text(item.get("Processo")),
    ))
    save_json(in_transit_file(company_id), records)

    summary = {
        "empresa": COMPANIES[company_id],
        "arquivo": file_name,
        "aba": sheet_name,
        "linhas_lidas": len(raw_records),
        "itens_importados": len(records),
        "linhas_ignoradas": skipped,
        "referencias_excluidas": "Base substituida pela ultima importacao",
        "importado_em": now,
    }
    logs = load_json(import_log_file(company_id), [])
    logs.append({"tipo": "em_transito", **summary})
    save_json(import_log_file(company_id), logs)
    return summary


def import_price_table(company_id: str, file_name: str, file_bytes: bytes):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    raw_records, sheet_name = read_spreadsheet_records(file_name, file_bytes)
    costing_by_reference = costing_info_by_reference(company_id)
    prices_by_key = {}
    skipped = 0
    now = datetime.now().isoformat(timespec="seconds")

    for raw in raw_records:
        product_code = normalize_record(
            get_field(raw, "PR_COD", "Referencia", "Referência", "Produto", "Codigo", "Código")
        )
        product_code = str(product_code or "").strip()
        if not product_code:
            skipped += 1
            continue

        id_tab = normalize_record(get_field(raw, "ID_TAB", "ID TAB", "Tabela"))
        id_est = normalize_record(get_field(raw, "ID_EST", "ID EST", "Estoque"))
        lc_cod = normalize_record(get_field(raw, "LC_COD", "LC COD", "Lista", "Local"))
        key = "|".join([
            normalize_text(id_tab),
            normalize_text(id_est),
            normalize_text(lc_cod),
            normalize_text(product_code),
        ]).strip("|")
        if not key:
            skipped += 1
            continue

        costing_info = classification_for_reference(product_code, costing_by_reference)
        record = {
            "ID_TAB": id_tab,
            "ID_EST": id_est,
            "LC_COD": lc_cod,
            "PR_COD": product_code,
            "PR_DES": normalize_record(get_field(raw, "PR_DES", "Descricao", "Descrição", "Descricao item", "Descrição item")),
            "AGRP": normalize_record(get_field(raw, "AGRP")) or costing_info.get("AGRP") or "",
            "Origem": normalize_record(get_field(raw, "Origem")) or costing_info.get("Origem") or "",
            "VAL_VEND": optional_number_value(get_field(raw, "VAL_VEND", "Valor Venda", "V. Venda", "Preco", "Preço")),
            "PER_DESC": optional_number_value(get_field(raw, "PER_DESC", "Percentual Desconto", "% Desconto", "Desconto")),
            "PER_COM": optional_number_value(get_field(raw, "PER_COM", "Percentual Comissao", "% Comissao", "Comissao")),
            "Aliquota IPI": optional_number_value(get_field(raw, "Aliquota IPI", "Aliquota de IPI", "Aliq IPI", "IPI", "ALIQ_IPI")) or 0.0,
            "Custo BRL Indice": costing_info.get("custo") if costing_info.get("custo") is not None else "",
            "_empresa_id": company_id,
            "_empresa_nome": COMPANIES[company_id],
            "_atualizado_em": now,
        }
        prices_by_key[key] = record

    records = sorted(prices_by_key.values(), key=lambda item: (
        normalize_text(item.get("ID_TAB")),
        normalize_text(item.get("PR_COD")),
    ))
    save_json(price_table_file(company_id), records)

    summary = {
        "empresa": COMPANIES[company_id],
        "arquivo": file_name,
        "aba": sheet_name,
        "linhas_lidas": len(raw_records),
        "itens_importados": len(records),
        "linhas_ignoradas": skipped,
        "referencias_excluidas": "Base substituida pela ultima importacao",
        "importado_em": now,
        "lembrete_ipi": (
            "Atualize as aliquotas do IPI em Cadastros > Tabela de Precos > Atualizar IPI. "
            "Caminho no ERP: ERP > Manutencao > Produtos/Servicos > Estoque e clica em Pesquisar, "
            "dai voce pode colocar estado ativo, e outros filtros que queira."
        ),
    }
    logs = load_json(import_log_file(company_id), [])
    logs.append({"tipo": "tabela_precos", **summary})
    save_json(import_log_file(company_id), logs)
    return summary


def import_products(company_id: str, file_name: str, file_bytes: bytes):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    raw_records, sheet_name = read_spreadsheet_records(file_name, file_bytes)
    products_by_key = {}
    skipped = 0
    now = datetime.now().isoformat(timespec="seconds")

    for raw in raw_records:
        reference = normalize_record(get_field(raw, "Referencia", "Referência", "Refer�ncia", "PR_COD", "Codigo", "Código"))
        product_id = normalize_record(get_field(raw, "ID", "Id", "Codigo ID", "Código ID"))
        key = normalize_text(reference) or f"ID:{normalize_text(product_id)}"
        if not key:
            skipped += 1
            continue
        record = {
            "ID": normalize_record(product_id),
            "Referencia": normalize_record(reference),
            "Descricao resumida": normalize_record(get_field(raw, "Descricao resumida", "Descrição resumida", "Descri��o resumida")),
            "Descricao": normalize_record(get_field(raw, "Descricao", "Descrição", "Descri��o")),
            "Dados tecnicos": normalize_record(get_field(raw, "Dados tecnicos", "Dados técnicos", "Dados t�cnicos")),
            "Descricao completa": normalize_record(get_field(raw, "Descricao completa", "Descrição completa", "Descri��o completa")),
            "Marca": normalize_record(get_field(raw, "Marca")),
            "So se usuario especificar": normalize_record(get_field(raw, "So se usuario especificar", "Só se usuário especificar", "S� se usu�rio especificar")),
            "Palavra chave": normalize_record(get_field(raw, "Palavra chave")),
            "Palavras positivas": normalize_record(get_field(raw, "Palavras positivas")),
            "Palavras negativas": normalize_record(get_field(raw, "Palavras negativas")),
            "Regras de variaveis": normalize_record(get_field(raw, "Regras de variaveis", "Regras de variáveis", "Regras de vari�veis")),
            "Score minimo": optional_number_value(get_field(raw, "Score minimo", "Score mínimo", "Score m�nimo")),
            "_empresa_id": company_id,
            "_empresa_nome": COMPANIES[company_id],
            "_atualizado_em": now,
        }
        record = normalize_product_record_case(record)
        products_by_key[key] = record

    records = sorted(products_by_key.values(), key=lambda item: (
        normalize_text(item.get("Referencia")),
        normalize_text(item.get("ID")),
    ))
    save_json(products_file(company_id), records)

    summary = {
        "empresa": COMPANIES[company_id],
        "arquivo": file_name,
        "aba": sheet_name,
        "linhas_lidas": len(raw_records),
        "itens_importados": len(records),
        "linhas_ignoradas": skipped,
        "referencias_excluidas": "Base substituida pela ultima importacao",
        "importado_em": now,
    }
    logs = load_json(import_log_file(company_id), [])
    logs.append({"tipo": "cadastro_produtos", **summary})
    save_json(import_log_file(company_id), logs)
    return summary


def import_mercado_livre_ads(company_id: str, file_name: str, file_bytes: bytes, import_type: str = "anuncios"):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    import_type = mercado_livre_import_type(import_type)
    type_name = MERCADO_LIVRE_IMPORT_TYPES.get(import_type, "Anuncios")
    if import_type == "ficha_tecnica":
        raw_records, sheet_name, header_row, skipped = read_mercado_livre_technical_sheet_records(file_name, file_bytes)
    else:
        raw_records, sheet_name, header_row, skipped = read_mercado_livre_ads_records(file_name, file_bytes)
    now = datetime.now().isoformat(timespec="seconds")
    records = []
    for index, record in enumerate(raw_records, start=1):
        clean_record = {
            key: normalize_record(value)
            for key, value in record.items()
        }
        clean_record["_empresa_id"] = company_id
        clean_record["_empresa_nome"] = COMPANIES[company_id]
        clean_record["_importado_em"] = now
        clean_record["_linha_importacao"] = index
        clean_record["_tipo_importacao"] = import_type
        records.append(clean_record)

    imports = mercado_livre_load_imports(company_id)
    selected_import = next((item for item in imports if item.get("tipo", "anuncios") == import_type), None)
    import_id = selected_import.get("id") if selected_import else import_type
    import_name = selected_import.get("nome") if selected_import else type_name
    existing_records = load_json(mercado_livre_import_file(company_id, import_id), []) if selected_import else []

    ad_code_counts = Counter(
        ad_code
        for ad_code, _product_number in (
            mercado_livre_record_identity_parts(record)
            for record in [*existing_records, *records]
        )
        if ad_code
    )
    merged_by_key = {}
    merged_records = []
    for existing in existing_records:
        key = mercado_livre_record_key(existing, ad_code_counts)
        if not key:
            continue
        if key not in merged_by_key:
            merged_by_key[key] = existing
            merged_records.append(existing)

    inserted = 0
    updated = 0
    unchanged = 0
    skipped_without_key = 0
    for record in records:
        key = mercado_livre_record_key(record, ad_code_counts)
        if not key:
            skipped_without_key += 1
            continue
        current = merged_by_key.get(key)
        if current is None:
            record["_criado_em"] = now
            record["_atualizado_em"] = now
            merged_by_key[key] = record
            merged_records.append(record)
            inserted += 1
            continue

        changed = False
        for field, value in record.items():
            if field in ("_importado_em", "_linha_importacao"):
                continue
            if current.get(field) != value:
                current[field] = value
                changed = True
        current["_empresa_id"] = company_id
        current["_empresa_nome"] = COMPANIES[company_id]
        current["_tipo_importacao"] = import_type
        current["_ultima_importacao_em"] = now
        if changed:
            current["_atualizado_em"] = now
            updated += 1
        else:
            unchanged += 1

    save_json(mercado_livre_import_file(company_id, import_id), merged_records)
    metadata = {
        "id": import_id,
        "nome": import_name,
        "arquivo": file_name,
        "aba": sheet_name,
        "tipo": import_type,
        "tipo_nome": type_name,
        "linha_cabecalho": header_row,
        "linhas": len(merged_records),
        "ultima_importacao_linhas": len(records),
        "incluidos_ultima_importacao": inserted,
        "alterados_ultima_importacao": updated,
        "sem_alteracao_ultima_importacao": unchanged,
        "importado_em": now,
    }
    if selected_import:
        selected_import.update(metadata)
    else:
        imports.append(metadata)
    save_json(mercado_livre_imports_file(company_id), imports)

    summary = {
        "empresa": COMPANIES[company_id],
        "import_id": import_id,
        "nome_importacao": import_name,
        "arquivo": file_name,
        "aba": sheet_name,
        "tipo": import_type,
        "tipo_nome": type_name,
        "linha_cabecalho": header_row,
        "linhas_lidas": len(raw_records) + skipped,
        "linhas_importadas": inserted + updated,
        "linhas_incluidas": inserted,
        "linhas_alteradas": updated,
        "linhas_sem_alteracao": unchanged,
        "linhas_ignoradas": skipped + skipped_without_key,
        "linhas_sem_chave": skipped_without_key,
        "total_anuncios_empresa": len(merged_records),
        "total_registros_empresa": len(merged_records),
        "importado_em": now,
    }
    logs = load_json(import_log_file(company_id), [])
    logs.append({"tipo": f"mercado_livre_{import_type}", **summary})
    save_json(import_log_file(company_id), logs)
    return summary


def import_mercado_livre_sales(company_id: str, file_name: str, file_bytes: bytes):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")
    if company_id not in MERCADO_LIVRE_COMPANIES:
        raise ValueError("Empresa invalida para Mercado Livre.")

    raw_records, sheet_name, header_row, skipped = read_mercado_livre_sales_records(file_name, file_bytes)
    now = datetime.now().isoformat(timespec="seconds")
    existing_records = load_json(mercado_livre_sales_file(company_id), [])
    by_key = {}
    merged_records = []
    for record in existing_records:
        key = mercado_livre_sales_key(record)
        if not key:
            continue
        if key not in by_key:
            by_key[key] = record
            merged_records.append(record)

    inserted = 0
    updated = 0
    unchanged = 0
    skipped_without_key = 0
    for index, record in enumerate(raw_records, start=1):
        clean_record = {key: normalize_record(value) for key, value in record.items()}
        clean_record["_empresa_id"] = company_id
        clean_record["_empresa_nome"] = COMPANIES[company_id]
        clean_record["_importado_em"] = now
        clean_record["_linha_importacao"] = index
        key = mercado_livre_sales_key(clean_record)
        if not key:
            skipped_without_key += 1
            continue
        current = by_key.get(key)
        if current is None:
            clean_record["_criado_em"] = now
            clean_record["_atualizado_em"] = now
            by_key[key] = clean_record
            merged_records.append(clean_record)
            inserted += 1
            continue
        changed = False
        for field, value in clean_record.items():
            if field in ("_importado_em", "_linha_importacao"):
                continue
            if current.get(field) != value:
                current[field] = value
                changed = True
        current["_ultima_importacao_em"] = now
        if changed:
            current["_atualizado_em"] = now
            updated += 1
        else:
            unchanged += 1

    save_json(mercado_livre_sales_file(company_id), merged_records)
    summary = {
        "empresa": COMPANIES[company_id],
        "arquivo": file_name,
        "aba": sheet_name,
        "linha_cabecalho": header_row,
        "linhas_lidas": len(raw_records) + skipped,
        "linhas_importadas": inserted + updated,
        "linhas_incluidas": inserted,
        "linhas_alteradas": updated,
        "linhas_sem_alteracao": unchanged,
        "linhas_ignoradas": skipped + skipped_without_key,
        "linhas_sem_chave": skipped_without_key,
        "total_vendas_empresa": len(merged_records),
        "importado_em": now,
    }
    logs = load_json(import_log_file(company_id), [])
    logs.append({"tipo": "mercado_livre_vendas", **summary})
    save_json(import_log_file(company_id), logs)
    return summary


def stock_export_csv(company_id: str) -> bytes:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    records = stock_records_with_last_exit(company_id)
    records = sorted(
        records,
        key=lambda record: (
            -number_value(record.get("Total")),
            normalize_text(record.get("Referencia")),
        ),
    )
    output = BytesIO()
    output.write(b"\xef\xbb\xbf")
    text_output = TextIOWrapper(output, encoding="utf-8", newline="", write_through=True)
    writer = csv.writer(text_output, delimiter=";")
    output.write("sep=;\r\n".encode("utf-8"))
    writer.writerow([label for _key, label in STOCK_TABLE_COLUMNS])
    for record in records:
        writer.writerow([
            br_decimal(record.get(key)) if key in ("Quantidade", "V.Unitario", "Total")
            else record.get(key) if record.get(key) is not None else ""
            for key, _label in STOCK_TABLE_COLUMNS
        ])
    text_output.flush()
    return output.getvalue()


def stock_export_xlsx(company_id: str) -> bytes:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    records = stock_records_with_last_exit(company_id)
    records = sorted(
        records,
        key=lambda record: (
            -number_value(record.get("Total")),
            normalize_text(record.get("Referencia")),
        ),
    )
    return export_records_xlsx(records, [key for key, _label in STOCK_TABLE_COLUMNS], "Estoque")


def in_transit_export_xlsx(company_id: str) -> bytes:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    records = sorted(
        load_json(in_transit_file(company_id), []),
        key=lambda record: (
            table_sort_value(record, "Previsao"),
            normalize_text(record.get("Referencia")),
            normalize_text(record.get("Processo")),
        ),
    )
    return export_records_xlsx(records, [key for key, _label in IN_TRANSIT_TABLE_COLUMNS], "Em Transito")


def price_table_export_xlsx(company_id: str) -> bytes:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    records = sorted(
        price_table_records_with_cost(company_id, save=True),
        key=lambda record: (
            normalize_text(record.get("PR_COD")),
            normalize_text(record.get("ID_TAB")),
            normalize_text(record.get("ID_EST")),
        ),
    )
    return export_records_xlsx(records, [key for key, _label in PRICE_TABLE_COLUMNS], "Tabela de Precos")


def update_price_table_ipi(company_id: str, file_name: str, file_bytes: bytes):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    raw_records, sheet_name = read_spreadsheet_records(file_name, file_bytes)
    ipi_by_reference = {}
    skipped = 0
    for raw in raw_records:
        reference = normalize_record(get_field(raw, "PR_COD", "Referencia", "Referência", "Codigo", "Código"))
        reference_key = normalize_text(reference)
        ipi_value = optional_number_value(get_field(raw, "FI_ALI", "Aliquota IPI", "Aliquota de IPI", "IPI"))
        if not reference_key or ipi_value is None:
            skipped += 1
            continue
        ipi_by_reference[reference_key] = ipi_value

    records = load_json(price_table_file(company_id), [])
    now = datetime.now().isoformat(timespec="seconds")
    updated = 0
    not_found = 0
    matched_references = set()
    for record in records:
        reference_key = normalize_text(record.get("PR_COD"))
        if not reference_key:
            continue
        old_value = record.get("Aliquota IPI")
        if old_value in (None, "") and record.get("Aliquota de IPI") not in (None, ""):
            record["Aliquota IPI"] = record.get("Aliquota de IPI")
        if reference_key not in ipi_by_reference:
            continue
        matched_references.add(reference_key)
        next_value = round(float(ipi_by_reference[reference_key]), 6)
        if optional_number_value(record.get("Aliquota IPI")) != next_value:
            record["Aliquota IPI"] = next_value
            record["_ipi_atualizado_em"] = now
            updated += 1
    not_found = max(0, len(ipi_by_reference) - len(matched_references))
    save_json(price_table_file(company_id), records)

    summary = {
        "empresa": COMPANIES[company_id],
        "arquivo": file_name,
        "aba": sheet_name,
        "linhas_lidas": len(raw_records),
        "referencias_com_ipi": len(ipi_by_reference),
        "itens_atualizados": updated,
        "referencias_nao_encontradas": not_found,
        "linhas_ignoradas": skipped,
        "ultima_atualizacao_ipi": now,
        "mensagem": "Aliquotas de IPI atualizadas na Tabela de Precos.",
    }
    logs = load_json(import_log_file(company_id), [])
    logs.append({"tipo": "tabela_precos_ipi", **summary})
    save_json(import_log_file(company_id), logs)
    return summary


def products_export_xlsx(company_id: str) -> bytes:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    records = sorted(
        load_json(products_file(company_id), []),
        key=lambda record: (
            normalize_text(record.get("Referencia")),
            normalize_text(record.get("ID")),
        ),
    )
    return export_records_xlsx(records, [key for key, _label in PRODUCT_TABLE_COLUMNS], "Cadastro de Produtos")


def clients_export_xlsx(company_id: str, query: str = "") -> bytes:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    records = [
        record
        for record in load_json(clients_file(company_id), [])
        if value_matches(record, query)
    ]
    return export_records_xlsx(records, [key for key, _label in CLIENT_TABLE_COLUMNS], "Clientes")


def price_table_records_with_cost(company_id: str, save: bool = False) -> list[dict]:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    records = load_json(price_table_file(company_id), [])
    costing_by_reference = costing_info_by_reference(company_id)
    updated = False
    now = datetime.now().isoformat(timespec="seconds")
    for record in records:
        if record.get("Aliquota IPI") in (None, "") and record.get("Aliquota de IPI") not in (None, ""):
            record["Aliquota IPI"] = record.get("Aliquota de IPI")
            updated = True
        costing_info = classification_for_reference(record.get("PR_COD"), costing_by_reference)
        cost = costing_info.get("custo")
        next_value = round(float(cost), 6) if cost is not None else ""
        if record.get("Custo BRL Indice") != next_value:
            record["Custo BRL Indice"] = next_value
            record["_custo_tabela_atualizado_em"] = now
            updated = True
    if save and updated:
        save_json(price_table_file(company_id), records)
    return records


def costing_info_by_reference(company_id: str) -> dict:
    info = {}
    for record in load_json(costing_file(company_id), []):
        key = normalize_text(record.get("Referencia"))
        if not key:
            continue
        info[key] = {
            "custo": optional_number_value(record.get("Custo BRL Indice")),
            "AGRP": normalize_record(record.get("AGRP")),
            "Origem": normalize_record(record.get("Origem")) or "Importado",
        }
    for record in load_json(costing_fabricated_file(company_id), []):
        key = normalize_text(record.get("Referencia"))
        if not key or key in info:
            continue
        info[key] = {
            "custo": optional_number_value(record.get("Custo BRL Indice")),
            "AGRP": normalize_record(record.get("AGRP")),
            "Origem": normalize_record(record.get("Origem")) or "Fabricado BR",
        }
    return info


def all_costing_info_by_reference(preferred_company_id: str = "") -> dict:
    ordered_companies = []
    if preferred_company_id in COMPANIES:
        ordered_companies.append(preferred_company_id)
    ordered_companies.extend(company_id for company_id in COMPANIES if company_id not in ordered_companies)

    info = {}
    for company_id in ordered_companies:
        for reference, record in costing_info_by_reference(company_id).items():
            if reference and reference not in info:
                info[reference] = record
    return info


def missing_product_classification(value) -> bool:
    text = str(normalize_record(value) or "").strip().upper()
    return not text or text in {"-1", "SEM AGRP", "SEM CADASTRO DE CUSTEIO"}


def usable_product_classification(agrp, origin) -> bool:
    return (
        not missing_product_classification(agrp)
        and not missing_product_classification(origin)
    )


def reference_lookup_keys(value) -> list[str]:
    original = str(normalize_record(value) or "").strip()
    keys = []

    def add(candidate):
        key = normalize_text(candidate)
        if key and key not in keys:
            keys.append(key)

    add(original)
    candidates = [original]
    for _index in range(4):
        next_candidates = []
        for candidate in candidates:
            stripped_und = re.sub(r"(?i)(?:[-_\s]+UND)$", "", candidate).strip()
            stripped_numeric = re.sub(r"(?:[-_\s]+[0-9]+)$", "", candidate).strip()
            for stripped in (stripped_und, stripped_numeric):
                if stripped and stripped != candidate:
                    add(stripped)
                    next_candidates.append(stripped)
        candidates = next_candidates
        if not candidates:
            break
    return keys


def classification_for_reference(reference, costing_by_reference: dict) -> dict:
    for key in reference_lookup_keys(reference):
        info = costing_by_reference.get(key)
        if info:
            return info
    return {}


def sales_classification_info_by_reference(sales: list[dict]) -> dict:
    info = {}
    for sale in sales:
        if not usable_product_classification(sale.get("AGRP"), sale.get("Origem")):
            continue
        reference = sale.get("PR_COD") or sale.get("Referencia")
        for key in reference_lookup_keys(reference):
            if key and key not in info:
                info[key] = {
                    "AGRP": normalize_record(sale.get("AGRP")),
                    "Origem": normalize_record(sale.get("Origem")),
                    "fonte": "vendas_classificadas",
                }
    return info


def refresh_sales_classification(company_id: str, save: bool = True) -> dict:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    sales = load_json(sales_file(company_id), [])
    costing_by_reference = all_costing_info_by_reference(company_id)
    sales_by_reference = sales_classification_info_by_reference(sales)
    updated_rows = 0
    agrp_updated = 0
    origin_updated = 0
    no_costing = 0
    used_sales_reference = 0
    now = datetime.now().isoformat(timespec="seconds")

    for sale in sales:
        costing_info = classification_for_reference(sale.get("PR_COD"), costing_by_reference)
        source = "custeio"
        if not costing_info:
            costing_info = classification_for_reference(sale.get("PR_COD"), sales_by_reference)
            source = "vendas_classificadas"
        row_updated = False

        if costing_info:
            if source == "vendas_classificadas":
                used_sales_reference += 1
            agrp = costing_info.get("AGRP")
            origin = costing_info.get("Origem")
            if agrp and normalize_record(sale.get("AGRP")) != agrp:
                sale["AGRP"] = agrp
                agrp_updated += 1
                row_updated = True
            if origin and normalize_record(sale.get("Origem")) != origin:
                sale["Origem"] = origin
                origin_updated += 1
                row_updated = True
        else:
            no_costing += 1
            if missing_product_classification(sale.get("AGRP")):
                sale["AGRP"] = "Sem AGRP"
                agrp_updated += 1
                row_updated = True
            if missing_product_classification(sale.get("Origem")):
                sale["Origem"] = "Sem cadastro de custeio"
                origin_updated += 1
                row_updated = True

        if row_updated:
            sale["_classificacao_atualizada_em"] = now
            sale["_classificacao_fonte"] = source
            updated_rows += 1

    if save and updated_rows:
        save_json(sales_file(company_id), sales)

    return {
        "itens_verificados": len(sales),
        "itens_reclassificados": updated_rows,
        "agrp_atualizados": agrp_updated,
        "origem_atualizadas": origin_updated,
        "itens_sem_custeio": no_costing,
        "itens_corrigidos_por_vendas_ja_classificadas": used_sales_reference,
        "atualizado_em": now,
    }


def stock_summary_payload(company_id: str):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    records = load_json(stock_file(company_id), [])
    total_value = sum(number_value(record.get("Total")) for record in records)
    with_cost = sum(1 for record in records if number_value(record.get("V.Unitario")) > 0)
    return {
        "empresa": COMPANIES[company_id],
        "total_sku": len(records),
        "sku_com_custo": with_cost,
        "sku_sem_custo": len(records) - with_cost,
        "valor_total_estoque": round(total_value, 3),
    }


def recalculate_stock_value(company_id: str):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    records = load_json(stock_file(company_id), [])
    costing_by_reference = costing_info_by_reference(company_id)
    updated = 0
    without_cost = 0
    without_quantity = 0
    now = datetime.now().isoformat(timespec="seconds")

    for record in records:
        key = normalize_text(record.get("Referencia"))
        quantity = optional_number_value(record.get("Quantidade"))
        costing_info = costing_by_reference.get(key)
        unit_cost = costing_info.get("custo") if costing_info else None
        if costing_info:
            record["AGRP"] = costing_info.get("AGRP") or "Sem AGRP"
            record["Origem"] = costing_info.get("Origem")
        else:
            record["AGRP"] = record.get("AGRP") or "Sem AGRP"
            record["Origem"] = record.get("Origem") or "Sem cadastro de custeio"
        if unit_cost is None:
            record["V.Unitario"] = 0.0
            record["Total"] = 0.0
            without_cost += 1
            continue
        if quantity is None:
            record["V.Unitario"] = round(unit_cost, 3)
            record["Total"] = 0.0
            without_quantity += 1
            continue
        record["V.Unitario"] = round(unit_cost, 3)
        record["Total"] = round(unit_cost * quantity, 3)
        record["_valor_recalculado_em"] = now
        updated += 1

    save_json(stock_file(company_id), records)
    return {
        "empresa": COMPANIES[company_id],
        "referencias_estoque": len(records),
        "referencias_atualizadas": updated,
        "referencias_sem_custo": without_cost,
        "referencias_sem_quantidade": without_quantity,
        "valor_total_estoque": round(sum(number_value(record.get("Total")) for record in records), 3),
        "recalculado_em": now,
    }


def missing_stock_cost_export_xlsx(company_id: str) -> bytes:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    costing_references = {
        normalize_text(record.get("Referencia"))
        for record in load_json(costing_file(company_id), [])
        if normalize_text(record.get("Referencia"))
    }
    rows_by_reference = {}
    for record in load_json(stock_file(company_id), []):
        key = normalize_text(record.get("Referencia"))
        if not key or key in costing_references or key in rows_by_reference:
            continue
        rows_by_reference[key] = {
            "Referencia": normalize_record(record.get("Referencia")),
            "Descricao item": normalize_record(record.get("Descricao")),
            "AGRP": normalize_record(record.get("AGRP")),
            "Origem": "Importado",
            "Preco Dollar": "",
            "Custo BRL Indice": "",
            "Custo BRL Ultima Importacao": "",
        }
    rows = sorted(rows_by_reference.values(), key=lambda item: normalize_text(item.get("Referencia")))
    return export_records_xlsx(rows, [key for key, _label in COSTING_TABLE_COLUMNS], "Log Sem Custo")


def import_costing(company_id: str, file_name: str, file_bytes: bytes, indice_custeio):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    indice = optional_number_value(indice_custeio)
    if indice is None:
        raise ValueError("Informe o Indice de Custeio.")

    raw_records, sheet_name = read_spreadsheet_records(file_name, file_bytes)
    skipped_origin = 0
    skipped_reference = 0
    skipped_duplicates = 0
    skipped_existing = 0
    updated_missing_prices = 0
    seen_references = set()
    existing_records = load_json(costing_file(company_id), [])
    existing_by_reference = {
        normalize_text(record.get("Referencia")): record
        for record in existing_records
        if normalize_text(record.get("Referencia"))
    }
    imported_records = list(existing_records)

    for raw in raw_records:
        origin = str(normalize_record(get_field(raw, "Origem")) or "").strip()
        if normalize_text(origin) != "IMPORTADO":
            skipped_origin += 1
            continue

        reference = normalize_record(get_field(raw, "Referencia", "Refer\u00eancia"))
        reference = str(reference or "").strip()
        if not reference:
            skipped_reference += 1
            continue

        key = normalize_text(reference)
        if key in seen_references:
            skipped_duplicates += 1
            continue
        seen_references.add(key)
        dollar_price = optional_number_value(get_field(raw, "Preco Dollar", "Pre\u00e7o Dollar"))
        indexed_cost = round(indice * dollar_price, 6) if dollar_price is not None else None
        if key in existing_by_reference:
            existing_record = existing_by_reference[key]
            if optional_number_value(existing_record.get("Preco Dollar")) is None and dollar_price is not None:
                existing_record["Preco Dollar"] = dollar_price
                existing_record["Custo BRL Indice"] = indexed_cost
                existing_record["Origem"] = existing_record.get("Origem") or "Importado"
                if not normalize_record(existing_record.get("Descricao item")):
                    existing_record["Descricao item"] = normalize_record(get_field(raw, "Descricao item", "Descri\u00e7\u00e3o item", "Descricao"))
                if not normalize_record(existing_record.get("AGRP")):
                    existing_record["AGRP"] = normalize_record(get_field(raw, "AGRP"))
                if not normalize_record(existing_record.get("Custo BRL Ultima Importacao")):
                    existing_record["Custo BRL Ultima Importacao"] = normalize_record(get_field(
                        raw,
                        "Custo BRL Ultima Importacao",
                "Custo BRL Ultima Importa\u00e7\u00e3o",
                    ))
                existing_record["_indice_custeio"] = indice
                existing_record["_atualizado_em"] = datetime.now().isoformat(timespec="seconds")
                updated_missing_prices += 1
            else:
                skipped_existing += 1
            continue

        new_record = {
            "Referencia": reference,
            "Descricao item": normalize_record(get_field(raw, "Descricao item", "Descri\u00e7\u00e3o item", "Descricao", "Descri\u00e7\u00e3o")),
            "AGRP": normalize_record(get_field(raw, "AGRP")),
            "Origem": "Importado",
            "Preco Dollar": dollar_price,
            "Custo BRL Indice": indexed_cost,
            "Custo BRL Ultima Importacao": normalize_record(get_field(
                raw,
                "Custo BRL Ultima Importacao",
                "Custo BRL Ultima Importa\u00e7\u00e3o",
            )),
            "_empresa_id": company_id,
            "_empresa_nome": COMPANIES[company_id],
            "_indice_custeio": indice,
            "_atualizado_em": datetime.now().isoformat(timespec="seconds"),
        }
        imported_records.append(new_record)
        existing_by_reference[key] = new_record

    imported_records.sort(key=lambda item: normalize_text(item.get("Referencia")))
    save_json(costing_file(company_id), imported_records)
    save_json(costing_config_file(company_id), {
        "indice_custeio": indice,
        "atualizado_em": datetime.now().isoformat(timespec="seconds"),
    })

    summary = {
        "empresa": COMPANIES[company_id],
        "arquivo": file_name,
        "aba": sheet_name,
        "indice_custeio": indice,
        "linhas_lidas": len(raw_records),
        "referencias_importadas": len(imported_records) - len(existing_records),
        "referencias_atualizadas_preco": updated_missing_prices,
        "total_referencias_base": len(imported_records),
        "linhas_ignoradas_origem": skipped_origin,
        "linhas_ignoradas_sem_referencia": skipped_reference,
        "linhas_ignoradas_duplicadas": skipped_duplicates,
        "referencias_ja_existentes": skipped_existing,
        "importado_em": datetime.now().isoformat(timespec="seconds"),
    }
    logs = load_json(costing_import_log_file(company_id), [])
    logs.append(summary)
    save_json(costing_import_log_file(company_id), logs)
    return summary


def recalculate_costing(company_id: str, indice_custeio):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    indice = optional_number_value(indice_custeio)
    if indice is None:
        raise ValueError("Informe o Indice de Custeio.")

    records = load_json(costing_file(company_id), [])
    now = datetime.now().isoformat(timespec="seconds")
    recalculated = 0
    without_price = 0
    for record in records:
        dollar_price = optional_number_value(record.get("Preco Dollar"))
        if dollar_price is None:
            record["Custo BRL Indice"] = None
            without_price += 1
        else:
            record["Custo BRL Indice"] = round(indice * dollar_price, 6)
            recalculated += 1
        record["_indice_custeio"] = indice
        record["_atualizado_em"] = now

    save_json(costing_file(company_id), records)
    save_json(costing_config_file(company_id), {
        "indice_custeio": indice,
        "atualizado_em": now,
    })
    return {
        "empresa": COMPANIES[company_id],
        "indice_custeio": indice,
        "referencias_recalculadas": recalculated,
        "referencias_sem_preco_dollar": without_price,
        "total_referencias_base": len(records),
        "recalculado_em": now,
    }


def import_costing_fabricated(company_id: str, file_name: str, file_bytes: bytes, cost_increase):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    increase = optional_number_value(cost_increase)
    if increase is None:
        increase = 0

    raw_records, sheet_name = read_spreadsheet_records(file_name, file_bytes)
    skipped_origin = 0
    skipped_reference = 0
    skipped_duplicates = 0
    seen_references = set()
    existing_records = load_json(costing_fabricated_file(company_id), [])
    existing_by_reference = {
        normalize_text(record.get("Referencia")): record
        for record in existing_records
        if normalize_text(record.get("Referencia"))
    }
    records = list(existing_records)
    imported = 0
    updated = 0
    now = datetime.now().isoformat(timespec="seconds")

    for raw in raw_records:
        origin = str(normalize_record(get_field(raw, "Origem")) or "").strip()
        if normalize_text(origin) != "FABRICADO BR":
            skipped_origin += 1
            continue

        reference = normalize_record(get_field(raw, "Referencia", "Refer\u00eancia", "Codigo do produto", "C\u00f3digo do produto"))
        reference = str(reference or "").strip()
        if not reference:
            skipped_reference += 1
            continue

        key = normalize_text(reference)
        if key in seen_references:
            skipped_duplicates += 1
            continue
        seen_references.add(key)

        base_cost = last_cost if last_cost is not None else price_value
        indexed_cost = round(base_cost * (1 + (increase / 100)), 6) if base_cost is not None else None

        payload = {
            "Referencia": reference,
            "Descricao item": normalize_record(get_field(raw, "Descricao item", "Descri\u00e7\u00e3o item", "Descricao", "Descri\u00e7\u00e3o")),
            "AGRP": normalize_record(get_field(raw, "AGRP")),
            "Origem": "Fabricado BR",
            "Preco Dollar": price_value,
            "Custo BRL Indice": indexed_cost,
            "Custo BRL Ultima Importacao": last_cost,
            "_empresa_id": company_id,
            "_empresa_nome": COMPANIES[company_id],
            "_percentual_aumento_custo": increase,
            "_atualizado_em": now,
        }

        if key in existing_by_reference:
            existing_by_reference[key].update(payload)
            updated += 1
        else:
            records.append(payload)
            existing_by_reference[key] = payload
            imported += 1

    records.sort(key=lambda item: normalize_text(item.get("Referencia")))
    save_json(costing_fabricated_file(company_id), records)

    return {
        "empresa": COMPANIES[company_id],
        "arquivo": file_name,
        "aba": sheet_name,
        "percentual_aumento_custo": increase,
        "linhas_lidas": len(raw_records),
        "referencias_importadas": imported,
        "referencias_atualizadas": updated,
        "total_referencias_base": len(records),
        "linhas_ignoradas_origem": skipped_origin,
        "linhas_ignoradas_sem_referencia": skipped_reference,
        "linhas_ignoradas_duplicadas": skipped_duplicates,
        "importado_em": now,
    }


def recalculate_costing_fabricated(company_id: str, cost_increase):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    increase = optional_number_value(cost_increase)
    if increase is None:
        raise ValueError("Informe o percentual de aumento de custo.")

    records = load_json(costing_fabricated_file(company_id), [])
    now = datetime.now().isoformat(timespec="seconds")
    recalculated = 0
    without_cost = 0
    for record in records:
        base_cost = optional_number_value(record.get("Custo BRL Ultima Importacao"))
        if base_cost is None:
            base_cost = optional_number_value(record.get("Preco Dollar"))
        if base_cost is None:
            base_cost = optional_number_value(record.get("Custo BRL Indice"))
        if base_cost is None:
            record["Custo BRL Indice"] = None
            without_cost += 1
        else:
            record["Custo BRL Indice"] = round(base_cost * (1 + (increase / 100)), 6)
            recalculated += 1
        record["_percentual_aumento_custo"] = increase
        record["_atualizado_em"] = now

    save_json(costing_fabricated_file(company_id), records)
    return {
        "empresa": COMPANIES[company_id],
        "percentual_aumento_custo": increase,
        "referencias_recalculadas": recalculated,
        "referencias_sem_custo_base": without_cost,
        "total_referencias_base": len(records),
        "recalculado_em": now,
    }


def br_decimal(value) -> str:
    number = optional_number_value(value)
    if number is None:
        return ""
    text = f"{number:.3f}"
    return text.replace(".", ",")


def costing_export_csv(company_id: str) -> bytes:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    records = load_json(costing_file(company_id), [])
    output = BytesIO()
    output.write(b"\xef\xbb\xbf")
    text_output = TextIOWrapper(output, encoding="utf-8", newline="", write_through=True)
    writer = csv.writer(text_output, delimiter=";")
    output.write("sep=;\r\n".encode("utf-8"))
    writer.writerow([label for _key, label in COSTING_TABLE_COLUMNS])
    for record in records:
        writer.writerow([
            br_decimal(record.get(key)) if key in ("Preco Dollar", "Custo BRL Indice", "Custo BRL Ultima Importacao")
            else record.get(key) if record.get(key) is not None else ""
            for key, _label in COSTING_TABLE_COLUMNS
        ])
    text_output.flush()
    return output.getvalue()


def costing_export_xlsx(company_id: str, costing_kind: str) -> bytes:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    if costing_kind == "fabricated":
        records = load_json(costing_fabricated_file(company_id), [])
        sheet_name = "Custeio Fabricado"
    else:
        records = load_json(costing_file(company_id), [])
        sheet_name = "Custeio Importados"

    rows = [
        {
            label: export_cell_value(record.get(key))
            for key, label in COSTING_TABLE_COLUMNS
        }
        for record in records
    ]
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe = pd.DataFrame(rows, columns=[label for _key, label in COSTING_TABLE_COLUMNS])
        dataframe.to_excel(writer, index=False, sheet_name=sheet_name)
        worksheet = writer.sheets[sheet_name]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 14), 55)
    return output.getvalue()


def export_records_xlsx(records: list[dict], columns: list[str], sheet_name: str) -> bytes:
    rows = [
        {
            column: export_cell_value(record.get(column))
            for column in columns
        }
        for record in records
    ]
    output = BytesIO()
    clean_sheet_name = sheet_name[:31] or "Relatorio"
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe = pd.DataFrame(rows, columns=columns)
        dataframe.to_excel(writer, index=False, sheet_name=clean_sheet_name)
        worksheet = writer.sheets[clean_sheet_name]
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions
        for column_cells in worksheet.columns:
            max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 14), 55)
    return output.getvalue()


def export_csv(records: list[dict], columns: list[str]) -> bytes:
    output = BytesIO()
    output.write(b"\xef\xbb\xbf")
    text_output = TextIOWrapper(output, encoding="utf-8", newline="", write_through=True)
    writer = csv.writer(text_output, delimiter=";")
    output.write("sep=;\r\n".encode("utf-8"))
    writer.writerow(columns)
    for record in records:
        writer.writerow([export_cell_value(record.get(column)) for column in columns])
    text_output.flush()
    return output.getvalue()


def export_cell_value(value):
    value = normalize_record(value)
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return br_decimal(value)
    return value


def parse_br_date(value):
    value = str(value or "").strip()
    if not value:
        return None
    for date_format in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(value, date_format).date()
        except ValueError:
            continue
    return None


def sale_date_value(record: dict):
    value = normalize_record(record.get("NF_EMI"))
    if not value:
        return None
    try:
        return pd.to_datetime(value).date()
    except Exception:
        return None


def sales_export_csv(company_id: str, filters: dict | None = None) -> bytes:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    filters = filters or {}
    records = load_json(sales_file(company_id), [])
    start_date = parse_br_date(filters.get("start_date"))
    end_date = parse_br_date(filters.get("end_date"))
    client_query = normalize_text(filters.get("client"))
    product_query = normalize_text(filters.get("product"))

    filtered_records = []
    for record in records:
        sale_date = sale_date_value(record)
        if start_date and (not sale_date or sale_date < start_date):
            continue
        if end_date and (not sale_date or sale_date > end_date):
            continue
        if client_query:
            client_text = normalize_text(
                f"{record.get('CL_NOM') or ''} {record.get('ID_CL') or ''} {record.get('CGC') or ''}"
            )
            if client_query not in client_text:
                continue
        if product_query:
            product_text = normalize_text(
                f"{record.get('PR_COD') or ''} {record.get('PR_DES') or ''}"
            )
            if product_query not in product_text:
                continue
        filtered_records.append(record)

    preferred_columns = [key for key, _label in SALES_TABLE_COLUMNS]
    all_columns = []
    for column in preferred_columns:
        if any(column in record for record in filtered_records):
            all_columns.append(column)
    for record in filtered_records:
        for column in record:
            if column.startswith("_") or column in all_columns:
                continue
            all_columns.append(column)
    return export_csv(filtered_records, all_columns)


def sales_export_xlsx(company_id: str, filters: dict | None = None) -> bytes:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    filters = filters or {}
    records = load_json(sales_file(company_id), [])
    start_date = parse_br_date(filters.get("start_date"))
    end_date = parse_br_date(filters.get("end_date"))
    client_query = normalize_text(filters.get("client"))
    product_query = normalize_text(filters.get("product"))

    filtered_records = []
    for record in records:
        sale_date = sale_date_value(record)
        if start_date and (not sale_date or sale_date < start_date):
            continue
        if end_date and (not sale_date or sale_date > end_date):
            continue
        if client_query:
            client_text = normalize_text(
                f"{record.get('CL_NOM') or ''} {record.get('ID_CL') or ''} {record.get('CGC') or ''}"
            )
            if client_query not in client_text:
                continue
        if product_query:
            product_text = normalize_text(
                f"{record.get('PR_COD') or ''} {record.get('PR_DES') or ''}"
            )
            if product_query not in product_text:
                continue
        filtered_records.append(record)

    preferred_columns = [key for key, _label in SALES_TABLE_COLUMNS]
    all_columns = []
    for column in preferred_columns:
        if any(column in record for record in filtered_records):
            all_columns.append(column)
    for record in filtered_records:
        for column in record:
            if column.startswith("_") or column in all_columns:
                continue
            all_columns.append(column)
    return export_records_xlsx(filtered_records, all_columns, "Vendas")


def recalculate_sales_cost(company_id: str, start_date_value: str, end_date_value: str):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    start_date = parse_br_date(start_date_value)
    end_date = parse_br_date(end_date_value)
    if not start_date or not end_date:
        raise ValueError("Informe Data inicial e Data final para recalcular o custo.")
    if end_date < start_date:
        raise ValueError("Data final nao pode ser menor que a data inicial.")

    sales = load_json(sales_file(company_id), [])
    costing_records = load_json(costing_file(company_id), [])
    costing_by_reference = {
        normalize_text(record.get("Referencia")): optional_number_value(record.get("Custo BRL Indice"))
        for record in costing_records
        if normalize_text(record.get("Referencia"))
    }

    in_period = 0
    updated = 0
    not_found = 0
    without_quantity = 0
    now = datetime.now().isoformat(timespec="seconds")

    for sale in sales:
        sale_date = sale_date_value(sale)
        if not sale_date or sale_date < start_date or sale_date > end_date:
            continue

        in_period += 1
        sale["PR_CUSN"] = 0.0
        product_key = normalize_text(sale.get("PR_COD"))
        unit_cost = costing_by_reference.get(product_key)
        quantity = optional_number_value(sale.get("PR_QTD"))

        if unit_cost is None:
            not_found += 1
            continue
        if quantity is None:
            without_quantity += 1
            continue

        sale["PR_CUSN"] = round(unit_cost * quantity, 3)
        sale["_custo_recalculado_em"] = now
        updated += 1

    save_json(sales_file(company_id), sales)
    return {
        "empresa": COMPANIES[company_id],
        "data_inicial": start_date.strftime("%d/%m/%Y"),
        "data_final": end_date.strftime("%d/%m/%Y"),
        "itens_periodo": in_period,
        "itens_atualizados": updated,
        "itens_sem_custeio": not_found,
        "itens_sem_quantidade": without_quantity,
        "recalculado_em": now,
    }


def missing_sales_cost_export_xlsx(company_id: str, start_date_value: str, end_date_value: str) -> bytes:
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    start_date = parse_br_date(start_date_value)
    end_date = parse_br_date(end_date_value)
    if not start_date or not end_date:
        raise ValueError("Informe Data inicial e Data final para gerar o log sem custo.")
    if end_date < start_date:
        raise ValueError("Data final nao pode ser menor que a data inicial.")

    sales = load_json(sales_file(company_id), [])
    costing_references = {
        normalize_text(record.get("Referencia"))
        for record in load_json(costing_file(company_id), [])
        if normalize_text(record.get("Referencia"))
    }
    rows_by_reference = {}
    for sale in sales:
        sale_date = sale_date_value(sale)
        if not sale_date or sale_date < start_date or sale_date > end_date:
            continue
        product_key = normalize_text(sale.get("PR_COD"))
        if not product_key or product_key in costing_references or product_key in rows_by_reference:
            continue
        rows_by_reference[product_key] = {
            "Referencia": normalize_record(sale.get("PR_COD")),
            "Descricao item": normalize_record(sale.get("PR_DES")),
            "AGRP": normalize_record(sale.get("AGRP")),
            "Origem": "Importado",
            "Preco Dollar": "",
            "Custo BRL Indice": "",
            "Custo BRL Ultima Importacao": "",
        }
    rows = sorted(rows_by_reference.values(), key=lambda item: normalize_text(item.get("Referencia")))
    return export_records_xlsx(rows, [key for key, _label in COSTING_TABLE_COLUMNS], "Log Sem Custo")


def import_sales(company_id: str, file_name: str, file_bytes: bytes):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    workbook = pd.ExcelFile(BytesIO(file_bytes))
    sheet_name = "LucratividadeVenda" if "LucratividadeVenda" in workbook.sheet_names else workbook.sheet_names[0]
    frame = pd.read_excel(workbook, sheet_name=sheet_name)
    frame = frame.where(pd.notnull(frame), None)

    existing = load_json(sales_file(company_id), [])
    existing_keys = {item.get("_dedupe_key") or record_key(company_id, item) for item in existing}
    costing_by_reference = all_costing_info_by_reference(company_id)

    inserted = []
    ignored = 0

    for raw_record in frame.to_dict(orient="records"):
        record = {str(key): normalize_record(value) for key, value in raw_record.items()}
        costing_info = classification_for_reference(record.get("PR_COD"), costing_by_reference)
        if missing_product_classification(record.get("AGRP")) and costing_info.get("AGRP"):
            record["AGRP"] = costing_info.get("AGRP")
        if missing_product_classification(record.get("Origem")) and costing_info.get("Origem"):
            record["Origem"] = costing_info.get("Origem")
        if missing_product_classification(record.get("AGRP")):
            record["AGRP"] = "Sem AGRP"
        if missing_product_classification(record.get("Origem")):
            record["Origem"] = "Sem cadastro de custeio"
        key = record_key(company_id, record)

        if key in existing_keys:
            ignored += 1
            continue

        record["_empresa_id"] = company_id
        record["_empresa_nome"] = COMPANIES[company_id]
        record["_dedupe_key"] = key
        record["_importado_em"] = datetime.now().isoformat(timespec="seconds")
        inserted.append(record)
        existing_keys.add(key)

    all_records = existing + inserted
    if inserted:
        try:
            save_json(sales_file(company_id), all_records)
        except PermissionError as exc:
            raise ValueError(
                "Nao foi possivel gravar as notas fiscais porque o Windows bloqueou o arquivo de vendas. "
                "Feche outras janelas do CRM, planilhas, backups ou sincronizacao que possam estar usando a pasta data e tente importar novamente."
            ) from exc
        classification_summary = refresh_sales_classification(company_id)
        all_records = load_json(sales_file(company_id), [])
        sync_vendors_from_sales(company_id)
    else:
        classification_summary = {"atualizados": 0, "sem_classificacao": 0}

    notes = {
        str(item.get("ID_NF") or item.get("NF_NUM") or "")
        for item in inserted
        if item.get("ID_NF") or item.get("NF_NUM")
    }
    summary = {
        "empresa": COMPANIES[company_id],
        "arquivo": file_name,
        "aba": sheet_name,
        "linhas_lidas": int(len(frame.index)),
        "linhas_importadas": len(inserted),
        "linhas_ignoradas": ignored,
        "notas_importadas": len(notes),
        "total_registros_empresa": len(all_records),
        "classificacao_atualizada": classification_summary,
        "importado_em": datetime.now().isoformat(timespec="seconds"),
    }

    logs = load_json(import_log_file(company_id), [])
    logs.append(summary)
    try:
        save_json(import_log_file(company_id), logs)
    except PermissionError:
        summary["aviso_log"] = "Importacao processada, mas o Windows bloqueou a gravacao do log de importacao."

    return summary


def import_clients(company_id: str, file_name: str, file_bytes: bytes):
    if company_id not in COMPANIES:
        raise ValueError("Empresa invalida.")

    workbook = pd.ExcelFile(BytesIO(file_bytes))
    sheet_name = workbook.sheet_names[0]
    frame = pd.read_excel(workbook, sheet_name=sheet_name)
    frame = frame.where(pd.notnull(frame), None)

    existing = load_json(clients_file(company_id), [])
    existing_keys = {item.get("_dedupe_key") or client_key(company_id, item) for item in existing}

    inserted = []
    ignored = 0
    skipped_non_clients = 0

    for raw_record in frame.to_dict(orient="records"):
        record = {str(key): normalize_record(value) for key, value in raw_record.items()}

        if not is_client_record(record):
            skipped_non_clients += 1
            continue

        key = client_key(company_id, record)

        if key in existing_keys:
            ignored += 1
            continue

        record["_empresa_id"] = company_id
        record["_empresa_nome"] = COMPANIES[company_id]
        record["_dedupe_key"] = key
        record["_documento_normalizado"] = normalize_identifier(record.get("CGC"))
        record["_importado_em"] = datetime.now().isoformat(timespec="seconds")
        inserted.append(record)
        existing_keys.add(key)

    all_records = existing + inserted
    save_json(clients_file(company_id), all_records)

    summary = {
        "empresa": COMPANIES[company_id],
        "arquivo": file_name,
        "aba": sheet_name,
        "linhas_lidas": int(len(frame.index)),
        "clientes_importados": len(inserted),
        "clientes_ignorados": ignored,
        "clientes_descartados": skipped_non_clients,
        "total_clientes_empresa": len(all_records),
        "importado_em": datetime.now().isoformat(timespec="seconds"),
    }

    logs = load_json(client_import_log_file(company_id), [])
    logs.append(summary)
    save_json(client_import_log_file(company_id), logs)

    return summary


class CRMHandler(BaseHTTPRequestHandler):
    server_version = "CRM/0.1"

    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path

        if path == "/api/companies":
            return self.send_json([
                {"id": company_id, "name": name}
                for company_id, name in COMPANIES.items()
            ])

        if path == "/api/dashboard/sales-family":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            selected_year = params.get("year", ["all"])[0]
            selected_origin = params.get("origin", ["all"])[0]
            try:
                payload = cached_payload(
                    ("dashboard_sales_family", company_id, selected_year, selected_origin),
                    dashboard_dependencies(company_id),
                    lambda: dashboard_sales_family_payload(company_id, selected_year, selected_origin),
                )
                return self.send_json(payload)
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/table/sales", "/api/table/clients", "/api/table/stock", "/api/table/in-transit", "/api/table/costing", "/api/table/costing-fabricated", "/api/table/prices", "/api/table/products"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            query = params.get("q", [""])[0].strip()
            sort_key = params.get("sort_key", [""])[0]
            sort_dir = params.get("sort_dir", [""])[0]
            try:
                limit = int(params.get("limit", ["200"])[0])
                table_name = path.rsplit("/", 1)[-1]
                return self.send_json(table_payload(company_id, table_name, query, limit, sort_key, sort_dir))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/slow-items":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            query = params.get("q", [""])[0].strip()
            try:
                return self.send_json(slow_items_payload(company_id, query))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/mercado-livre/overview":
            try:
                return self.send_json(mercado_livre_overview_payload())
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/mercado-livre/ads":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["onix"])[0]
            query = params.get("q", [""])[0].strip()
            import_id = params.get("import_id", [""])[0]
            import_type = params.get("type", [""])[0]
            try:
                limit = int(params.get("limit", ["300"])[0])
                return self.send_json(mercado_livre_ads_payload(company_id, query, limit, import_id, import_type))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/mercado-livre/general":
            params = parse_qs(parsed.query)
            query = params.get("q", [""])[0].strip()
            try:
                limit = int(params.get("limit", ["300"])[0])
                return self.send_json(mercado_livre_general_payload(query, limit))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/mercado-livre/sales":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["onix"])[0]
            query = params.get("q", [""])[0].strip()
            try:
                limit = int(params.get("limit", ["300"])[0])
                return self.send_json(mercado_livre_sales_payload(company_id, query, limit))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/mercado-livre/sales-dashboard":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["agrupado"])[0]
            year_filter = params.get("year", ["todos"])[0]
            try:
                return self.send_json(mercado_livre_sales_dashboard_payload(company_id, year_filter))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/export-mercado-livre-summary", "/api/mercado-livre/summary/export"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["onix"])[0]
            try:
                content = mercado_livre_summary_export_xlsx(company_id)
                filename = f"mercado_livre_resumo_{company_id}.xlsx"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/export-mercado-livre-general", "/api/mercado-livre/general/export"):
            try:
                content = mercado_livre_general_export_xlsx()
                filename = "mercado_livre_gestao_geral.xlsx"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/export-slow-items", "/api/export-slow-items/", "/api/slow-items/export"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            query = params.get("q", [""])[0].strip()
            try:
                content = slow_items_export_xlsx(company_id, query)
                filename = f"itens_sem_giro_{company_id}.xlsx"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/export-costing":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            try:
                content = costing_export_xlsx(company_id, "imported")
                filename = f"custeio_importados_{company_id}.xlsx"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/export-costing-xlsx":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            costing_kind = params.get("kind", ["imported"])[0]
            try:
                content = costing_export_xlsx(company_id, costing_kind)
                filename_prefix = "custeio_fabricado" if costing_kind == "fabricated" else "custeio_importados"
                filename = f"{filename_prefix}_{company_id}.xlsx"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/export-sales", "/api/export-sales/", "/api/sales/export"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            try:
                content = sales_export_xlsx(company_id, {
                    "start_date": params.get("start_date", [""])[0],
                    "end_date": params.get("end_date", [""])[0],
                    "client": params.get("client", [""])[0],
                    "product": params.get("product", [""])[0],
                })
                filename = f"notas_fiscais_vendas_{company_id}.xlsx"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/export-stock", "/api/export-stock/", "/api/stock/export"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            try:
                content = stock_export_xlsx(company_id)
                filename = f"estoque_{company_id}.xlsx"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/export-in-transit", "/api/export-in-transit/", "/api/in-transit/export"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            try:
                content = in_transit_export_xlsx(company_id)
                filename = f"em_transito_{company_id}.xlsx"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/export-prices", "/api/export-prices/", "/api/prices/export"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            try:
                content = price_table_export_xlsx(company_id)
                filename = f"tabela_precos_{company_id}.xlsx"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/export-products", "/api/export-products/", "/api/products/export"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            try:
                content = products_export_xlsx(company_id)
                filename = f"cadastro_produtos_{company_id}.xlsx"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/export-clients-xlsx", "/api/export-clients", "/api/clients/export"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            query = params.get("q", [""])[0]
            try:
                content = clients_export_xlsx(company_id, query)
                filename = f"clientes_{company_id}.xlsx"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/export-stock-missing-cost", "/api/export-stock-missing-cost/", "/api/stock/missing-cost"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            try:
                content = missing_stock_cost_export_xlsx(company_id)
                filename = f"log_sem_custo_estoque_{company_id}.xlsx"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/stock-summary", "/api/stock/summary"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            try:
                return self.send_json(stock_summary_payload(company_id))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/recalculate-costing", "/api/recalculate-costing/", "/api/costing/recalculate"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            indice = params.get("indice_custeio", [""])[0]
            try:
                return self.send_json(recalculate_costing(company_id, indice))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/recalculate-costing-fabricated", "/api/recalculate-costing-fabricated/", "/api/costing-fabricated/recalculate"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            increase = params.get("cost_increase", [""])[0]
            try:
                return self.send_json(recalculate_costing_fabricated(company_id, increase))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/recalculate-stock-value", "/api/recalculate-stock-value/", "/api/stock/recalculate-value"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            try:
                return self.send_json(recalculate_stock_value(company_id))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/recalculate-sales-cost", "/api/recalculate-sales-cost/", "/api/sales/recalculate-cost"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            start_date_param = params.get("start_date", [""])[0]
            end_date_param = params.get("end_date", [""])[0]
            try:
                return self.send_json(recalculate_sales_cost(company_id, start_date_param, end_date_param))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/export-sales-missing-cost", "/api/export-sales-missing-cost/", "/api/sales/missing-cost"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            start_date_param = params.get("start_date", [""])[0]
            end_date_param = params.get("end_date", [""])[0]
            try:
                content = missing_sales_cost_export_xlsx(company_id, start_date_param, end_date_param)
                filename = f"log_sem_custo_{company_id}.xlsx"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendors":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            query = params.get("q", [""])[0].strip()
            try:
                return self.send_json(vendors_payload(company_id, query))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendors/assignment-options":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            try:
                return self.send_json(vendor_assignment_options_payload(company_id))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendors/client-search":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            query = params.get("q", [""])[0]
            try:
                return self.send_json(vendor_client_search_payload(company_id, query))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendor-page-links":
            try:
                return self.send_json(vendor_page_links_payload())
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/home/vendor-panel":
            token = self.headers.get("X-Auth-Token", "")
            params = parse_qs(parsed.query)
            company_id = params.get("company", [""])[0]
            vendor_id_value = params.get("vendor_id", [""])[0]
            try:
                return self.send_json(home_vendor_panel_payload(token, company_id, vendor_id_value))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/regions":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            try:
                return self.send_json(region_options_payload(company_id))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendor-goals":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            vendor_id_value = params.get("vendor_id", [""])[0]
            year = params.get("year", [CURRENT_YEAR])[0]
            try:
                payload = cached_payload(
                    ("vendor_goals", company_id, vendor_id_value, str(year)),
                    vendor_dashboard_dependencies(company_id),
                    lambda: vendor_goals_payload(company_id, vendor_id_value, year),
                )
                return self.send_json(payload)
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendor-monthly-items":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            year = params.get("year", [CURRENT_YEAR])[0]
            month = params.get("month", [date.today().month])[0]
            try:
                return self.send_json(vendor_monthly_items_payload(company_id, year, month))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/product-search":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            query = params.get("q", [""])[0]
            try:
                return self.send_json(product_search_payload(company_id, query))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/quotes/context":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            vendor_id_value = params.get("vendor_id", [""])[0]
            try:
                return self.send_json(quote_context_payload(company_id, vendor_id_value))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/quotes/client-search":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            query = params.get("q", [""])[0]
            try:
                return self.send_json(quote_client_search_payload(company_id, query))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/quotes/product-search":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            query = params.get("q", [""])[0]
            try:
                return self.send_json(quote_product_search_payload(company_id, query))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/quick-consult/products":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            query = params.get("q", [""])[0]
            try:
                return self.send_json(quick_consult_products_payload(company_id, query))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/quotes/item-preview":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            reference = params.get("reference", [""])[0]
            quantity = params.get("quantity", ["1"])[0]
            discount = params.get("discount", ["0"])[0]
            try:
                return self.send_json(quote_item_preview_payload(company_id, reference, quantity, discount))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/quotes/search":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            vendor_id_value = params.get("vendor_id", [""])[0]
            query = params.get("q", [""])[0]
            limit = int(optional_number_value(params.get("limit", [20])[0]) or 20)
            page = int(optional_number_value(params.get("page", [1])[0]) or 1)
            try:
                return self.send_json(commercial_document_search_payload(company_id, vendor_id_value, query, "orcamento", limit, page))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/orders/search":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            vendor_id_value = params.get("vendor_id", [""])[0]
            query = params.get("q", [""])[0]
            limit = int(optional_number_value(params.get("limit", [20])[0]) or 20)
            page = int(optional_number_value(params.get("page", [1])[0]) or 1)
            try:
                return self.send_json(commercial_document_search_payload(company_id, vendor_id_value, query, "pedido", limit, page))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/quotes/pdf":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            quote_id = params.get("id", [""])[0]
            try:
                quote = quote_by_id(company_id, quote_id)
                content = quote_pdf_bytes(quote)
                filename = f"orcamento_{quote.get('numero') or quote_id}.pdf"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/pdf")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/orders/pdf":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            order_id = params.get("id", [""])[0]
            try:
                order = order_by_id(company_id, order_id)
                content = quote_pdf_bytes(order)
                filename = f"pedido_{order.get('numero') or order_id}.pdf"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/pdf")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/block-reasons":
            try:
                return self.send_json(block_reasons_payload())
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/client-lookup":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            code = params.get("code", [""])[0]
            try:
                return self.send_json(client_lookup_payload(company_id, code))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/client-block-status":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            code = params.get("code", [""])[0]
            try:
                return self.send_json(client_block_status_payload(company_id, code))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/blocked-clients":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            query = params.get("q", [""])[0].strip()
            try:
                return self.send_json(blocked_clients_payload(company_id, query))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/dashboard/clients-by-uf":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            try:
                payload = cached_payload(
                    ("dashboard_clients_by_uf", company_id),
                    dashboard_dependencies(company_id),
                    lambda: clients_by_uf_payload(company_id),
                )
                return self.send_json(payload)
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/dashboard/sales-annual":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            selected_year = params.get("year", ["all"])[0]
            selected_origin = params.get("origin", ["all"])[0]
            selected_agrp = params.get("agrp", ["all"])[0]
            try:
                payload = cached_payload(
                    ("dashboard_sales_annual", company_id, selected_year, selected_origin, selected_agrp),
                    dashboard_dependencies(company_id),
                    lambda: dashboard_sales_annual_payload(company_id, selected_year, selected_origin, selected_agrp),
                )
                return self.send_json(payload)
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/dashboard/sales-family":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            selected_year = params.get("year", ["all"])[0]
            selected_origin = params.get("origin", ["all"])[0]
            try:
                payload = cached_payload(
                    ("dashboard_sales_family", company_id, selected_year, selected_origin),
                    dashboard_dependencies(company_id),
                    lambda: dashboard_sales_family_payload(company_id, selected_year, selected_origin),
                )
                return self.send_json(payload)
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/dashboard/vendor-regions", "/api/vendor-regions-data"):
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            vendor_id_value = params.get("vendor_id", [""])[0]
            token = self.headers.get("X-Auth-Token", "")
            try:
                dashboard_vendor_access_context(company_id, vendor_id_value, token)
                payload = cached_payload(
                    ("vendor_regions", company_id, vendor_id_value),
                    vendor_dashboard_dependencies(company_id),
                    lambda: vendor_regions_payload(company_id, vendor_id_value),
                )
                return self.send_json(payload)
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendor-page-data":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            vendor_id_value = params.get("vendor_id", [""])[0]
            try:
                payload = cached_payload(
                    ("vendor_page", company_id, vendor_id_value),
                    vendor_dashboard_dependencies(company_id),
                    lambda: vendor_page_payload(company_id, vendor_id_value),
                )
                return self.send_json(payload)
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendor-region-clients":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            vendor_id_value = params.get("vendor_id", [""])[0]
            try:
                return self.send_json(vendor_region_clients_payload(company_id, vendor_id_value, params))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendor-day-by-day":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            vendor_id_value = params.get("vendor_id", [""])[0]
            day_key = params.get("date", [""])[0] or date.today().isoformat()
            try:
                payload = cached_payload(
                    ("vendor_day_by_day", company_id, vendor_id_value, day_key),
                    vendor_dashboard_dependencies(company_id),
                    lambda: vendor_day_by_day_payload(company_id, vendor_id_value, day_key),
                )
                return self.send_json(payload)
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendor-day-by-day/client":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            vendor_id_value = params.get("vendor_id", [""])[0]
            client_id = params.get("client_id", [""])[0]
            day_key = params.get("date", [""])[0]
            try:
                return self.send_json(vendor_day_by_day_client_payload(company_id, vendor_id_value, client_id, day_key))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendor-day-by-day/agenda":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            vendor_id_value = params.get("vendor_id", [""])[0]
            try:
                return self.send_json(vendor_day_by_day_agenda_payload(company_id, vendor_id_value))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendor-day-by-day/daily-contacts":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            vendor_id_value = params.get("vendor_id", [""])[0]
            try:
                return self.send_json(vendor_day_by_day_daily_contacts_payload(company_id, vendor_id_value, params))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendor-day-by-day/report":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            vendor_id_value = params.get("vendor_id", [""])[0]
            try:
                return self.send_json(vendor_day_by_day_report_payload(company_id, vendor_id_value, params))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/follow-up/options":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            token = self.headers.get("X-Auth-Token", "")
            try:
                return self.send_json(follow_up_options_payload(company_id, token))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/follow-up":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            token = self.headers.get("X-Auth-Token", "")
            try:
                return self.send_json(follow_up_payload(company_id, params, token))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/follow-up/export-xlsx":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            token = self.headers.get("X-Auth-Token", "") or params.get("token", [""])[0]
            try:
                content = follow_up_export_xlsx(company_id, params, token)
                filename = f"follow_up_atendimentos_{company_id}.xlsx"
                self.send_response(HTTPStatus.OK)
                self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
                self.send_header("Cache-Control", "no-store")
                self.send_header("Content-Length", str(len(content)))
                self.end_headers()
                self.wfile.write(content)
                return
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendor-region-management":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            vendor_id_value = params.get("vendor_id", [""])[0]
            try:
                return self.send_json(vendor_region_management_payload(company_id, vendor_id_value))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/prospect/options":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            uf = params.get("uf", [""])[0]
            try:
                return self.send_json(prospect_options_payload(company_id, uf))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/prospect/clients":
            params = parse_qs(parsed.query)
            company_id = params.get("company", ["ionlab"])[0]
            query = params.get("q", [""])[0]
            uf = params.get("uf", [""])[0]
            city = params.get("city", [""])[0]
            try:
                return self.send_json(prospect_client_search_payload(company_id, query, uf, city))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/prospect":
            try:
                return self.send_json(prospect_payload(parse_qs(parsed.query)))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/auth/session":
            token = self.headers.get("X-Auth-Token", "")
            try:
                return self.send_json(auth_session_payload(token))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/users":
            params = parse_qs(parsed.query)
            query = params.get("q", [""])[0].strip()
            try:
                return self.send_json(users_payload(query))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path.startswith("/api/stats/"):
            company_id = path.rsplit("/", 1)[-1]
            if company_id not in COMPANIES:
                return self.send_json({"error": "Empresa invalida."}, HTTPStatus.BAD_REQUEST)
            records = load_json(sales_file(company_id), [])
            clients = load_json(clients_file(company_id), [])
            logs = load_json(import_log_file(company_id), [])
            client_logs = load_json(client_import_log_file(company_id), [])
            note_ids = {str(item.get("ID_NF") or item.get("NF_NUM") or "") for item in records}
            note_ids.discard("")
            return self.send_json({
                "empresa": COMPANIES[company_id],
                "linhas": len(records),
                "notas": len(note_ids),
                "clientes": len(clients),
                "ultima_importacao": logs[-1] if logs else None,
                "ultima_importacao_clientes": client_logs[-1] if client_logs else None,
            })

        if path.startswith("/api/"):
            return self.send_json({"error": "Rota nao encontrada."}, HTTPStatus.NOT_FOUND)

        if path in ("/", "/index.html") or path.startswith("/vendedor/"):
            return self.serve_static("index.html")

        return self.serve_static(unquote(path.lstrip("/")))

    def do_POST(self):
        path = urlparse(self.path).path
        if path == "/api/auth/login":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
                return self.send_json(auth_login_payload(payload))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.UNAUTHORIZED)

        if path == "/api/auth/logout":
            try:
                token = self.headers.get("X-Auth-Token", "")
                return self.send_json(auth_logout_payload(token))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/auth/change-password":
            try:
                token = self.headers.get("X-Auth-Token", "")
                content_length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
                return self.send_json(change_password_payload(payload, token))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/users":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
                return self.send_json(save_user_payload(payload, self.headers.get("X-Auth-Token", "")))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendors":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
                return self.send_json(save_vendor_payload(payload))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendor-goals":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
                return self.send_json(save_vendor_goals_payload(payload))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendor-monthly-items":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
                return self.send_json(save_vendor_monthly_items_payload(payload))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/slow-items":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
                return self.send_json(save_slow_items_payload(payload))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/regions":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
                action = str(payload.get("action") or "save")
                if action == "delete":
                    return self.send_json(delete_region_rule_payload(str(payload.get("company") or "ionlab"), str(payload.get("id") or "")))
                return self.send_json(save_region_rule_payload(payload))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendor-region-exclusions":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
                return self.send_json(save_vendor_region_exclusions_payload(payload))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/vendor-day-by-day/client":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
                return self.send_json(save_vendor_day_by_day_client_payload(payload))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/block-reasons":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
                return self.send_json(save_block_reason_payload(payload))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/blocked-clients":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
                return self.send_json(save_blocked_client_payload(payload))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path in ("/api/recalculate-costing", "/api/recalculate-costing/", "/api/costing/recalculate"):
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
                return self.send_json(recalculate_costing(
                    payload.get("company", ""),
                    payload.get("indice_custeio"),
                ))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/quotes":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
                return self.send_json(save_quote_payload(payload))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path == "/api/orders":
            try:
                content_length = int(self.headers.get("Content-Length", "0"))
                payload = json.loads(self.rfile.read(content_length).decode("utf-8") or "{}")
                return self.send_json(save_order_payload(payload))
            except Exception as exc:
                return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

        if path not in ("/api/import-sales", "/api/import-clients", "/api/import-stock", "/api/import-in-transit", "/api/import-prices", "/api/import-prices-ipi", "/api/import-products", "/api/import-costing", "/api/import-costing-fabricated", "/api/import-mercado-livre-ads", "/api/import-mercado-livre-sales"):
            return self.send_json({"error": "Rota nao encontrada."}, HTTPStatus.NOT_FOUND)

        try:
            content_length = int(self.headers.get("Content-Length", "0"))
            fields, files = parse_multipart(self.headers, self.rfile.read(content_length))
            company_id = fields.get("company", "")
            upload = files.get("file")

            if not upload or not upload["content"]:
                return self.send_json({"error": "Selecione um arquivo para importar."}, HTTPStatus.BAD_REQUEST)

            if path == "/api/import-sales":
                summary = import_sales(company_id, upload["filename"], upload["content"])
            elif path == "/api/import-clients":
                summary = import_clients(company_id, upload["filename"], upload["content"])
            elif path == "/api/import-prices":
                summary = import_price_table(company_id, upload["filename"], upload["content"])
            elif path == "/api/import-prices-ipi":
                summary = update_price_table_ipi(company_id, upload["filename"], upload["content"])
            elif path == "/api/import-products":
                summary = import_products(company_id, upload["filename"], upload["content"])
            elif path == "/api/import-in-transit":
                summary = import_in_transit(company_id, upload["filename"], upload["content"])
            elif path == "/api/import-costing":
                summary = import_costing(company_id, upload["filename"], upload["content"], fields.get("indice_custeio"))
            elif path == "/api/import-costing-fabricated":
                summary = import_costing_fabricated(company_id, upload["filename"], upload["content"], fields.get("cost_increase"))
            elif path == "/api/import-mercado-livre-ads":
                summary = import_mercado_livre_ads(company_id, upload["filename"], upload["content"], fields.get("kind") or fields.get("type") or "anuncios")
            elif path == "/api/import-mercado-livre-sales":
                summary = import_mercado_livre_sales(company_id, upload["filename"], upload["content"])
            else:
                summary = import_stock(company_id, upload["filename"], upload["content"])
            return self.send_json(summary)
        except Exception as exc:
            return self.send_json({"error": str(exc)}, HTTPStatus.BAD_REQUEST)

    def serve_static(self, requested_path: str):
        file_path = (STATIC_DIR / requested_path).resolve()
        if STATIC_DIR.resolve() not in file_path.parents and file_path != STATIC_DIR.resolve():
            self.send_error(HTTPStatus.FORBIDDEN)
            return

        if file_path.is_dir():
            file_path = file_path / "index.html"

        if not file_path.exists():
            self.send_error(HTTPStatus.NOT_FOUND)
            return

        content_type = CONTENT_TYPES.get(file_path.suffix.lower(), "application/octet-stream")
        content = file_path.read_bytes()
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", content_type)
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def send_json(self, payload, status=HTTPStatus.OK):
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Cache-Control", "no-store")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def log_message(self, format, *args):
        return


def ensure_data_dirs():
    for company_id in COMPANIES:
        company_dir(company_id).mkdir(parents=True, exist_ok=True)


def server_log(message: str):
    try:
        SERVER_LOG_FILE.parent.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().isoformat(timespec="seconds")
        with SERVER_LOG_FILE.open("a", encoding="utf-8") as log:
            log.write(f"[{timestamp}] {message}\n")
    except Exception:
        pass


class CRMHTTPServer(ThreadingHTTPServer):
    def handle_error(self, request, client_address):
        server_log(f"Erro atendendo {client_address}:\n{traceback.format_exc()}")
        super().handle_error(request, client_address)


def warm_start_caches():
    for company_id in COMPANIES:
        try:
            product_search_index(company_id)
            server_log(f"Catalogo de busca preparado para {company_id}.")
        except Exception:
            server_log(f"Falha ao preparar catalogo de busca para {company_id}:\n{traceback.format_exc()}")


def main():
    ensure_data_dirs()
    port = int(sys.argv[1]) if len(sys.argv) > 1 else 8765
    try:
        server = CRMHTTPServer(("127.0.0.1", port), CRMHandler)
        server_log(f"CRM iniciado em http://127.0.0.1:{port}")
        print(f"CRM iniciado em http://127.0.0.1:{port}")
        threading.Thread(target=warm_start_caches, daemon=True).start()
        server.serve_forever()
    except Exception:
        server_log(f"Erro fatal:\n{traceback.format_exc()}")
        raise


if __name__ == "__main__":
    main()
